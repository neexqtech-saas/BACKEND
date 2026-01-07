"""
Leave Control Views
Optimized for high-traffic, low-cost, future-proof architecture
All queries O(1) or using proper database indexes
"""

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from django.db.models import Q
from datetime import datetime
from decimal import Decimal
from .models import LeaveType, EmployeeLeaveBalance, LeaveApplication
from .serializers import (
    LeaveTypeSerializer, LeaveTypeUpdateSerializer,
    EmployeeLeaveBalanceSerializer, EmployeeLeaveBalanceUpdateSerializer,
    LeaveApplicationSerializer, LeaveApplicationUpdateSerializer
)
from AuthN.models import AdminProfile, UserProfile, BaseUserModel
from SiteManagement.models import Site
from utils.pagination_utils import CustomPagination
from utils.site_filter_utils import filter_queryset_by_site


def get_admin_and_site_for_leave(request, site_id, allow_user_role=True):
    """
    Optimized admin and site validation - O(1) queries with select_related
    Handles admin, organization, and user roles
    Returns: (admin, site, None) tuple or (None, None, Response with error)
    """
    user = request.user
    
    # Fast path for admin role - O(1) query
    if user.role == 'admin':
        admin_id = user.id
        admin = user
        # Single O(1) query with index on (id, created_by_admin, is_active)
        try:
            site = Site.objects.only('id', 'site_name', 'created_by_admin_id', 'is_active').get(
                id=site_id, 
                created_by_admin_id=admin_id, 
                is_active=True
            )
            return admin, site, None
        except Site.DoesNotExist:
            return None, None, Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Site not found or you don't have permission to access this site",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
    
    # Organization role - O(1) queries with select_related
    elif user.role == 'organization':
        admin_id = request.query_params.get('admin_id')
        if not admin_id:
            return None, None, Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                "data": None
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate UUID format
        try:
            import uuid
            uuid.UUID(str(admin_id))
        except (ValueError, AttributeError):
            return None, None, Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": f"Invalid admin_id format: {admin_id}. Must be a valid UUID.",
                "data": None
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Single O(1) query with select_related to avoid N+1 - uses index on (id, role)
        try:
            admin = BaseUserModel.objects.select_related('own_admin_profile').only(
                'id', 'role', 'email'
            ).get(id=admin_id, role='admin')
        except BaseUserModel.DoesNotExist:
            return None, None, Response({
                "status": status.HTTP_404_NOT_FOUND,
                "message": f"Admin with ID {admin_id} not found in the system",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        
        # O(1) query - verify admin belongs to organization using select_related
        admin_profile = AdminProfile.objects.select_related('user', 'organization').only(
            'id', 'user_id', 'organization_id'
        ).filter(
            user_id=admin_id,
            organization_id=user.id
        ).first()
        
        if not admin_profile:
            return None, None, Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Admin does not belong to your organization",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Single O(1) query with index on (id, created_by_admin, is_active)
        try:
            site = Site.objects.only('id', 'site_name', 'created_by_admin_id', 'is_active').get(
                id=site_id, 
                created_by_admin_id=admin_id, 
                is_active=True
            )
            return admin, site, None
        except Site.DoesNotExist:
            return None, None, Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Site not found or you don't have permission to access this site",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
    
    # User role - O(1) queries with select_related
    elif user.role == 'user' and allow_user_role:
        # O(1) query - Get admin from current assignment using utility
        from SiteManagement.models import EmployeeAdminSiteAssignment
        
        assignment = EmployeeAdminSiteAssignment.objects.select_related('admin', 'site').filter(
            employee_id=user.id,
            site_id=site_id,
            is_active=True
        ).only('id', 'admin_id', 'site_id', 'employee_id', 'is_active').first()
        
        if not assignment:
            return None, None, Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "You are not assigned to this site",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        admin = assignment.admin
        admin_id = admin.id
        
        # O(1) query - Validate site exists
        try:
            site = Site.objects.only('id', 'site_name', 'is_active').get(
                id=site_id, 
                is_active=True
            )
            return admin, site, None
        except Site.DoesNotExist:
            return None, None, Response({
                "status": status.HTTP_404_NOT_FOUND,
                "message": "Site not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
    
    else:
        return None, None, Response({
            "status": status.HTTP_403_FORBIDDEN,
            "message": "Unauthorized access. Only admin, organization, and user roles can access this endpoint",
            "data": None
        }, status=status.HTTP_403_FORBIDDEN)


class LeaveTypeAPIView(APIView):
    """Leave Type CRUD operations"""
    
    def get(self, request, site_id, pk=None):
        """Get leave type(s) - O(1) queries with index optimization"""
        try:
            admin, site, error_response = get_admin_and_site_for_leave(request, site_id, allow_user_role=True)
            if error_response:
                return error_response
            
            admin_id = admin.id
            
            if pk:
                # Single O(1) query using index leavetype_id_adm_idx (id, admin)
                leave = LeaveType.objects.filter(
                    admin_id=admin_id, 
                    id=pk, 
                    is_active=True
                ).select_related('admin').only(
                    'id', 'admin_id', 'site_id', 'name', 'code', 'description',
                    'default_count', 'is_paid', 'is_active', 'created_at', 'updated_at',
                    'admin__email'
                ).first()
                
                if not leave:
                    return Response({
                        "status": status.HTTP_404_NOT_FOUND,
                        "message": "Leave type not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
                
                # O(1) site check
                if site_id and leave.site_id != site_id:
                    return Response({
                        "status": status.HTTP_404_NOT_FOUND,
                        "message": "Leave type not found for this site"
                    }, status=status.HTTP_404_NOT_FOUND)
                
                serializer = LeaveTypeSerializer(leave)
                return Response({
                    "status": status.HTTP_200_OK,
                    "message": "Leave type fetched successfully",
                    "data": serializer.data
                })
            
            # List query with index optimization - uses leavetype_adm_active_idx (admin, is_active)
            leaves = LeaveType.objects.filter(
                admin_id=admin_id, 
                is_active=True
            ).select_related('admin').only(
                'id', 'admin_id', 'site_id', 'name', 'code', 'description',
                'default_count', 'is_paid', 'is_active', 'created_at', 'updated_at',
                'admin__email'
            )
            
            # Filter by site - O(1) with index leavetype_site_adm_active_idx
            if site_id:
                leaves = leaves.filter(site_id=site_id)
            
            leaves = leaves.order_by('name')
            
            serializer = LeaveTypeSerializer(leaves, many=True)
            return Response({
                "status": status.HTTP_200_OK,
                "message": "Leave types fetched successfully",
                "data": serializer.data
            })
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def post(self, request, site_id, pk=None):
        """Create leave type - Optimized"""
        try:
            admin, site, error_response = get_admin_and_site_for_leave(request, site_id, allow_user_role=False)
            if error_response:
                return error_response
            
            # O(1) query - Get admin profile with select_related
            admin_profile = AdminProfile.objects.select_related('user', 'organization').filter(
                user_id=admin.id
            ).only('id', 'user_id', 'organization_id').first()
            
            if not admin_profile:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Admin profile not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            data = request.data.copy()
            data['admin'] = admin.user_id
            # Set site if provided
            if site_id:
                data['site'] = site_id

            serializer = LeaveTypeSerializer(data=data)
            if not serializer.is_valid():
                # Check for unique constraint error in serializer errors
                errors = serializer.errors
                if errors.get('non_field_errors'):
                    non_field_errors = errors['non_field_errors']
                    if any('unique' in str(err).lower() for err in non_field_errors):
                        code = data.get('code', '')
                        return Response({
                            "status": status.HTTP_400_BAD_REQUEST,
                            "message": f"A leave type with code '{code}' already exists. Please use a different code or set the existing one to inactive.",
                            "data": {
                                "code": [f"A leave type with this code already exists."]
                            }
                        }, status=status.HTTP_400_BAD_REQUEST)
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "Validation failed",
                    "data": serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            serializer.save()
            return Response({
                "status": status.HTTP_201_CREATED,
                "message": "Leave type created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def put(self, request, site_id, pk=None):
        # Get admin_id based on role
        if request.user.role == 'admin':
            admin_id = request.user.id
            admin_user = request.user
        elif request.user.role == 'organization':
            # Organization role: get admin_id from query params
            admin_id = request.query_params.get('admin_id')
            if not admin_id:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate admin exists and belongs to organization
            try:
                admin_user = BaseUserModel.objects.get(id=admin_id, role='admin')
                # Verify admin belongs to organization
                from AuthN.models import AdminProfile
                admin_profile = AdminProfile.objects.filter(
                    user=admin_user,
                    organization=request.user
                ).first()
                if not admin_profile:
                    return Response({
                        "status": status.HTTP_403_FORBIDDEN,
                        "message": "Admin does not belong to your organization",
                        "data": None
                    }, status=status.HTTP_403_FORBIDDEN)
            except BaseUserModel.DoesNotExist:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Admin not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        leave = get_object_or_404(LeaveType, admin__id=admin_id, id=pk)
        
        # Validate site belongs to admin
        try:
            site = Site.objects.get(id=site_id, created_by_admin=admin_user, is_active=True)
            if leave.site_id != site_id:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Leave type not found for this site"
                }, status=status.HTTP_404_NOT_FOUND)
        except Site.DoesNotExist:
            return Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Site not found or you don't have permission to access this site",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        serializer = LeaveTypeUpdateSerializer(leave, data=request.data)
        if not serializer.is_valid():
            # Check for unique constraint error in serializer errors
            errors = serializer.errors
            if errors.get('non_field_errors'):
                non_field_errors = errors['non_field_errors']
                if any('unique' in str(err).lower() for err in non_field_errors):
                    code = request.data.get('code', leave.code)
                    return Response({
                        "status": status.HTTP_400_BAD_REQUEST,
                        "message": f"A leave type with code '{code}' already exists. Please use a different code or set the existing one to inactive.",
                        "data": {
                            "code": [f"A leave type with this code already exists."]
                        }
                    }, status=status.HTTP_400_BAD_REQUEST)
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "Validation failed",
                "data": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        serializer.save()
        return Response({
            "status": status.HTTP_200_OK,
            "message": "Leave type updated successfully",
            "data": serializer.data
        })

    def patch(self, request, site_id, pk=None):
        # Get admin_id based on role
        if request.user.role == 'admin':
            admin_id = request.user.id
            admin_user = request.user
        elif request.user.role == 'organization':
            # Organization role: get admin_id from query params
            admin_id = request.query_params.get('admin_id')
            if not admin_id:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate admin exists and belongs to organization
            try:
                admin_user = BaseUserModel.objects.get(id=admin_id, role='admin')
                # Verify admin belongs to organization
                from AuthN.models import AdminProfile
                admin_profile = AdminProfile.objects.filter(
                    user=admin_user,
                    organization=request.user
                ).first()
                if not admin_profile:
                    return Response({
                        "status": status.HTTP_403_FORBIDDEN,
                        "message": "Admin does not belong to your organization",
                        "data": None
                    }, status=status.HTTP_403_FORBIDDEN)
            except BaseUserModel.DoesNotExist:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Admin not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        leave = get_object_or_404(LeaveType, admin__id=admin_id, id=pk)
        
        # Validate site belongs to admin
        try:
            site = Site.objects.get(id=site_id, created_by_admin=admin_user, is_active=True)
            if leave.site_id != site_id:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Leave type not found for this site"
                }, status=status.HTTP_404_NOT_FOUND)
        except Site.DoesNotExist:
            return Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Site not found or you don't have permission to access this site",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        serializer = LeaveTypeSerializer(leave, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "status": status.HTTP_200_OK,
                "message": "Leave type updated successfully",
                "data": serializer.data
            })
        return Response({
            "status": status.HTTP_400_BAD_REQUEST,
            "message": "Validation failed",
            "data": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, site_id, pk=None):
        # Get admin_id based on role
        if request.user.role == 'admin':
            admin_id = request.user.id
            admin_user = request.user
        elif request.user.role == 'organization':
            # Organization role: get admin_id from query params
            admin_id = request.query_params.get('admin_id')
            if not admin_id:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate admin exists and belongs to organization
            try:
                admin_user = BaseUserModel.objects.get(id=admin_id, role='admin')
                # Verify admin belongs to organization
                from AuthN.models import AdminProfile
                admin_profile = AdminProfile.objects.filter(
                    user=admin_user,
                    organization=request.user
                ).first()
                if not admin_profile:
                    return Response({
                        "status": status.HTTP_403_FORBIDDEN,
                        "message": "Admin does not belong to your organization",
                        "data": None
                    }, status=status.HTTP_403_FORBIDDEN)
            except BaseUserModel.DoesNotExist:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Admin not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        leave = get_object_or_404(LeaveType, admin__id=admin_id, id=pk)
        
        # Validate site belongs to admin
        try:
            site = Site.objects.get(id=site_id, created_by_admin=admin_user, is_active=True)
            if leave.site_id != site_id:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Leave type not found for this site"
                }, status=status.HTTP_404_NOT_FOUND)
        except Site.DoesNotExist:
            return Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Site not found or you don't have permission to access this site",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Check if leave type is assigned to any employee
        # Get unique employee count (not total records)
        assigned_employees = EmployeeLeaveBalance.objects.filter(
            leave_type=leave
        ).values_list('user', flat=True).distinct()
        
        assigned_count = assigned_employees.count()
        
        if assigned_count > 0:
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": f"Cannot delete '{leave.name}'. This leave type is currently assigned to {assigned_count} employee(s). Please unassign this leave type from all employees before deleting.",
                "data": {
                    "assigned_employees_count": assigned_count,
                    "total_assignments": EmployeeLeaveBalance.objects.filter(
                        leave_type=leave
                    ).count(),
                    "leave_type_name": leave.name
                }
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if leave type has any applications
        applications_count = LeaveApplication.objects.filter(
            leave_type=leave
        ).count()
        
        if applications_count > 0:
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": f"Cannot delete '{leave.name}'. This leave type has {applications_count} leave application(s) associated with it. Deleting it would result in loss of historical data.",
                "data": {
                    "applications_count": applications_count,
                    "leave_type_name": leave.name
                }
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Safe to delete
        leave.is_active = False
        leave.save()
        return Response({
            "status": status.HTTP_200_OK,
            "message": f"{leave.name} deleted successfully",
            "data": None
        })


class EmployeeLeaveBalanceAPIView(APIView):
    """
    Employee Leave Balance - View & Update Only
    Supports both admin_id (all employees) and user_id (specific user)
    Note: For assigning leaves, use AssignLeaveAPIView
    """
    
    def get(self, request, site_id, user_id=None, pk=None):
        """
        GET /leave-balances/<site_id>/?year=2025 -> All employees under admin (year optional)
        GET /leave-balances/<site_id>/<user_id>/?year=2025 -> Specific employee (year optional)
        GET /leave-balances/<site_id>/<user_id>/<pk>/ -> Specific balance
        """
        # Get admin_id based on role
        if request.user.role == 'admin':
            admin_id = request.user.id
            admin = request.user
        elif request.user.role == 'organization':
            # Organization role: get admin_id from query params
            admin_id = request.query_params.get('admin_id')
            if not admin_id:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate admin exists and belongs to organization
            try:
                admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                # Verify admin belongs to organization
                admin_profile = AdminProfile.objects.filter(
                    user=admin,
                    organization=request.user
                ).first()
                if not admin_profile:
                    return Response({
                        "status": status.HTTP_403_FORBIDDEN,
                        "message": "Admin does not belong to your organization",
                        "data": None
                    }, status=status.HTTP_403_FORBIDDEN)
            except BaseUserModel.DoesNotExist:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Admin not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
        elif request.user.role == 'user':
            # Employee role: can only view their own leave balances
            if user_id and str(user_id) != str(request.user.id):
                return Response({
                    "status": status.HTTP_403_FORBIDDEN,
                    "message": "You can only view your own leave balances",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Set user_id to logged-in user
            user_id = request.user.id
            
            # Validate that employee is assigned to this site
            from SiteManagement.models import EmployeeAdminSiteAssignment
            from datetime import date
            from django.db.models import Q
            
            assignment = EmployeeAdminSiteAssignment.objects.filter(
                employee_id=user_id,
                site_id=site_id,
                is_active=True
            ).select_related('admin', 'site').first()
            
            if not assignment:
                return Response({
                    "status": status.HTTP_403_FORBIDDEN,
                    "message": "You are not assigned to this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            admin = assignment.admin
            admin_id = admin.id
            
            # Validate site exists
            try:
                site = Site.objects.get(id=site_id, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Site not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Unauthorized access. Only admin, organization, and user roles can access this endpoint",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Validate site belongs to admin (for admin and organization roles only)
        if request.user.role != 'user':
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": status.HTTP_403_FORBIDDEN,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
        
        # Get year from query params (optional)
        year = request.GET.get('year')
        
        # Specific balance by ID
        if pk:
            # Single O(1) query using index leavebal_id_user_idx
            if user_id:
                balance = EmployeeLeaveBalance.objects.filter(
                    user_id=user_id, 
                    id=pk
                ).select_related('user', 'user__own_user_profile', 'leave_type').only(
                    'id', 'user_id', 'leave_type_id', 'year', 'assigned', 'used',
                    'leave_type__site_id', 'leave_type__name', 'leave_type__code',
                    'user__email', 'user__own_user_profile__user_name'
                ).first()
            else:
                balance = EmployeeLeaveBalance.objects.filter(
                    id=pk
                ).select_related('user', 'user__own_user_profile', 'leave_type').only(
                    'id', 'user_id', 'leave_type_id', 'year', 'assigned', 'used',
                    'leave_type__site_id', 'leave_type__name', 'leave_type__code',
                    'user__email', 'user__own_user_profile__user_name'
                ).first()
            
            if not balance:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Leave balance not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) site check - access via select_related
            if site_id and balance.leave_type.site_id != site_id:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Leave balance not found for this site"
                }, status=status.HTTP_404_NOT_FOUND)
            
            serializer = EmployeeLeaveBalanceSerializer(balance)
            return Response({
                "status": status.HTTP_200_OK,
                "message": "Leave balance fetched successfully",
                "data": serializer.data
            })
        
        # Admin viewing all employees' balances
        if admin_id and not user_id:
            # Get all users under this admin using utility
            from utils.Employee.assignment_utils import get_employee_ids_under_admin
            admin_users = get_employee_ids_under_admin(admin_id, active_only=True, site_id=site_id)
            
            # Build query with optional year filter
            query = {
                'user__id__in': admin_users
            }
            if year:
                query['year'] = int(year)
            
            balances = EmployeeLeaveBalance.objects.filter(
                **query
            ).select_related('user', 'user__own_user_profile', 'leave_type').only(
                'id', 'user_id', 'leave_type_id', 'year', 'assigned', 'used',
                'leave_type__site_id', 'leave_type__name', 'leave_type__code',
                'user__email', 'user__own_user_profile__user_name'
            )
            
            # Filter by site if provided (filter by leave_type's site) - O(1) with index
            if site_id:
                balances = balances.filter(leave_type__site_id=site_id)
            
            balances = balances.order_by('-year', 'user__email', 'leave_type__name')
            
            serializer = EmployeeLeaveBalanceSerializer(balances, many=True)
            
            message = f"All employees leave balances fetched successfully"
            if year:
                message += f" for year {year}"
            
            return Response({
                "status": status.HTTP_200_OK,
                "message": message,
                "data": serializer.data
            })
        
        # Specific user's balances
        if user_id:
            # Build query with optional year filter
            query = {
                'user__id': user_id
            }
            if year:
                query['year'] = int(year)
            
            balances = EmployeeLeaveBalance.objects.filter(**query).select_related(
                'user', 'user__own_user_profile', 'leave_type'
            ).only(
                'id', 'user_id', 'leave_type_id', 'year', 'assigned', 'used',
                'leave_type__name', 'leave_type__code',
                'user__email', 'user__own_user_profile__user_name'
            ).order_by('-year', 'leave_type__name')
            serializer = EmployeeLeaveBalanceSerializer(balances, many=True)
            
            message = f"Leave balances fetched successfully"
            if year:
                message += f" for year {year}"
            
            return Response({
                "status": status.HTTP_200_OK,
                "message": message,
                "data": serializer.data
            })
        
        return Response({
            "status": status.HTTP_400_BAD_REQUEST,
            "message": "Invalid request",
            "data": None
        }, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request, admin_id=None, user_id=None, pk=None):
        """
        Update leave balance - Only 'assigned' field can be updated
        Cannot change user, leave_type, year, used, etc.
        """
        leave_balance = get_object_or_404(EmployeeLeaveBalance, id=pk)
        
        # Only allow updating 'assigned' field
        if 'assigned' not in request.data:
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "Only 'assigned' field can be updated",
                "data": None
            }, status=status.HTTP_400_BAD_REQUEST)
        
        assigned = float(request.data['assigned'])
        
        # Validation: assigned cannot exceed default_count
        leave_type = leave_balance.leave_type
        if assigned > float(leave_type.default_count):
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": f"Cannot assign {assigned} days for {leave_type.name}. Maximum allowed is {leave_type.default_count} days",
                "data": None
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validation: assigned should be >= used (cannot reduce below used)
        from decimal import Decimal
        used_value = Decimal(str(leave_balance.used))
        assigned_decimal = Decimal(str(assigned))
        
        if assigned_decimal < used_value:
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": f"Cannot assign {assigned} days. Employee has already used {leave_balance.used} days",
                "data": None
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Update only assigned field
        leave_balance.assigned = assigned
        leave_balance.save()
        
        serializer = EmployeeLeaveBalanceSerializer(leave_balance)
        return Response({
            "status": status.HTTP_200_OK,
            "message": "Leave balance updated successfully",
            "data": serializer.data
        })

    def delete(self, request, admin_id=None, user_id=None, pk=None):
        """Delete leave balance (unassign leave from employee)"""
        leave_balance = get_object_or_404(EmployeeLeaveBalance, id=pk)
        
        # Check if employee has used any leaves
        if leave_balance.used > 0:
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": f"Cannot unassign {leave_balance.leave_type.name}. Employee has already used {leave_balance.used} days.",
                "data": None
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Hard delete the record
        leave_type_name = leave_balance.leave_type.name
        leave_balance.delete()
        
        return Response({
            "status": status.HTTP_200_OK,
            "message": f"{leave_type_name} unassigned successfully",
            "data": None
        })


class AssignLeaveAPIView(APIView):
    """
    Flexible Leave Assignment API - Single aur Bulk dono handle karta hai
    Auto-assigns default_count if 'assigned' not provided
    
    Single Format (assigned optional):
    {
        "year": 2024,
        "leave_type": 1
    }
    OR
    {
        "year": 2024,
        "leave_type": 1,
        "assigned": 12
    }
    
    Bulk Format (assigned optional):
    {
        "year": 2024,
        "leaves": [
            {"leave_type": 1},
            {"leave_type": 2, "assigned": 10}
        ]
    }
    """
    
    def post(self, request, site_id, user_id=None):
        """Assign leave(s) to a user - single ya bulk"""
        # Get admin_id based on role
        if request.user.role == 'admin':
            admin_id = request.user.id
            admin = request.user
        elif request.user.role == 'organization':
            # Organization role: get admin_id from query params
            admin_id = request.query_params.get('admin_id')
            if not admin_id:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate admin exists and belongs to organization
            try:
                admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                # Verify admin belongs to organization
                admin_profile = AdminProfile.objects.filter(
                    user=admin,
                    organization=request.user
                ).first()
                if not admin_profile:
                    return Response({
                        "status": status.HTTP_403_FORBIDDEN,
                        "message": "Admin does not belong to your organization",
                        "data": None
                    }, status=status.HTTP_403_FORBIDDEN)
            except BaseUserModel.DoesNotExist:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Admin not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
        elif request.user.role == 'user':
            # Employee role: can only assign leaves to themselves
            if not user_id:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "user_id is required in URL",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if str(user_id) != str(request.user.id):
                return Response({
                    "status": status.HTTP_403_FORBIDDEN,
                    "message": "You can only assign leaves to yourself",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            user_id = request.user.id
            
            # Validate that employee is assigned to this site
            from SiteManagement.models import EmployeeAdminSiteAssignment
            from datetime import date
            from django.db.models import Q
            
            assignment = EmployeeAdminSiteAssignment.objects.filter(
                employee_id=user_id,
                site_id=site_id,
                is_active=True
            ).select_related('admin', 'site').first()
            
            if not assignment:
                return Response({
                    "status": status.HTTP_403_FORBIDDEN,
                    "message": "You are not assigned to this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            admin = assignment.admin
            admin_id = admin.id
            
            # Validate site exists
            try:
                site = Site.objects.get(id=site_id, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Site not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Unauthorized access. Only admin, organization, and user roles can access this endpoint",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Validate site belongs to admin (for admin and organization roles only)
        if request.user.role != 'user':
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": status.HTTP_403_FORBIDDEN,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
        
        # Get user_id from URL or request body
        target_user_id = user_id or request.data.get('user_id')
        if not target_user_id:
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "user_id is required",
                "data": None
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get user directly from BaseUserModel
        from AuthN.models import BaseUserModel
        try:
            user = BaseUserModel.objects.get(id=target_user_id, role='user')
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": f"User with ID {target_user_id} not found or not a valid employee",
                "data": None
            }, status=status.HTTP_400_BAD_REQUEST)
        
        year = request.data.get('year')
        
        if not year:
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "Year is required",
                "data": None
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if single or bulk format
        leaves_list = []
        
        # Single leave format check
        if 'leave_type' in request.data:
            leaves_list.append({
                'leave_type': request.data.get('leave_type'),
                'assigned': request.data.get('assigned')  # Can be None
            })
        # Bulk leaves format check
        elif 'leaves' in request.data:
            leaves_list = request.data.get('leaves', [])
        else:
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "Either 'leave_type' or 'leaves' array is required",
                "data": None
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not leaves_list:
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "At least one leave type is required",
                "data": None
            }, status=status.HTTP_400_BAD_REQUEST)
        
        created_balances = []
        errors = []
        
        for leave_data in leaves_list:
            leave_type_id = leave_data.get('leave_type')
            assigned = leave_data.get('assigned')  # Can be None
            
            if not leave_type_id:
                errors.append({
                    "error": "leave_type is required for each entry"
                })
                continue
            
            # Get leave type to fetch default_count if assigned not provided
            try:
                leave_type = LeaveType.objects.get(id=leave_type_id, is_active=True)
                
                # If assigned not provided, use default_count from leave type
                if assigned is None:
                    assigned = leave_type.default_count
                else:
                    assigned = float(assigned)
                    
                    # Validation: assigned should not exceed default_count
                    if assigned > float(leave_type.default_count):
                        errors.append({
                            "leave_type": leave_type_id,
                            "leave_type_name": leave_type.name,
                            "leave_type_code": leave_type.code,
                            "error": f"Cannot assign {assigned} days. Maximum allowed is {leave_type.default_count} days (default count)"
                        })
                        continue
                    
            except LeaveType.DoesNotExist:
                errors.append({
                    "leave_type": leave_type_id,
                    "error": "Leave type not found or inactive"
                })
                continue
            
            # Check if balance already exists
            existing_balance = EmployeeLeaveBalance.objects.filter(
                user__id=target_user_id,
                leave_type_id=leave_type_id,
                year=year
            ).first()
            
            if existing_balance:
                errors.append({
                    "leave_type": leave_type_id,
                    "leave_type_name": leave_type.name,
                    "error": f"Balance already exists for {leave_type.name} ({leave_type.code}) in {year}"
                })
                continue
            
            # Create new balance
            try:
                balance = EmployeeLeaveBalance.objects.create(
                    user=user,
                    leave_type_id=leave_type_id,
                    year=year,
                    assigned=assigned,
                    used=0
                )
                created_balances.append({
                    "id": balance.id,
                    "leave_type": leave_type_id,
                    "leave_type_name": leave_type.name,
                    "leave_type_code": leave_type.code,
                    "assigned": float(assigned),
                    "used": 0,
                    "balance": float(assigned)
                })
            except Exception as e:
                errors.append({
                    "leave_type": leave_type_id,
                    "leave_type_name": leave_type.name,
                    "error": str(e)
                })
        
        if created_balances:
            return Response({
                "status": status.HTTP_201_CREATED,
                "message": f"Successfully assigned {len(created_balances)} leave type(s)",
                "data": {
                    "created": created_balances,
                    "errors": errors
                }
            }, status=status.HTTP_201_CREATED)
        else:
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "Failed to assign leaves",
                "data": {
                    "created": [],
                    "errors": errors
                }
            }, status=status.HTTP_400_BAD_REQUEST)


class LeaveApplicationAPIView(APIView):
    """
    Leave Application CRUD operations
    Supports both admin_id (all employees) and user_id (specific user)
    """
    pagination_class = CustomPagination
    
    def generate_excel_export(self, leave_data, year):
        """Generate Excel export for leave applications"""
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from io import BytesIO
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leave Applications"
        
        headers = [
            "Employee Name", "Custom Employee ID", "Email", "Leave Type", "Leave Type Code",
            "From Date", "To Date", "Total Days", "Day Type", "Reason", "Status",
            "Applied At", "Reviewed At", "Reviewed By", "Comments"
        ]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Header Row
        for col, head in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=head)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")
        
        # Helper function to convert values to Excel-compatible format
        def to_excel_value(val):
            if val is None:
                return "N/A"
            if hasattr(val, 'strftime'):
                return val.strftime('%Y-%m-%d')
            try:
                if not isinstance(val, (str, int, float, bool)):
                    return str(val)
            except:
                return "N/A"
            return val
        
        # Data Rows
        for i, leave in enumerate(leave_data, 2):
            row = [
                to_excel_value(leave.get("user_name", "N/A")),
                to_excel_value(leave.get("custom_employee_id", "N/A")),
                to_excel_value(leave.get("user_email", "N/A")),
                to_excel_value(leave.get("leave_type_name", "N/A")),
                to_excel_value(leave.get("leave_type_code", "N/A")),
                to_excel_value(leave.get("from_date", "N/A")),
                to_excel_value(leave.get("to_date", "N/A")),
                to_excel_value(leave.get("total_days", 0)),
                to_excel_value(leave.get("leave_day_type", "N/A")),
                to_excel_value(leave.get("reason", "N/A")),
                to_excel_value(leave.get("status", "N/A")),
                to_excel_value(leave.get("applied_at", "N/A")),
                to_excel_value(leave.get("reviewed_at", "N/A")),
                to_excel_value(leave.get("reviewed_by_email", "N/A")),
                to_excel_value(leave.get("comments", "N/A")),
            ]
            for col, val in enumerate(row, 1):
                ws.cell(row=i, column=col).value = to_excel_value(val)
        
        # Auto width
        for col in ws.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[letter].width = max_len + 2
        
        # Save in memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="leave_applications_{year}.xlsx"'
        
        return response

    def get_year_date_range(self, organization_id, year):
        """
        Calculate date range based on organization's leave year type
        Returns: (start_date, end_date) tuple
        """
        from AuthN.models import OrganizationSettings, BaseUserModel
        from datetime import date
        
        try:
            org = BaseUserModel.objects.get(id=organization_id, role='organization')
            org_settings = OrganizationSettings.objects.filter(organization=org).first()
            
            if not org_settings:
                # Default to calendar year if no settings
                return date(year, 1, 1), date(year, 12, 31)
            
            leave_year_type = org_settings.leave_year_type
            start_month = org_settings.leave_year_start_month or 1
            
            if leave_year_type == 'calendar':
                # Calendar Year: Jan 1 to Dec 31
                start_date = date(year, 1, 1)
                end_date = date(year, 12, 31)
            elif leave_year_type == 'financial':
                # Financial Year (India): Apr 1 to Mar 31 next year
                # If year is 2025, it means Apr 1, 2025 to Mar 31, 2026
                start_date = date(year, 4, 1)
                end_date = date(year + 1, 3, 31)
            elif leave_year_type == 'custom':
                # Custom year based on start_month
                start_date = date(year, start_month, 1)
                # End date is last day of month before start_month next year
                if start_month == 1:
                    end_date = date(year, 12, 31)
                else:
                    import calendar
                    end_month = start_month - 1
                    end_year = year + 1
                    last_day = calendar.monthrange(end_year, end_month)[1]
                    end_date = date(end_year, end_month, last_day)
            else:
                # Default to calendar
                start_date = date(year, 1, 1)
                end_date = date(year, 12, 31)
            
            return start_date, end_date
        except Exception as e:
            # Fallback to calendar year
            return date(year, 1, 1), date(year, 12, 31)

    def get(self, request, site_id, user_id=None, pk=None):
        """
        GET /leave-applications/<site_id>/?year=2025 -> All employees' applications (year REQUIRED)
        GET /leave-applications/<site_id>/<user_id>/?year=2025 -> Specific employee's applications (year REQUIRED)
        GET /leave-applications/<site_id>/<user_id>/<pk>/ -> Specific application (year not needed)
        
        Year filtering respects organization's leave year type (calendar/financial/custom)
        """
        # Get admin_id based on role
        if request.user.role == 'admin':
            admin_id = request.user.id
            admin = request.user
        elif request.user.role == 'organization':
            # Organization role: get admin_id from query params
            admin_id = request.query_params.get('admin_id')
            if not admin_id:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate admin exists and belongs to organization
            try:
                admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                # Verify admin belongs to organization
                admin_profile = AdminProfile.objects.filter(
                    user=admin,
                    organization=request.user
                ).first()
                if not admin_profile:
                    return Response({
                        "status": status.HTTP_403_FORBIDDEN,
                        "message": "Admin does not belong to your organization",
                        "data": None
                    }, status=status.HTTP_403_FORBIDDEN)
            except BaseUserModel.DoesNotExist:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Admin not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
        elif request.user.role == 'user':
            # Employee role: can only view their own leave applications
            if user_id and str(user_id) != str(request.user.id):
                return Response({
                    "status": status.HTTP_403_FORBIDDEN,
                    "message": "You can only view your own leave applications",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Set user_id to logged-in user
            user_id = request.user.id
            
            # Validate that employee is assigned to this site
            from SiteManagement.models import EmployeeAdminSiteAssignment
            from datetime import date
            from django.db.models import Q
            
            assignment = EmployeeAdminSiteAssignment.objects.filter(
                employee_id=user_id,
                site_id=site_id,
                is_active=True
            ).select_related('admin', 'site').first()
            
            if not assignment:
                return Response({
                    "status": status.HTTP_403_FORBIDDEN,
                    "message": "You are not assigned to this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            admin = assignment.admin
            admin_id = admin.id
            
            # Validate site exists
            try:
                site = Site.objects.get(id=site_id, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Site not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Unauthorized access. Only admin, organization, and user roles can access this endpoint",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Validate site belongs to admin (for admin and organization roles only)
        if request.user.role != 'user':
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": status.HTTP_403_FORBIDDEN,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
        
        # Specific leave application by ID (no year needed) - O(1) query
        if pk:
            if user_id:
                leave = LeaveApplication.objects.filter(
                    user_id=user_id, 
                    id=pk
                ).select_related(
                    'user', 'user__own_user_profile', 'leave_type', 'reviewed_by', 'admin', 'organization'
                ).only(
                    'id', 'admin_id', 'site_id', 'organization_id', 'user_id', 'leave_type_id',
                    'from_date', 'to_date', 'total_days', 'leave_day_type', 'reason',
                    'status', 'applied_at', 'reviewed_at', 'reviewed_by_id', 'comments',
                    'user__email', 'user__own_user_profile__user_name', 'user__own_user_profile__custom_employee_id',
                    'leave_type__name', 'leave_type__code', 'reviewed_by__email'
                ).first()
            else:
                leave = LeaveApplication.objects.filter(
                    id=pk
                ).select_related(
                    'user', 'user__own_user_profile', 'leave_type', 'reviewed_by', 'admin', 'organization'
                ).only(
                    'id', 'admin_id', 'site_id', 'organization_id', 'user_id', 'leave_type_id',
                    'from_date', 'to_date', 'total_days', 'leave_day_type', 'reason',
                    'status', 'applied_at', 'reviewed_at', 'reviewed_by_id', 'comments',
                    'user__email', 'user__own_user_profile__user_name', 'user__own_user_profile__custom_employee_id',
                    'leave_type__name', 'leave_type__code', 'reviewed_by__email'
                ).first()
            
            if not leave:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Leave application not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            serializer = LeaveApplicationSerializer(leave)
            return Response({
                "status": status.HTTP_200_OK,
                "message": "Leave application fetched successfully",
                "data": serializer.data
            })
        
        # Year parameter is REQUIRED for listing
        year_param = request.GET.get('year')
        if not year_param:
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "Year parameter is required. Please provide ?year=2025",
                "data": None
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate and parse year
        try:
            year = int(year_param)
            if year < 1900 or year > 2100:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "Invalid year. Year must be between 1900 and 2100",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
        except ValueError:
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "Invalid year format. Year must be a valid integer (e.g., 2025)",
                "data": None
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get organization_id from admin or user
        from AuthN.models import UserProfile
        # Validate admin and site if provided
        if admin_id:
            admin_profile = UserProfile.objects.filter(user_id=admin_id).first()
            if admin_profile:
                org_id = admin_profile.organization_id
            else:
                org_id = None
        elif user_id:
            user_profile = UserProfile.objects.filter(user_id=user_id).first()
            if user_profile:
                org_id = user_profile.organization_id
            else:
                org_id = None
        else:
            org_id = None
        
        # Get date filter from query params or use year range
        from datetime import date, datetime, timedelta
        from_date_param = request.GET.get('from_date')
        to_date_param = request.GET.get('to_date')
        from_date = None
        to_date = None
        
        if from_date_param or to_date_param:
            # Use provided date range
            if from_date_param:
                try:
                    from_date = datetime.strptime(from_date_param, "%Y-%m-%d").date()
                except ValueError:
                    return Response({
                        "status": status.HTTP_400_BAD_REQUEST,
                        "message": "Invalid from_date format. Use YYYY-MM-DD",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
            else:
                # Default to 10 days ago if from_date not provided
                from_date = date.today() - timedelta(days=10)
            
            if to_date_param:
                try:
                    to_date = datetime.strptime(to_date_param, "%Y-%m-%d").date()
                except ValueError:
                    return Response({
                        "status": status.HTTP_400_BAD_REQUEST,
                        "message": "Invalid to_date format. Use YYYY-MM-DD",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
            else:
                # Default to today if to_date not provided
                to_date = date.today()
            
            date_filter = {
                'from_date__gte': from_date,
                'from_date__lte': to_date
            }
        else:
            # Calculate date range based on organization's leave year type
            if org_id:
                start_date, end_date = self.get_year_date_range(org_id, year)
                date_filter = {
                    'from_date__gte': start_date,
                    'from_date__lte': end_date
                }
            else:
                # Fallback to calendar year if no org settings
                start_date = date(year, 1, 1)
                end_date = date(year, 12, 31)
                date_filter = {
                    'from_date__gte': start_date,
                    'from_date__lte': end_date
                }
        
        # Admin viewing all employees' applications
        if admin_id and not user_id:
            # Get employees under this admin using utility
            from utils.Employee.assignment_utils import get_employee_ids_under_admin
            admin_users = get_employee_ids_under_admin(admin_id, active_only=True, site_id=site_id)
            leaves = LeaveApplication.objects.filter(
                user_id__in=admin_users,
                **date_filter
            ).select_related('user', 'user__own_user_profile', 'leave_type', 'reviewed_by', 'admin', 'organization').only(
                'id', 'admin_id', 'site_id', 'organization_id', 'user_id', 'leave_type_id',
                'from_date', 'to_date', 'total_days', 'leave_day_type', 'reason',
                'status', 'applied_at', 'reviewed_at', 'reviewed_by_id', 'comments',
                'user__email', 'user__own_user_profile__user_name', 'user__own_user_profile__custom_employee_id',
                'leave_type__name', 'leave_type__code', 'reviewed_by__email'
            )
            
            # Filter by site if provided - O(1) with index leaveapp_site_adm_st_app_idx
            if site_id:
                leaves = leaves.filter(site_id=site_id)
            
            leaves = leaves.order_by('-applied_at')
            
            # Apply search filter if provided
            search_query = request.GET.get('search', '').strip()
            if search_query:
                leaves = leaves.filter(
                    Q(user__own_user_profile__user_name__icontains=search_query) |
                    Q(user__email__icontains=search_query) |
                    Q(user__own_user_profile__custom_employee_id__icontains=search_query) |
                    Q(leave_type__name__icontains=search_query) |
                    Q(reason__icontains=search_query)
                )
            
            # Apply status filter if provided
            status_filter = request.GET.get('status', '').strip()
            if status_filter and status_filter.lower() != 'all':
                leaves = leaves.filter(status=status_filter.lower())
            
            # Check for Excel export
            export = request.GET.get('export') == 'true'
            if export:
                # Get all leaves without pagination for export (respects search and status filters)
                serializer = LeaveApplicationSerializer(leaves, many=True)
                return self.generate_excel_export(serializer.data, year)
            
            # Apply pagination
            paginator = self.pagination_class()
            paginated_leaves = paginator.paginate_queryset(leaves, request)
            serializer = LeaveApplicationSerializer(paginated_leaves, many=True)
            pagination_data = paginator.get_paginated_response(serializer.data)
            
            # Determine which date range to return in response
            if from_date_param or to_date_param:
                response_date_range = {
                    "start_date": from_date.isoformat(),
                    "end_date": to_date.isoformat()
                }
            else:
                response_date_range = {
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat()
                }
            
            return Response({
                "status": status.HTTP_200_OK,
                "message": f"All employees leave applications for year {year}",
                "year": year,
                "date_range": response_date_range,
                "data": serializer.data,
                **pagination_data
            })
        
        # Specific user's applications
        if user_id:
            leaves = LeaveApplication.objects.filter(
                user_id=user_id,
                **date_filter
            ).select_related('user', 'user__own_user_profile', 'leave_type', 'reviewed_by', 'admin', 'organization').only(
                'id', 'admin_id', 'site_id', 'organization_id', 'user_id', 'leave_type_id',
                'from_date', 'to_date', 'total_days', 'leave_day_type', 'reason',
                'status', 'applied_at', 'reviewed_at', 'reviewed_by_id', 'comments',
                'user__email', 'user__own_user_profile__user_name', 'user__own_user_profile__custom_employee_id',
                'leave_type__name', 'leave_type__code', 'reviewed_by__email'
            )
            
            # Filter by site if provided - O(1) with index
            if site_id:
                leaves = leaves.filter(site_id=site_id)
            
            leaves = leaves.order_by('-applied_at')
            
            # Apply search filter if provided
            search_query = request.GET.get('search', '').strip()
            if search_query:
                leaves = leaves.filter(
                    Q(leave_type__name__icontains=search_query) |
                    Q(reason__icontains=search_query)
                )
            
            # Apply range/limit parameter (default 50)
            limit_param = request.GET.get('limit') or request.GET.get('range')
            if limit_param:
                try:
                    limit = int(limit_param)
                    if limit > 0:
                        leaves = leaves[:limit]
                except ValueError:
                    pass  # Invalid limit, use default
            
            # If no limit specified, default to 50
            if not limit_param:
                leaves = leaves[:50]
            
            serializer = LeaveApplicationSerializer(leaves, many=True)
            
            # Determine which date range to return in response
            if from_date_param or to_date_param:
                response_date_range = {
                    "start_date": from_date.isoformat(),
                    "end_date": to_date.isoformat()
                }
            else:
                response_date_range = {
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat()
                }
            
            return Response({
                "status": status.HTTP_200_OK,
                "message": f"Leave applications fetched successfully for year {year}",
                "year": year,
                "date_range": response_date_range,
                "data": serializer.data
            })
        
        return Response({
            "status": status.HTTP_400_BAD_REQUEST,
            "message": "Invalid request",
            "data": None
        }, status=status.HTTP_400_BAD_REQUEST)

    def post(self, request, site_id, user_id=None, pk=None):
        """Create leave application"""
        # Get admin_id based on role
        if request.user.role == 'admin':
            admin_id = request.user.id
            admin = request.user
        elif request.user.role == 'organization':
            # Organization role: get admin_id from query params
            admin_id = request.query_params.get('admin_id')
            if not admin_id:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate admin exists and belongs to organization
            try:
                admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                # Verify admin belongs to organization
                admin_profile = AdminProfile.objects.filter(
                    user=admin,
                    organization=request.user
                ).first()
                if not admin_profile:
                    return Response({
                        "status": status.HTTP_403_FORBIDDEN,
                        "message": "Admin does not belong to your organization",
                        "data": None
                    }, status=status.HTTP_403_FORBIDDEN)
            except BaseUserModel.DoesNotExist:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Admin not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Validate site belongs to admin
        try:
            site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
        except Site.DoesNotExist:
            return Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Site not found or you don't have permission to access this site",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Get user_id from URL or request body
        target_user_id = user_id or request.data.get('user_id')
        if not target_user_id:
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "user_id is required",
                "data": None
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get user from BaseUserModel
        from AuthN.models import BaseUserModel
        base_user = get_object_or_404(BaseUserModel, id=target_user_id, role='user')
        
        # Try to get UserProfile for organization_id and admin_id from assignments
        try:
            user_profile = UserProfile.objects.get(user=base_user)
            org_id_val = user_profile.organization_id
            
            # Get current active admin using utility
            from utils.Employee.assignment_utils import get_current_admin_for_employee
            current_admin = get_current_admin_for_employee(base_user)
            admin_id_val = str(current_admin.id) if current_admin else None
        except UserProfile.DoesNotExist:
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "User profile not found. Please complete user setup.",
                "data": None
            }, status=status.HTTP_400_BAD_REQUEST)
        
        data = request.data.copy()
        data['user'] = target_user_id
        data['admin'] = admin_id_val
        data['organization'] = org_id_val
        # Set site if provided
        if site_id:
            data['site'] = site_id

        from_date = data.get('from_date')
        to_date = data.get('to_date')

        if from_date and to_date:
            from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
            to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
            total_days = (to_date_obj - from_date_obj).days + 1
            
            # If single day and half day, set to 0.5
            leave_day_type = data.get('leave_day_type')
            if not leave_day_type:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "leave_day_type is required",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if total_days == 1 and leave_day_type != 'full_day':
                total_days = Decimal('0.5')
            
            data['total_days'] = total_days

            # Check for overlapping leave applications - O(1) with index leaveapp_user_dates_status_idx
            overlapping_leaves = LeaveApplication.objects.filter(
                user_id=target_user_id,
                from_date__lte=to_date_obj,
                to_date__gte=from_date_obj
            ).exclude(status='cancelled').only('id', 'from_date', 'to_date', 'status', 'leave_type_id')
            
            if overlapping_leaves.exists():
                existing = overlapping_leaves.first()
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": f"Leave already applied for overlapping dates ({existing.from_date} to {existing.to_date}). Please choose different dates.",
                    "data": {
                        "existing_leave": {
                            "from_date": existing.from_date,
                            "to_date": existing.to_date,
                            "status": existing.status,
                            "leave_type": existing.leave_type.name
                        }
                    }
                }, status=status.HTTP_400_BAD_REQUEST)

            # Validate and fetch leave balance - O(1) with index leavebal_user_type_year_idx
            balance = EmployeeLeaveBalance.objects.filter(
                user_id=target_user_id,
                leave_type_id=data['leave_type'],
                year=from_date_obj.year
            ).select_related('leave_type').only('id', 'user_id', 'leave_type_id', 'year', 'assigned', 'used').first()

            if not balance:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "Leave balance not found for this year",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if balance.balance < total_days:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": f"Insufficient leave balance. Available: {balance.balance}, Required: {total_days}",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)

        serializer = LeaveApplicationSerializer(data=data)
        if serializer.is_valid():
            leave_app = serializer.save()

            # Sync balance based on all pending + approved leaves
            if from_date and to_date:
                from .balance_sync import sync_leave_balance
                sync_leave_balance(
                    user_id=target_user_id,
                    leave_type_id=leave_app.leave_type.id,
                    year=from_date_obj.year
                )

            return Response({
                "status": status.HTTP_201_CREATED,
                "message": "Leave application created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)
        
        return Response({
            "status": status.HTTP_400_BAD_REQUEST,
            "message": "Validation failed",
            "data": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request, site_id, user_id=None, pk=None):
        """Update leave application - Handle status changes and balance updates"""
        # Get admin_id based on role
        if request.user.role == 'admin':
            admin_id = request.user.id
            admin = request.user
        elif request.user.role == 'organization':
            # Organization role: get admin_id from query params
            admin_id = request.query_params.get('admin_id')
            if not admin_id:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate admin exists and belongs to organization
            try:
                admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                # Verify admin belongs to organization
                admin_profile = AdminProfile.objects.filter(
                    user=admin,
                    organization=request.user
                ).first()
                if not admin_profile:
                    return Response({
                        "status": status.HTTP_403_FORBIDDEN,
                        "message": "Admin does not belong to your organization",
                        "data": None
                    }, status=status.HTTP_403_FORBIDDEN)
            except BaseUserModel.DoesNotExist:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Admin not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Validate site belongs to admin
        try:
            site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
        except Site.DoesNotExist:
            return Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Site not found or you don't have permission to access this site",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Single O(1) query using index leaveapp_id_adm_idx
        leave = LeaveApplication.objects.filter(
            id=pk
        ).select_related('user', 'leave_type').only(
            'id', 'site_id', 'user_id', 'leave_type_id', 'from_date', 'status', 'total_days'
        ).first()
        
        if not leave:
            return Response({
                "status": status.HTTP_404_NOT_FOUND,
                "message": "Leave application not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        
        # O(1) site check
        if site_id and leave.site_id != site_id:
            return Response({
                "status": status.HTTP_404_NOT_FOUND,
                "message": "Leave application not found for this site"
            }, status=status.HTTP_404_NOT_FOUND)
        
        old_status = leave.status
        new_status = request.data.get('status')
        
        serializer = LeaveApplicationUpdateSerializer(leave, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            
            # Sync balance based on all pending + approved leaves
            if new_status and old_status != new_status:
                from .balance_sync import sync_leave_balance
                sync_leave_balance(
                    user_id=leave.user.id,
                    leave_type_id=leave.leave_type.id,
                    year=leave.from_date.year
                )
            
            return Response({
                "status": status.HTTP_200_OK,
                "message": f"Leave application {new_status or 'updated'} successfully",
                "data": serializer.data
            })
        return Response({
            "status": status.HTTP_400_BAD_REQUEST,
            "message": "Validation failed",
            "data": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, site_id, user_id=None, pk=None):
        """Cancel leave application - Only pending leaves can be cancelled"""
        # Get admin_id based on role
        if request.user.role == 'admin':
            admin_id = request.user.id
            admin = request.user
        elif request.user.role == 'organization':
            # Organization role: get admin_id from query params
            admin_id = request.query_params.get('admin_id')
            if not admin_id:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate admin exists and belongs to organization
            try:
                admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                # Verify admin belongs to organization
                admin_profile = AdminProfile.objects.filter(
                    user=admin,
                    organization=request.user
                ).first()
                if not admin_profile:
                    return Response({
                        "status": status.HTTP_403_FORBIDDEN,
                        "message": "Admin does not belong to your organization",
                        "data": None
                    }, status=status.HTTP_403_FORBIDDEN)
            except BaseUserModel.DoesNotExist:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Admin not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Validate site belongs to admin
        try:
            site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
        except Site.DoesNotExist:
            return Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Site not found or you don't have permission to access this site",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        leave = get_object_or_404(LeaveApplication, id=pk)
        
        # Filter by site if provided
        if site_id and leave.site_id != site_id:
            return Response({
                "status": status.HTTP_404_NOT_FOUND,
                "message": "Leave application not found for this site"
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Only pending leaves can be cancelled
        if leave.status != 'pending':
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": f"Cannot cancel {leave.status} leave. Only pending leaves can be cancelled.",
                "data": {
                    "current_status": leave.status,
                    "from_date": leave.from_date,
                    "to_date": leave.to_date
                }
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Store info before cancelling
        user_id = leave.user.id
        leave_type_id = leave.leave_type.id
        year = leave.from_date.year
        total_days = leave.total_days
        
        # Mark as cancelled
        leave.status = 'cancelled'
        leave.save()
        
        # Sync balance based on all pending + approved leaves
        from .balance_sync import sync_leave_balance
        sync_leave_balance(user_id, leave_type_id, year)
        
        return Response({
            "status": status.HTTP_200_OK,
            "message": "Leave application cancelled successfully. Balance restored.",
            "data": {
                "cancelled_leave": {
                    "id": leave.id,
                    "from_date": leave.from_date,
                    "to_date": leave.to_date,
                    "total_days": total_days,
                    "status": leave.status
                },
                "restored_balance": float(total_days)
            }
        })
