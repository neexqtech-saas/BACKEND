"""
Payroll System Views
Optimized for high-traffic, low-cost, future-proof architecture
All queries O(1) or using proper database indexes
"""
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.db import models as db_models
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from .models import (
    EmployeePayrollConfig,
    PayslipGenerator, ProfessionalTaxRule, OrganizationPayrollSettings,
    SalaryComponent, SalaryStructure, SalaryStructureItem, EmployeeBankInfo, EmployeeAdvance,
    EmployeeCustomMonthlyEarning, EmployeeCustomMonthlyDeduction, GeneratedPayrollRecord
)
from .serializers import (
    EmployeePayrollConfigCreateUpdateSerializer,
    EmployeePayrollConfigResponseSerializer,
    PayslipGeneratorSerializer,
    PayslipGeneratorCreateSerializer,
    PayslipGeneratorUpdateSerializer,
    PayslipGeneratorListSerializer,
    ProfessionalTaxRuleSerializer,
    OrganizationPayrollSettingsSerializer,
    OrganizationPayrollSettingsCreateSerializer,
    OrganizationPayrollSettingsUpdateSerializer,
    SalaryStructureResponseSerializer,
    SalaryStructureCreateSerializer,
    SalaryStructureUpdateSerializer,
    DeductionsUpdateItemSerializer,
    EmployeeBankInfoSerializer,
    EmployeeBankInfoCreateSerializer,
    EmployeeBankInfoUpdateSerializer,
    EmployeeBankInfoListSerializer,
    EmployeeAdvanceSerializer,
    EmployeeAdvanceCreateSerializer,
    EmployeeAdvanceListSerializer,
)
from AuthN.models import BaseUserModel, UserProfile, AdminProfile
from SiteManagement.models import Site
from utils.site_filter_utils import filter_queryset_by_site
from decimal import Decimal
from .utils import (
    calculate_pf_employee, calculate_pf_employer,
    calculate_esi_employee, calculate_esi_employer,
    calculate_pt, calculate_gratuity,
    is_employee_component, is_employer_component,
    is_basic_component, is_da_component, is_special_allowance_component,
    calculate_payroll_breakdown_optimized,
    prefetch_payroll_data,
    get_statutory_components_map,
    build_statutory_components_list,
    get_all_employee_payroll_details
)
from calendar import monthrange


# ==================== HELPER FUNCTIONS ====================

def get_admin_and_site_for_payroll(request, site_id):
    """
    Optimized admin and site validation - O(1) queries with select_related
    Returns: (admin, site, None) tuple or (None, None, Response with error)
    """
    user = request.user
    
    # Fast path for admin role - O(1) query
    if user.role == 'admin':
        admin_id = user.id
        admin = user
        # Single O(1) query with index on (id, created_by_admin, is_active)
        try:
            site = Site.objects.only('id', 'site_name', 'created_by_admin_id', 'is_active').get(
                id=site_id, 
                created_by_admin_id=admin_id, 
                is_active=True
            )
            return admin, site, None
        except Site.DoesNotExist:
            return None, None, Response({
                "response": False,
                "message": "Site not found or you don't have permission to access this site",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
    
    # Organization role - O(1) queries with select_related
    elif user.role == 'organization':
        admin_id = request.query_params.get('admin_id')
        if not admin_id:
            return None, None, Response({
                "response": False,
                "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                "data": None
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Single O(1) query with select_related to avoid N+1 - uses index on (id, role)
        try:
            admin = BaseUserModel.objects.only(
                'id', 'role', 'email'
            ).get(id=admin_id, role='admin')
        except BaseUserModel.DoesNotExist:
            return None, None, Response({
                "response": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        
        # O(1) query - verify admin belongs to organization using select_related
        admin_profile = AdminProfile.objects.select_related('user', 'organization').only(
            'id', 'user_id', 'organization_id'
        ).filter(
            user_id=admin_id,
            organization_id=user.id
        ).first()
        
        if not admin_profile:
            return None, None, Response({
                "response": False,
                "message": "Admin does not belong to your organization",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Single O(1) query with index on (id, created_by_admin, is_active)
        try:
            site = Site.objects.only('id', 'site_name', 'created_by_admin_id', 'is_active').get(
                id=site_id, 
                created_by_admin_id=admin_id, 
                is_active=True
            )
            return admin, site, None
        except Site.DoesNotExist:
            return None, None, Response({
                "response": False,
                "message": "Site not found or you don't have permission to access this site",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
    
    else:
        return None, None, Response({
            "response": False,
            "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
            "data": None
        }, status=status.HTTP_403_FORBIDDEN)

def get_or_create_statutory_component(organization, code, name, component_type, statutory_type=None):
    """Get or create a statutory salary component"""
    component, created = SalaryComponent.objects.get_or_create(
        organization=organization,
        code=code,
        defaults={
            'name': name,
            'component_type': component_type,
            'statutory_type': statutory_type,
            'is_active': True,
        }
    )
    # Update existing component if needed (in case settings changed)
    if not created:
        component.name = name
        component.component_type = component_type
        component.statutory_type = statutory_type
        component.is_active = True
        component.save()
    return component


def ensure_statutory_components_for_settings(payroll_settings):
    """Ensure statutory components are created based on payroll settings"""
    organization = payroll_settings.organization
    
    # PF Components (if enabled)
    if payroll_settings.pf_enabled:
        get_or_create_statutory_component(
            organization, 'PF_EMP', 'Provident Fund (PF – Employee)', 
            SalaryComponent.DEDUCTION, 'PF'
        )
        get_or_create_statutory_component(
            organization, 'PF_EMPR', 'Provident Fund (PF – Employer)', 
            SalaryComponent.DEDUCTION, 'PF'
        )
    
    # ESIC Components (if enabled)
    if payroll_settings.esi_enabled:
        get_or_create_statutory_component(
            organization, 'ESIC_EMP', 'Employee State Insurance (ESI – Employee)', 
            SalaryComponent.DEDUCTION, 'ESI'
        )
        get_or_create_statutory_component(
            organization, 'ESIC_EMPR', 'Employee State Insurance (ESI – Employer)', 
            SalaryComponent.DEDUCTION, 'ESI'
        )
    
    # PT (if enabled)
    if payroll_settings.pt_enabled:
        get_or_create_statutory_component(
            organization, 'PT', 'Professional Tax (PT)', 
            SalaryComponent.DEDUCTION, 'PT'
        )
    
    # Gratuity (if enabled)
    if payroll_settings.gratuity_enabled:
        get_or_create_statutory_component(
            organization, 'GRATUITY', 'Gratuity', 
            SalaryComponent.DEDUCTION, 'GRATUITY'
        )


# ==================== PAYSLIP GENERATOR VIEWS ====================

class PayslipGeneratorAPIView(APIView):
    """Payslip Generator CRUD API View"""
    
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def get(self, request, site_id, pk=None):
        """Get payslip(s) for admin - O(1) queries with index optimization"""
        try:
            admin, site, error_response = get_admin_and_site_for_payroll(request, site_id)
            if error_response:
                return error_response
            
            admin_id = admin.id
            
            if pk:
                # Single O(1) query using index payslip_id_adm_idx (id, admin)
                payslip = PayslipGenerator.objects.filter(
                    id=pk,
                    admin_id=admin_id
                ).select_related('admin', 'employee').only(
                    'id', 'admin_id', 'site_id', 'employee_id', 'payslip_number', 'month', 'year',
                    'pay_date', 'paid_days', 'loss_of_pay_days', 'template', 'currency',
                    'company_name', 'employee_name', 'employee_code', 'designation',
                    'total_earnings', 'total_deductions', 'net_pay', 'created_at', 'updated_at'
                ).first()
                
                if not payslip:
                    return Response({
                        "response": False,
                        "message": "Payslip not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
                
                # O(1) site check
                if payslip.site_id != site_id:
                    return Response({
                        "response": False,
                        "message": "Payslip not found for this site",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
                
                serializer = PayslipGeneratorSerializer(payslip, context={'request': request})
                return Response({
                    "response": True,
                    "message": "Payslip fetched successfully",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)
            else:
                # List query with index optimization - uses payslip_adm_created_idx (admin, created_at)
                payslips = PayslipGenerator.objects.filter(
                    admin_id=admin_id
                ).select_related('admin', 'employee').only(
                    'id', 'admin_id', 'employee_id', 'payslip_number', 'month', 'year',
                    'pay_date', 'paid_days', 'loss_of_pay_days', 'template', 'currency',
                    'company_name', 'employee_name', 'employee_code', 'designation',
                    'total_earnings', 'total_deductions', 'net_pay', 'created_at', 'updated_at'
                )
                
                # Filter by site if provided
                payslips = filter_queryset_by_site(payslips, site_id, 'site')
                payslips = payslips.order_by('-created_at')
                
                serializer = PayslipGeneratorListSerializer(payslips, many=True)
                return Response({
                    "response": True,
                    "message": "Payslips fetched successfully",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)
                
        except BaseUserModel.DoesNotExist:
            return Response({
                "response": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except PayslipGenerator.DoesNotExist:
            return Response({
                "response": False,
                "message": "Payslip not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "response": False,
                "message": f"Error fetching payslip: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request, site_id):
        """Create payslip - Optimized"""
        try:
            admin, site, error_response = get_admin_and_site_for_payroll(request, site_id)
            if error_response:
                return error_response
            
            data = request.data.copy()
            data['admin_id'] = str(admin.id)
            # Set site
            data['site'] = str(site.id)
            
            serializer = PayslipGeneratorCreateSerializer(data=data)
            if serializer.is_valid():
                payslip = serializer.save()
                response_serializer = PayslipGeneratorSerializer(payslip, context={'request': request})
                return Response({
                    "response": True,
                    "message": "Payslip created successfully",
                    "data": response_serializer.data
                }, status=status.HTTP_201_CREATED)
            
            return Response({
                "response": False,
                "message": "Validation failed",
                "data": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "response": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "response": False,
                "message": f"Error creating payslip: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def put(self, request, site_id, pk=None):
        """Update payslip"""
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "response": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "response": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "response": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "response": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validate site belongs to admin
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "response": False,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            payslip = PayslipGenerator.objects.get(id=pk, admin=admin)
            
            serializer = PayslipGeneratorUpdateSerializer(payslip, data=request.data, partial=True)
            if serializer.is_valid():
                updated_payslip = serializer.save()
                response_serializer = PayslipGeneratorSerializer(updated_payslip, context={'request': request})
                return Response({
                    "response": True,
                    "message": "Payslip updated successfully",
                    "data": response_serializer.data
                }, status=status.HTTP_200_OK)
            
            return Response({
                "response": False,
                "message": "Validation failed",
                "data": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "response": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except PayslipGenerator.DoesNotExist:
            return Response({
                "response": False,
                "message": "Payslip not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "response": False,
                "message": f"Error updating payslip: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def delete(self, request, site_id, pk=None):
        """Delete payslip"""
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "response": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "response": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "response": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "response": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validate site belongs to admin
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "response": False,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            payslip = PayslipGenerator.objects.get(id=pk, admin=admin)
            
            # Filter by site
            if payslip.site_id != site_id:
                return Response({
                    "response": False,
                    "message": "Payslip not found for this site",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            payslip.delete()
            
            return Response({
                "response": True,
                "message": "Payslip deleted successfully",
                "data": None
            }, status=status.HTTP_200_OK)
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "response": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except PayslipGenerator.DoesNotExist:
            return Response({
                "response": False,
                "message": "Payslip not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "response": False,
                "message": f"Error deleting payslip: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== PROFESSIONAL TAX RULE VIEWS ====================

class ProfessionalTaxRuleAPIView(APIView):
    """Professional Tax Rule API View"""
    
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Get all professional tax rules"""
        try:
            rules = ProfessionalTaxRule.objects.filter(is_active=True).order_by('state_name', 'salary_from')
            serializer = ProfessionalTaxRuleSerializer(rules, many=True)
            return Response({
                "response": True,
                "message": "Professional tax rules fetched successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                "response": False,
                "message": f"Error fetching rules: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== ORGANIZATION PAYROLL SETTINGS VIEWS ====================

class OrganizationPayrollSettingsAPIView(APIView):
    """Organization Payroll Settings API View"""
    
    permission_classes = [IsAuthenticated]
    
    def get(self, request, org_id):
        """Get payroll settings for organization"""
        try:
            organization = BaseUserModel.objects.get(id=org_id, role='organization')
            settings = OrganizationPayrollSettings.objects.filter(organization=organization).first()
            
            if settings:
                serializer = OrganizationPayrollSettingsSerializer(settings)
                return Response({
                    "response": True,
                    "message": "Payroll settings fetched successfully",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    "response": False,
                    "message": "Payroll settings not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
                
        except BaseUserModel.DoesNotExist:
            return Response({
                "response": False,
                "message": "Organization not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "response": False,
                "message": f"Error fetching settings: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request, org_id):
        """Create payroll settings (or update if exists)"""
        try:
            organization = BaseUserModel.objects.get(id=org_id, role='organization')
            
            # Check if settings already exist
            existing_settings = OrganizationPayrollSettings.objects.filter(organization=organization).first()
            
            if existing_settings:
                # If settings exist, use PUT logic instead
                serializer = OrganizationPayrollSettingsUpdateSerializer(existing_settings, data=request.data, partial=True)
                if serializer.is_valid():
                    updated_settings = serializer.save()
                    # Ensure statutory components are created based on enabled settings
                    ensure_statutory_components_for_settings(updated_settings)
                    response_serializer = OrganizationPayrollSettingsSerializer(updated_settings)
                    return Response({
                        "response": True,
                        "message": "Payroll settings updated successfully",
                        "data": response_serializer.data
                    }, status=status.HTTP_200_OK)
                
                return Response({
                    "response": False,
                    "message": "Validation failed",
                    "data": serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Create new settings
            data = request.data.copy()
            data['organization_id'] = org_id
            
            serializer = OrganizationPayrollSettingsCreateSerializer(data=data)
            if serializer.is_valid():
                settings = serializer.save()
                # Ensure statutory components are created based on enabled settings
                ensure_statutory_components_for_settings(settings)
                response_serializer = OrganizationPayrollSettingsSerializer(settings)
                return Response({
                    "response": True,
                    "message": "Payroll settings created successfully",
                    "data": response_serializer.data
                }, status=status.HTTP_201_CREATED)
            
            return Response({
                "response": False,
                "message": "Validation failed",
                "data": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "response": False,
                "message": "Organization not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "response": False,
                "message": f"Error creating/updating settings: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def put(self, request, org_id):
        """Update payroll settings"""
        try:
            organization = BaseUserModel.objects.get(id=org_id, role='organization')
            settings = OrganizationPayrollSettings.objects.get(organization=organization)
            
            serializer = OrganizationPayrollSettingsUpdateSerializer(settings, data=request.data, partial=True)
            if serializer.is_valid():
                updated_settings = serializer.save()
                # Ensure statutory components are created based on enabled settings
                ensure_statutory_components_for_settings(updated_settings)
                response_serializer = OrganizationPayrollSettingsSerializer(updated_settings)
                return Response({
                    "response": True,
                    "message": "Payroll settings updated successfully",
                    "data": response_serializer.data
                }, status=status.HTTP_200_OK)
            
            return Response({
                "response": False,
                "message": "Validation failed",
                "data": serializer.errors,
                'status': False
            },
            status=status.HTTP_400_BAD_REQUEST
        )
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "response": False,
                "message": "Organization not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except OrganizationPayrollSettings.DoesNotExist:
            return Response({
                "response": False,
                "message": "Payroll settings not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "response": False,
                "message": f"Error updating settings: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== SALARY STRUCTURE UNIFIED API VIEW ====================

class SalaryStructureUnifiedAPIView(APIView):
    """
    Unified CRUD API for Salary Structure
    Single URL: /api/payroll/<org_id>/salary-structure/
    Handles: GET, POST, PUT, DELETE
    """
    
    permission_classes = [IsAuthenticated]
    
    def _get_or_create_component(self, organization, code, name, component_type, statutory_type=None):
        """Get or create a salary component"""
        component, created = SalaryComponent.objects.get_or_create(
            organization=organization,
            code=code,
            defaults={
                'name': name,
                'component_type': component_type,
                'statutory_type': statutory_type,
                'is_active': True,
            }
        )
        return component
    
    def _ensure_default_components(self, organization):
        """
        Ensure all default components exist (Basic, Special Allowance, Statutory)
        Optimized: Fetches existing components first, then creates missing ones
        """
        components = {}
        
        # Define all required component codes
        required_components = [
            ('BASIC', 'Basic Salary', SalaryComponent.EARNING, None),
            ('SPECIAL_ALLOWANCE', 'Special Allowance', SalaryComponent.EARNING, None),
        ]
        
        # Get payroll settings to check which statutory components are enabled
        try:
            payroll_settings = OrganizationPayrollSettings.objects.select_related('organization').get(
                organization=organization
            )
            
            # Add statutory components based on settings
            if payroll_settings.pf_enabled:
                required_components.extend([
                    ('PF_EMP', 'Provident Fund (PF – Employee)', SalaryComponent.DEDUCTION, 'PF'),
                    ('PF_EMPR', 'Provident Fund (PF – Employer)', SalaryComponent.DEDUCTION, 'PF'),
                ])
            
            if payroll_settings.esi_enabled:
                required_components.extend([
                    ('ESIC_EMP', 'Employee State Insurance (ESI – Employee)', SalaryComponent.DEDUCTION, 'ESI'),
                    ('ESIC_EMPR', 'Employee State Insurance (ESI – Employer)', SalaryComponent.DEDUCTION, 'ESI'),
                ])
            
            if payroll_settings.pt_enabled:
                required_components.append(('PT', 'Professional Tax (PT)', SalaryComponent.DEDUCTION, 'PT'))
            
            if payroll_settings.gratuity_enabled:
                required_components.append(('GRATUITY', 'Gratuity', SalaryComponent.DEDUCTION, 'GRATUITY'))
        except OrganizationPayrollSettings.DoesNotExist:
            # If no settings, skip statutory components
            pass
        
        # Optimized: Fetch all existing components in one query
        component_codes = [code for code, _, _, _ in required_components]
        existing_components = {
            comp.code: comp
            for comp in SalaryComponent.objects.filter(
                organization=organization,
                code__in=component_codes
            )
        }
        
        # Get or create components
        for code, name, component_type, statutory_type in required_components:
            if code in existing_components:
                components[code] = existing_components[code]
            else:
                # Create missing component
                components[code] = self._get_or_create_component(
                    organization, code, name, component_type, statutory_type
                )
        
        return components
    
    def _get_statutory_value(self, component, payroll_settings):
        """Get display value for statutory component"""
        if not component.statutory_type:
            return None
        
        if component.statutory_type == 'PF':
            if 'Employee' in component.name or 'EMP' in component.code:
                return f"{payroll_settings.pf_employee_percentage}%" if payroll_settings.pf_enabled else None
            elif 'Employer' in component.name or 'EMPR' in component.code:
                return f"{payroll_settings.pf_employer_percentage}%" if payroll_settings.pf_enabled else None
        
        elif component.statutory_type == 'ESI':
            if 'Employee' in component.name or 'EMP' in component.code:
                return f"{payroll_settings.esi_employee_percentage}%" if payroll_settings.esi_enabled else None
            elif 'Employer' in component.name or 'EMPR' in component.code:
                return f"{payroll_settings.esi_employer_percentage}%" if payroll_settings.esi_enabled else None
        
        elif component.statutory_type == 'GRATUITY':
            return f"{payroll_settings.gratuity_percentage}%" if payroll_settings.gratuity_enabled else None
        
        elif component.statutory_type == 'PT':
            return "Auto" if payroll_settings.pt_enabled else None
        
        return None
    
    def get(self, request, org_id):
        """
        GET - Fetch Salary Structure (Auto-grouped) or List all structures
        If list=true, returns list of all structures
        Otherwise returns single structure details
        """
        try:
            organization = BaseUserModel.objects.get(id=org_id, role='organization')
            
            # Check if listing all structures
            list_all = request.query_params.get('list', 'false').lower() == 'true'
            if list_all:
                # Optimized: Single query for all structures
                structures = SalaryStructure.objects.filter(
                    organization=organization,
                    is_active=True
                ).order_by('-is_default', '-created_at')
                
                # Convert to list in one go - no queries in loop
                structures_list = [
                    {
                        'structure_id': structure.id,
                        'name': structure.name,
                        'description': structure.description or '',
                        'is_default': structure.is_default,
                        'created_at': structure.created_at.isoformat() if structure.created_at else None,
                        'updated_at': structure.updated_at.isoformat() if structure.updated_at else None,
                    }
                    for structure in structures
                ]
                
                # Optimized: Fetch statutory components in one query + build list
                statutory_list = []
                try:
                    # Single query for payroll settings
                    payroll_settings = OrganizationPayrollSettings.objects.select_related('organization').get(
                        organization=organization
                    )
                    
                    # Single query to fetch all statutory components for organization
                    components_map = get_statutory_components_map(organization)
                    
                    # Build list from prefetched data - no queries
                    statutory_list = build_statutory_components_list(payroll_settings, components_map)
                except OrganizationPayrollSettings.DoesNotExist:
                    # If no settings, no statutory components are enabled
                    pass
                
                return Response({
                    "response": True,
                    "message": "Salary structures fetched successfully",
                    "data": {
                        "structures": structures_list,
                        "statutory_components": statutory_list
                    }
                }, status=status.HTTP_200_OK)
            
            # Get structure (default or first active)
            structure_id_param = request.query_params.get('structure_id')
            if structure_id_param:
                structure = SalaryStructure.objects.get(
                    id=structure_id_param,
                    organization=organization
                )
            else:
                # Get default structure or first active
                structure = SalaryStructure.objects.filter(
                    organization=organization,
                    is_active=True
                ).order_by('-is_default', 'id').first()
            
            if not structure:
                return Response({
                    "response": False,
                    "message": "No salary structure found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Get payroll settings for statutory values
            try:
                payroll_settings = OrganizationPayrollSettings.objects.get(organization=organization)
            except OrganizationPayrollSettings.DoesNotExist:
                payroll_settings = None
            
            # Get structure items - optimized with select_related
            items = SalaryStructureItem.objects.filter(
                structure=structure
            ).select_related('component', 'calculation_base').order_by('order', 'id')
            
            earnings = []
            deductions = []
            
            for item in items:
                component = item.component
                is_statutory = bool(component.statutory_type)
                is_special_allowance = component.code == 'SPECIAL_ALLOWANCE'
                
                # Determine editable
                editable = not is_statutory and not is_special_allowance
                
                # Get value display
                if is_statutory:
                    value = self._get_statutory_value(component, payroll_settings) if payroll_settings else "Auto"
                elif item.calculation_type == 'percentage':
                    value = f"{item.value}%" if item.value else "0%"
                elif item.calculation_type == 'fixed':
                    value = str(item.value) if item.value else "0"
                else:
                    value = "Auto"
                
                item_data = {
                    "component": component.code,
                    "label": component.name,
                    "calculation_type": item.calculation_type,
                    "value": value,
                    "editable": editable
                }
                
                if component.component_type == SalaryComponent.EARNING:
                    earnings.append(item_data)
                else:
                    deductions.append(item_data)
            
            response_data = {
                "structure_id": structure.id,
                "name": structure.name,
                "earnings": earnings,
                "deductions": deductions
            }
            
            serializer = SalaryStructureResponseSerializer(response_data)
            
            return Response({
                "response": True,
                "message": "Salary structure fetched successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "response": False,
                "message": "Organization not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except SalaryStructure.DoesNotExist:
            return Response({
                "response": False,
                "message": "Salary structure not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "response": False,
                "message": f"Error fetching structure: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request, org_id):
        """
        POST - Create Salary Structure (AUTO MAGIC)
        Auto-creates Basic, Special Allowance, and statutory deductions
        Also accepts custom earnings and deductions in the same request
        """
        try:
            organization = BaseUserModel.objects.get(id=org_id, role='organization')
            
            serializer = SalaryStructureCreateSerializer(data=request.data)
            if not serializer.is_valid():
                return Response({
                    "response": False,
                    "message": "Validation failed",
                    "data": serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check for duplicate name
            structure_name = serializer.validated_data['name']
            existing_structure = SalaryStructure.objects.filter(
                organization=organization,
                name=structure_name,
                is_active=True
            ).first()
            
            if existing_structure:
                return Response({
                    "response": False,
                    "message": f"Salary structure with name '{structure_name}' already exists. Please use a different name.",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            with transaction.atomic():
                # 1. Create Structure
                structure = SalaryStructure.objects.create(
                    organization=organization,
                    name=structure_name,
                    description=serializer.validated_data.get('description', ''),
                    is_default=serializer.validated_data.get('is_default', False),
                    is_active=True
                )
                
                # 2. Ensure all default components exist
                components = self._ensure_default_components(organization)
                
                # 3. Get custom earnings and deductions from request
                custom_earnings_data = serializer.validated_data.get('earnings', [])
                custom_deductions_data = serializer.validated_data.get('deductions', [])
                
                # 4. Create Structure Items
                order = 1
                
                # EARNINGS: Basic Salary (50% default, or from request)
                basic_data = None
                for earning_data in custom_earnings_data:
                    if earning_data['component'] == 'BASIC':
                        basic_data = earning_data
                        break
                
                basic_component = components['BASIC']
                if basic_data:
                    # Use custom Basic value
                    SalaryStructureItem.objects.create(
                        structure=structure,
                        component=basic_component,
                        calculation_type=basic_data['calculation_type'],
                        value=float(basic_data['value']),
                        order=order
                    )
                else:
                    # Default Basic (50%)
                    SalaryStructureItem.objects.create(
                        structure=structure,
                        component=basic_component,
                        calculation_type='percentage',
                        value=50.0,
                        order=order
                    )
                order += 1
                
                # EARNINGS: Special Allowance (0, locked) - always create
                special_allowance_component = components['SPECIAL_ALLOWANCE']
                SalaryStructureItem.objects.create(
                    structure=structure,
                    component=special_allowance_component,
                    calculation_type='fixed',
                    value=0.0,
                    order=order
                )
                order += 1
                
                # EARNINGS: Custom earnings (excluding BASIC and SPECIAL_ALLOWANCE)
                # Optimized: Pre-fetch existing custom earning components
                updated_earning_codes = {'BASIC', 'SPECIAL_ALLOWANCE'}
                custom_earning_codes = [
                    ed['component'] for ed in custom_earnings_data
                    if ed['component'] not in ['BASIC', 'SPECIAL_ALLOWANCE']
                ]
                
                # Fetch existing components in one query
                existing_custom_components = {
                    comp.code: comp
                    for comp in SalaryComponent.objects.filter(
                        organization=organization,
                        code__in=custom_earning_codes
                    )
                }
                
                for earning_data in custom_earnings_data:
                    component_code = earning_data['component']
                    
                    # Skip Basic and Special Allowance (already handled)
                    if component_code in ['BASIC', 'SPECIAL_ALLOWANCE']:
                        continue
                    
                    # Validate: cannot add statutory components as earnings
                    if component_code in components and components[component_code].statutory_type:
                        continue  # Skip statutory components
                    
                    # Get or create component (use prefetched if available)
                    if component_code in existing_custom_components:
                        component = existing_custom_components[component_code]
                        # Update if needed
                        if component.component_type != SalaryComponent.EARNING:
                            component.component_type = SalaryComponent.EARNING
                            component.save()
                    else:
                        component = self._get_or_create_component(
                            organization,
                            component_code,
                            earning_data.get('label', component_code.replace('_', ' ').title()),
                            SalaryComponent.EARNING
                        )
                    
                    # Validate component is not statutory
                    if component.statutory_type:
                        continue  # Skip statutory components
                    
                    # Create structure item
                    SalaryStructureItem.objects.create(
                        structure=structure,
                        component=component,
                        calculation_type=earning_data['calculation_type'],
                        value=float(earning_data['value']),
                        order=order
                    )
                    updated_earning_codes.add(component_code)
                    order += 1
                
                # DEDUCTIONS: Statutory components (auto, locked) - always create
                statutory_order = order
                for code, component in components.items():
                    if code not in ['BASIC', 'SPECIAL_ALLOWANCE']:
                        # Statutory deduction
                        SalaryStructureItem.objects.create(
                            structure=structure,
                            component=component,
                            calculation_type='auto',
                            value=None,
                            order=statutory_order
                        )
                        statutory_order += 1
                
                # DEDUCTIONS: Custom non-statutory deductions
                # Optimized: Fetch statutory deduction codes ONCE before loop
                    statutory_deduction_codes = set(
                        SalaryStructureItem.objects.filter(
                            structure=structure,
                            component__component_type=SalaryComponent.DEDUCTION,
                            component__statutory_type__isnull=False
                        ).values_list('component__code', flat=True)
                    )
                
                # Optimized: Pre-fetch existing custom deduction components
                custom_deduction_codes = [dd['component'] for dd in custom_deductions_data]
                existing_deduction_components = {
                    comp.code: comp
                    for comp in SalaryComponent.objects.filter(
                        organization=organization,
                        code__in=custom_deduction_codes
                    )
                }
                
                order = statutory_order
                for deduction_data in custom_deductions_data:
                    component_code = deduction_data['component']
                    
                    # Validate: cannot add statutory deductions
                    if component_code in statutory_deduction_codes:
                        continue  # Skip statutory deductions
                    
                    # Get or create component (non-statutory deduction)
                    # Use prefetched if available
                    if component_code in existing_deduction_components:
                        component = existing_deduction_components[component_code]
                        # Update if needed
                        if component.component_type != SalaryComponent.DEDUCTION:
                            component.component_type = SalaryComponent.DEDUCTION
                            component.save()
                    else:
                        component = self._get_or_create_component(
                            organization,
                            component_code,
                            deduction_data.get('label', component_code.replace('_', ' ').title()),
                            SalaryComponent.DEDUCTION
                        )
                    
                    # Validate component is not statutory
                    if component.statutory_type:
                        continue  # Skip statutory components
                    
                    # Create structure item
                    SalaryStructureItem.objects.create(
                        structure=structure,
                        component=component,
                        calculation_type=deduction_data['calculation_type'],
                        value=float(deduction_data['value']),
                        order=order
                    )
                    order += 1
                
                # Return created structure with same format as GET
                # Get payroll settings for statutory values
                try:
                    payroll_settings = OrganizationPayrollSettings.objects.get(organization=organization)
                except OrganizationPayrollSettings.DoesNotExist:
                    payroll_settings = None
                
                # Get structure items
                items = SalaryStructureItem.objects.filter(
                    structure=structure
                ).select_related('component').order_by('order', 'id')
                
                earnings = []
                deductions = []
                
                for item in items:
                    component = item.component
                    is_statutory = bool(component.statutory_type)
                    is_special_allowance = component.code == 'SPECIAL_ALLOWANCE'
                    
                    # Determine editable
                    editable = not is_statutory and not is_special_allowance
                    
                    # Get value display
                    if is_statutory:
                        value = self._get_statutory_value(component, payroll_settings) if payroll_settings else "Auto"
                    elif item.calculation_type == 'percentage':
                        value = f"{item.value}%" if item.value else "0%"
                    elif item.calculation_type == 'fixed':
                        value = str(item.value) if item.value else "0"
                    else:
                        value = "Auto"
                    
                    item_data = {
                        "component": component.code,
                        "label": component.name,
                        "calculation_type": item.calculation_type,
                        "value": value,
                        "editable": editable
                    }
                    
                    if component.component_type == SalaryComponent.EARNING:
                        earnings.append(item_data)
                    else:
                        deductions.append(item_data)
                
                response_data = {
                    "structure_id": structure.id,
                    "name": structure.name,
                    "earnings": earnings,
                    "deductions": deductions
                }
                
                serializer = SalaryStructureResponseSerializer(response_data)
                
                return Response({
                    "response": True,
                    "message": "Salary structure created successfully",
                    "data": serializer.data
                }, status=status.HTTP_201_CREATED)
                
        except BaseUserModel.DoesNotExist:
            return Response({
                "response": False,
                "message": "Organization not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "response": False,
                "message": f"Error creating structure: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def put(self, request, org_id):
        """
        PUT - Update Salary Structure (ONLY EARNINGS)
        Only earnings can be updated, deductions are locked
        """
        try:
            organization = BaseUserModel.objects.get(id=org_id, role='organization')
            
            serializer = SalaryStructureUpdateSerializer(data=request.data)
            if not serializer.is_valid():
                return Response({
                    "response": False,
                    "message": "Validation failed",
                    "data": serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get structure (default or first active)
            structure_id_param = request.query_params.get('structure_id')
            if structure_id_param:
                structure = SalaryStructure.objects.get(
                    id=structure_id_param,
                    organization=organization
                )
            else:
                structure = SalaryStructure.objects.filter(
                    organization=organization,
                    is_active=True
                ).order_by('-is_default', 'id').first()
            
            if not structure:
                return Response({
                    "response": False,
                    "message": "Salary structure not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Check for duplicate name if name is being updated
            new_name = serializer.validated_data.get('name')
            if new_name:
                # If updating name, check for duplicate (excluding current structure)
                existing_structure = SalaryStructure.objects.filter(
                    organization=organization,
                    name=new_name,
                    is_active=True
                ).exclude(id=structure.id).first()
                
                if existing_structure:
                    return Response({
                        "response": False,
                        "message": f"Salary structure with name '{new_name}' already exists. Please use a different name.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            with transaction.atomic():
                # Update name if provided
                if new_name:
                    structure.name = new_name
                    structure.save()
                
                earnings_data = serializer.validated_data['earnings']
                
                # Ensure default components exist
                components = self._ensure_default_components(organization)
                
                # Separate Basic from other earnings
                basic_data = None
                other_earnings = []
                
                for earning_data in earnings_data:
                    if earning_data['component'] == 'BASIC':
                        basic_data = earning_data
                    else:
                        other_earnings.append(earning_data)
                
                # Update Basic Salary (order 1)
                if basic_data:
                    basic_component = components.get('BASIC') or self._get_or_create_component(
                        organization, 'BASIC', 'Basic Salary', SalaryComponent.EARNING
                    )
                    basic_item, _ = SalaryStructureItem.objects.get_or_create(
                        structure=structure,
                        component=basic_component,
                        defaults={'order': 1}
                    )
                    basic_item.calculation_type = basic_data['calculation_type']
                    basic_item.value = basic_data['value']
                    basic_item.order = 1
                    basic_item.save()
                
                # Ensure Special Allowance exists (order 2, locked)
                special_allowance_component = components.get('SPECIAL_ALLOWANCE') or self._get_or_create_component(
                    organization, 'SPECIAL_ALLOWANCE', 'Special Allowance', SalaryComponent.EARNING
                )
                special_allowance_item, _ = SalaryStructureItem.objects.get_or_create(
                    structure=structure,
                    component=special_allowance_component,
                    defaults={
                        'calculation_type': 'fixed',
                        'value': 0.0,
                        'order': 2
                    }
                )
                # Keep Special Allowance locked (don't update)
                
                # Update other earnings (order 3+)
                updated_codes = {'BASIC', 'SPECIAL_ALLOWANCE'}
                order = 3
                
                for earning_data in other_earnings:
                    component_code = earning_data['component']
                    
                    # Validate component code
                    if component_code == 'SPECIAL_ALLOWANCE':
                        return Response({
                            "response": False,
                            "message": "Special Allowance cannot be updated. It is system-managed.",
                            "data": None
                        }, status=status.HTTP_400_BAD_REQUEST)
                    
                    # Get or create component
                    component = self._get_or_create_component(
                        organization,
                        component_code,
                        earning_data.get('label', component_code.replace('_', ' ').title()),
                        SalaryComponent.EARNING
                    )
                    
                    # Check if component is statutory
                    if component.statutory_type:
                        return Response({
                            "response": False,
                            "message": f"Component '{component.name}' is statutory and cannot be updated.",
                            "data": None
                        }, status=status.HTTP_400_BAD_REQUEST)
                    
                    # Get or create structure item
                    structure_item, created = SalaryStructureItem.objects.get_or_create(
                        structure=structure,
                        component=component,
                        defaults={
                            'calculation_type': earning_data['calculation_type'],
                            'value': earning_data['value'],
                            'order': order
                        }
                    )
                    
                    if not created:
                        # Update existing item
                        structure_item.calculation_type = earning_data['calculation_type']
                        structure_item.value = earning_data['value']
                        structure_item.order = order
                        structure_item.save()
                    
                    updated_codes.add(component_code)
                    order += 1
                
                # Delete earning items that are not in the update (except Basic and Special Allowance)
                SalaryStructureItem.objects.filter(
                    structure=structure,
                    component__component_type=SalaryComponent.EARNING
                ).exclude(component__code__in=updated_codes).delete()
                
                # Handle deductions (non-statutory only)
                deductions_data = serializer.validated_data.get('deductions', [])
                if deductions_data:
                    # Get existing statutory deduction codes to exclude
                    statutory_deduction_codes = set(
                        SalaryStructureItem.objects.filter(
                            structure=structure,
                            component__component_type=SalaryComponent.DEDUCTION,
                            component__statutory_type__isnull=False
                        ).values_list('component__code', flat=True)
                    )
                    
                    # Track updated deduction codes
                    updated_deduction_codes = set(statutory_deduction_codes)  # Keep statutory ones
                    deduction_order = 100  # Start after statutory deductions
                    
                    for deduction_data in deductions_data:
                        component_code = deduction_data['component']
                        
                        # Validate: cannot update statutory deductions
                        if component_code in statutory_deduction_codes:
                            return Response({
                                "response": False,
                                "message": f"Component '{component_code}' is statutory and cannot be updated.",
                                "data": None
                            }, status=status.HTTP_400_BAD_REQUEST)
                        
                        # Get or create component (non-statutory deduction)
                        component = self._get_or_create_component(
                            organization,
                            component_code,
                            deduction_data.get('label', component_code.replace('_', ' ').title()),
                            SalaryComponent.DEDUCTION  # Must be deduction type
                        )
                        
                        # Validate component is not statutory
                        if component.statutory_type:
                            return Response({
                                "response": False,
                                "message": f"Component '{component.name}' is statutory and cannot be updated.",
                                "data": None
                            }, status=status.HTTP_400_BAD_REQUEST)
                        
                        # Get or create structure item
                        structure_item, created = SalaryStructureItem.objects.get_or_create(
                            structure=structure,
                            component=component,
                            defaults={
                                'calculation_type': deduction_data['calculation_type'],
                                'value': deduction_data['value'],
                                'order': deduction_order
                            }
                        )
                        
                        if not created:
                            # Update existing item
                            structure_item.calculation_type = deduction_data['calculation_type']
                            structure_item.value = deduction_data['value']
                            structure_item.order = deduction_order
                            structure_item.save()
                        
                        updated_deduction_codes.add(component_code)
                        deduction_order += 1
                    
                    # Delete non-statutory deduction items that are not in the update
                    SalaryStructureItem.objects.filter(
                        structure=structure,
                        component__component_type=SalaryComponent.DEDUCTION,
                        component__statutory_type__isnull=True  # Only non-statutory
                    ).exclude(component__code__in=updated_deduction_codes).delete()
                
                # Return updated structure
                return self.get(request, org_id)
                
        except BaseUserModel.DoesNotExist:
            return Response({
                "response": False,
                "message": "Organization not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except SalaryStructure.DoesNotExist:
            return Response({
                "response": False,
                "message": "Salary structure not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "response": False,
                "message": f"Error updating structure: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def delete(self, request, org_id):
        """
        DELETE - Delete Salary Structure
        Safe delete: checks if structure is in use
        """
        try:
            organization = BaseUserModel.objects.get(id=org_id, role='organization')
            
            structure_id = request.query_params.get('structure_id')
            if not structure_id:
                return Response({
                    "response": False,
                    "message": "structure_id is required",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            structure = SalaryStructure.objects.get(
                id=structure_id,
                organization=organization
            )
            
            # Check if structure is being used by any EmployeePayrollConfig
            configs_using_structure = EmployeePayrollConfig.objects.filter(
                salary_structure=structure,
                is_active=True
            ).select_related('employee')
            
            if configs_using_structure.exists():
                # Get details of configs using this structure - Optimized batch query
                configs_list = list(configs_using_structure[:10])  # Limit to first 10 for message
                employee_ids = [config.employee_id for config in configs_list]
                
                # Batch fetch employee profiles - O(1) query instead of N queries
                employee_profiles = {
                    profile['user_id']: profile
                    for profile in UserProfile.objects.filter(user_id__in=employee_ids).values(
                        'user_id', 'user_name', 'custom_employee_id'
                    )
                }
                
                config_details = []
                for config in configs_list:
                    emp_id = config.employee_id
                    profile = employee_profiles.get(emp_id)
                    if profile:
                        employee_name = profile.get('user_name') or profile.get('custom_employee_id') or str(emp_id)
                    else:
                        employee_name = str(emp_id)
                    
                    config_details.append(
                        f"{employee_name} (Month: {config.effective_month}/{config.effective_year})"
                    )
                
                total_count = configs_using_structure.count()
                if total_count > 10:
                    config_details.append(f"... and {total_count - 10} more")
                
                return Response({
                    "response": False,
                    "message": f"Cannot delete salary structure '{structure.name}' because it is currently assigned to {total_count} employee payroll configuration(s). Please remove or update these configurations first.",
                    "data": {
                        "structure_id": structure.id,
                        "structure_name": structure.name,
                        "configs_count": total_count,
                        "configs_details": config_details
                    }
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get count of items before deletion for response message
            structure_items_count = SalaryStructureItem.objects.filter(structure=structure).count()
            structure_name = structure.name
            
            # Delete structure - CASCADE will automatically delete all related SalaryStructureItems
            structure.delete()  # This will cascade delete all related items
            
            # Note: SalaryStructureItem has on_delete=CASCADE, so items are automatically deleted
            
            return Response({
                "response": True,
                "message": f"Salary structure '{structure_name}' and all {structure_items_count} related components deleted successfully",
                "data": None
            }, status=status.HTTP_200_OK)
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "response": False,
                "message": "Organization not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except SalaryStructure.DoesNotExist:
            return Response({
                "response": False,
                "message": "Salary structure not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "response": False,
                "message": f"Error deleting structure: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== EMPLOYEE PAYROLL CONFIG VIEWS ====================

class EmployeePayrollConfigAPIView(APIView):
    """
    Unified API View for Employee Payroll Configuration
    Handles GET (list/detail), POST, PUT, DELETE
    URL patterns:
    - prefix/<admin_id>            → GET (list), POST
    - prefix/<admin_id>/<id>       → GET (detail), PUT, DELETE
    """
    
    permission_classes = [IsAuthenticated]
    
    def get_organization_from_admin(self, admin_id):
        """Get organization_id from admin_id - O(1) with select_related"""
        try:
            admin_profile = AdminProfile.objects.select_related('organization').get(user_id=admin_id)
            return admin_profile.organization
        except AdminProfile.DoesNotExist:
            return None
    
    def get(self, request, site_id, pk=None, employee_id=None):
        """GET - List all configs or get single config
        URL patterns:
        - prefix/<site_id>/?effective_month=X&effective_year=Y → List with filters
        - prefix/<site_id>/<int:pk>/ → Get by config ID
        - prefix/<site_id>/employee/<uuid:employee_id>/?effective_month=X&effective_year=Y → Get by employee ID
        """
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "response": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "response": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "response": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "response": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Common: Parse month/year params
            effective_month = request.query_params.get('effective_month')
            effective_year = request.query_params.get('effective_year')
            month = None
            year = None
            if effective_month:
                try:
                    month = int(effective_month)
                    if not (1 <= month <= 12):
                        month = None
                except ValueError:
                    pass
            if effective_year:
                try:
                    year = int(effective_year)
                except ValueError:
                    pass
            if not month or not year:
                from django.utils import timezone
                now = timezone.now()
                if not month:
                    month = now.month
                if not year:
                    year = now.year
            
            # Common: Transform employee data to response
            def transform_emp_data(emp_data):
                return {
                    "id": emp_data['payroll_config']['config_id'],
                    "employee_id": emp_data['employee_id'],
                    "salary_structure": emp_data['payroll_config']['salary_structure_id'],
                    "effective_month": emp_data['payroll_config']['effective_month'],
                    "effective_year": emp_data['payroll_config']['effective_year'],
                    "gross_salary": emp_data['payroll_config']['gross_salary'],
                    "gross_salary_monthly": emp_data['payroll_config'].get('gross_salary_monthly', emp_data['payroll_config']['gross_salary']),
                    "earnings": emp_data['calculated_breakdown']['earnings'],
                    "deductions": emp_data['calculated_breakdown']['deductions'],
                    "total_earnings": emp_data['calculated_breakdown']['total_earnings'],
                    "total_deductions": emp_data['calculated_breakdown']['total_deductions'],
                    "net_pay": emp_data['calculated_breakdown']['net_pay']
                }
            
            # Common: Transform config to response
            def transform_config(config, breakdown):
                # Calculate monthly gross salary (gross_salary is stored as annual in DB)
                from decimal import Decimal
                gross_salary_monthly = config.gross_salary / Decimal('12') if config.gross_salary else Decimal('0.00')
                return {
                    "id": config.id,
                    "employee_id": str(config.employee.id),
                    "salary_structure": config.salary_structure.id,
                    "effective_month": config.effective_month,
                    "effective_year": config.effective_year,
                    "gross_salary": str(config.gross_salary),
                    "gross_salary_monthly": str(round(gross_salary_monthly, 2)),
                    "earnings": breakdown['earnings'],
                    "deductions": breakdown['deductions'],
                    "total_earnings": str(breakdown['total_earnings']),
                    "total_deductions": str(breakdown['total_deductions']),
                    "net_pay": str(breakdown['net_pay'])
                }
            
            if pk:
                # Single O(1) query using index payconfig_id_adm_idx
                config = EmployeePayrollConfig.objects.filter(
                    id=pk, admin_id=admin_id, is_active=True
                ).select_related('employee', 'salary_structure').only(
                    'id', 'admin_id', 'site_id', 'employee_id', 'salary_structure_id',
                    'gross_salary', 'effective_month', 'effective_year', 'is_active'
                ).first()
                
                if not config:
                    return Response({
                        "status": False,
                        "message": "Employee payroll configuration not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
                
                # O(1) site check
                if config.site_id != site_id:
                    return Response({
                        "status": False,
                        "message": "Employee payroll configuration not found for this site",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
                
                employee_details_result = get_all_employee_payroll_details(
                    admin_id=admin_id,
                    month=config.effective_month,
                    year=config.effective_year,
                    employee_id=str(config.employee.id)
                )
                
                if 'error' in employee_details_result:
                    return Response({
                        "status": False,
                        "message": employee_details_result['error'],
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                employees = employee_details_result.get('employees', [])
                emp_data = None
                for emp in employees:
                    if emp.get('payroll_config', {}).get('config_id') == pk:
                        emp_data = emp
                        break
                
                if not emp_data or not emp_data.get('payroll_config'):
                    return Response({
                        "status": False,
                        "message": "Employee payroll configuration not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
                
                return Response({
                    "status": True,
                    "message": "Employee payroll configuration fetched successfully",
                    "data": transform_emp_data(emp_data)
                }, status=status.HTTP_200_OK)
            
            elif employee_id:
                # Get config by employee_id
                if effective_month and effective_year:
                    # Single config request
                    employee_details_result = get_all_employee_payroll_details(
                        admin_id=admin_id,
                        month=month,
                        year=year,
                        employee_id=employee_id
                    )
                    
                    if 'error' in employee_details_result:
                        return Response({
                            "status": False,
                            "message": employee_details_result['error'],
                            "data": None
                        }, status=status.HTTP_400_BAD_REQUEST)
                    
                    employees = employee_details_result.get('employees', [])
                    if not employees or not employees[0].get('payroll_config'):
                        return Response({
                            "status": False,
                            "message": "Employee payroll configuration not found for the specified month and year",
                            "data": None
                        }, status=status.HTTP_200_OK)
                    
                    return Response({
                        "status": True,
                        "message": "Employee payroll configuration fetched successfully",
                        "data": transform_emp_data(employees[0])
                    }, status=status.HTTP_200_OK)
                else:
                    # Return list of configs for this employee (multiple months/years)
                    try:
                        employee = BaseUserModel.objects.get(id=employee_id, role='user')
                    except BaseUserModel.DoesNotExist:
                        return Response({
                            "status": False,
                            "message": "Employee not found",
                            "data": None
                        }, status=status.HTTP_404_NOT_FOUND)
                    
                    query = EmployeePayrollConfig.objects.filter(
                        admin_id=admin_id,
                        employee=employee,
                        is_active=True
                    ).select_related('employee', 'salary_structure')
                    
                    if effective_month:
                        try:
                            month_val = int(effective_month)
                            if 1 <= month_val <= 12:
                                query = query.filter(effective_month=month_val)
                        except ValueError:
                            pass
                    
                    if effective_year:
                        try:
                            year_val = int(effective_year)
                            query = query.filter(effective_year=year_val)
                        except ValueError:
                            pass
                    
                    configs_list = list(query.order_by('-effective_year', '-effective_month'))
                    
                    if not configs_list:
                        return Response({
                            "status": True,
                            "message": "Employee payroll configurations fetched successfully",
                            "data": []
                        }, status=status.HTTP_200_OK)
                    
                    structure_items_map, payroll_settings_map, employee_states_map, pt_rules_cache = prefetch_payroll_data(configs_list)
                    response_data = []
                    for config in configs_list:
                        breakdown = calculate_payroll_breakdown_optimized(
                            config,
                            structure_items_map.get(config.salary_structure_id, []),
                            payroll_settings_map.get(config.salary_structure.organization_id),
                            employee_states_map.get(config.employee_id),
                            pt_rules_cache
                        )
                        response_data.append(transform_config(config, breakdown))
                    
                    return Response({
                        "status": True,
                        "message": "Employee payroll configurations fetched successfully",
                        "data": response_data
                    }, status=status.HTTP_200_OK)
            else:
                # Get all configs for employees under this admin
                employee_details_result = get_all_employee_payroll_details(
                    admin_id=admin_id,
                    month=month,
                    year=year,
                    employee_id=None
                )
                
                if 'error' in employee_details_result:
                    return Response({
                        "status": False,
                        "message": employee_details_result['error'],
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                response_data = [
                    transform_emp_data(emp_data)
                    for emp_data in employee_details_result['employees']
                    if emp_data.get('payroll_config') and emp_data.get('calculated_breakdown')
                ]
                
                return Response({
                    "status": True,
                    "message": "Employee payroll configurations fetched successfully",
                    "data": response_data
                }, status=status.HTTP_200_OK)
                
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error fetching employee payroll configuration: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @transaction.atomic
    def post(self, request, site_id):
        """POST - Create new EmployeePayrollConfig"""
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "status": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validate site belongs to admin
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get organization from admin
            organization = self.get_organization_from_admin(admin_id)
            if not organization:
                return Response({
                    "status": False,
                    "message": "Organization not found for admin",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Validate input
            serializer = EmployeePayrollConfigCreateUpdateSerializer(data=request.data)
            if not serializer.is_valid():
                return Response({
                    "status": False,
                    "message": "Validation failed",
                    "data": serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            data = serializer.validated_data
            
            # Verify employee exists and belongs to this admin
            try:
                employee = BaseUserModel.objects.get(id=data['employee_id'], role='user')
                from utils.Employee.assignment_utils import verify_employee_under_admin
                admin_obj = BaseUserModel.objects.get(id=admin_id, role='admin')
                if not verify_employee_under_admin(employee, admin_obj, raise_exception=False):
                    return Response({
                        "status": False,
                        "message": "Employee does not belong to this admin",
                        "data": None
                    }, status=status.HTTP_403_FORBIDDEN)
            except BaseUserModel.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Employee not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            except UserProfile.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Employee profile not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Verify salary structure exists and belongs to organization
            try:
                salary_structure = SalaryStructure.objects.get(
                    id=data['salary_structure_id'],
                    organization=organization
                )
            except SalaryStructure.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Salary structure not found or does not belong to organization",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Check for duplicate (employee, effective_month, effective_year)
            existing_config = EmployeePayrollConfig.objects.filter(
                employee=employee,
                effective_month=data['effective_month'],
                effective_year=data['effective_year'],
                is_active=True
            ).first()
            
            if existing_config:
                return Response({
                    "status": False,
                    "message": f"Configuration already exists for employee for {data['effective_month']}/{data['effective_year']}",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Create config
            config = EmployeePayrollConfig.objects.create(
                admin=admin,
                employee=employee,
                salary_structure=salary_structure,
                gross_salary=data['gross_salary'],
                effective_month=data['effective_month'],
                effective_year=data['effective_year'],
                pf_applicable=data.get('pf_applicable'),
                esi_applicable=data.get('esi_applicable'),
                pt_applicable=data.get('pt_applicable'),
                gratuity_applicable=data.get('gratuity_applicable')
            )
            
            # Calculate breakdown using prefetch_payroll_data utility
            configs_list = [config]
            structure_items_map, payroll_settings_map, employee_states_map, pt_rules_cache = prefetch_payroll_data(configs_list)
            
            breakdown = calculate_payroll_breakdown_optimized(
                config,
                structure_items_map.get(config.salary_structure_id, []),
                payroll_settings_map.get(config.salary_structure.organization_id),
                employee_states_map.get(config.employee_id),
                pt_rules_cache
            )
            
            # Calculate monthly gross salary (gross_salary is stored as annual in DB)
            from decimal import Decimal
            gross_salary_monthly = config.gross_salary / Decimal('12') if config.gross_salary else Decimal('0.00')
            response_data = {
                "id": config.id,
                "employee_id": str(config.employee.id),
                "salary_structure": config.salary_structure.id,
                "effective_month": config.effective_month,
                "effective_year": config.effective_year,
                "gross_salary": str(config.gross_salary),
                "gross_salary_monthly": str(round(gross_salary_monthly, 2)),
                "earnings": breakdown['earnings'],
                "deductions": breakdown['deductions'],
                "total_earnings": str(breakdown['total_earnings']),
                "total_deductions": str(breakdown['total_deductions']),
                "net_pay": str(breakdown['net_pay'])
            }
            
            return Response({
                "status": True,
                "message": "Employee payroll configuration created successfully",
                "data": response_data
            }, status=status.HTTP_201_CREATED)
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error creating employee payroll configuration: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @transaction.atomic
    def put(self, request, site_id, pk=None):
        """PUT - Update EmployeePayrollConfig"""
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "status": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validate site belongs to admin
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get config
            try:
                config = EmployeePayrollConfig.objects.select_related(
                    'employee', 'salary_structure'
                ).get(id=pk, admin=admin, is_active=True)
                
                # Filter by site if provided
                if site_id and config.site_id != site_id:
                    return Response({
                        "status": False,
                        "message": "Employee payroll configuration not found for this site",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            except EmployeePayrollConfig.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Employee payroll configuration not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Validate input
            serializer = EmployeePayrollConfigCreateUpdateSerializer(data=request.data)
            if not serializer.is_valid():
                return Response({
                    "status": False,
                    "message": "Validation failed",
                    "data": serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            data = serializer.validated_data
            
            # Verify employee exists and belongs to this admin (if changed)
            if data['employee_id'] != config.employee.id:
                try:
                    employee = BaseUserModel.objects.get(id=data['employee_id'], role='user')
                    from utils.Employee.assignment_utils import verify_employee_under_admin
                    admin_obj = BaseUserModel.objects.get(id=admin_id, role='admin')
                    if not verify_employee_under_admin(employee, admin_obj, raise_exception=False):
                        return Response({
                            "status": False,
                            "message": "Employee does not belong to this admin",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except (BaseUserModel.DoesNotExist, UserProfile.DoesNotExist):
                    return Response({
                        "status": False,
                        "message": "Employee not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            
            # Get organization from admin
            organization = self.get_organization_from_admin(admin_id)
            if not organization:
                return Response({
                    "status": False,
                    "message": "Organization not found for admin",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Verify salary structure exists and belongs to organization (if changed)
            if data['salary_structure_id'] != config.salary_structure.id:
                try:
                    salary_structure = SalaryStructure.objects.get(
                        id=data['salary_structure_id'],
                        organization=organization
                    )
                except SalaryStructure.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Salary structure not found or does not belong to organization",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                salary_structure = config.salary_structure
            
            # Check for duplicate (employee, effective_month, effective_year) if changed
            if (data['employee_id'] != config.employee.id or
                data['effective_month'] != config.effective_month or
                data['effective_year'] != config.effective_year):
                
                existing_config = EmployeePayrollConfig.objects.filter(
                    employee_id=data['employee_id'],
                    effective_month=data['effective_month'],
                    effective_year=data['effective_year'],
                    is_active=True
                ).exclude(id=pk).first()
                
                if existing_config:
                    return Response({
                        "status": False,
                        "message": f"Configuration already exists for employee for {data['effective_month']}/{data['effective_year']}",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Update config
            config.employee_id = data['employee_id']
            config.salary_structure = salary_structure
            config.gross_salary = data['gross_salary']
            config.effective_month = data['effective_month']
            config.effective_year = data['effective_year']
            config.pf_applicable = data.get('pf_applicable')
            config.esi_applicable = data.get('esi_applicable')
            config.pt_applicable = data.get('pt_applicable')
            config.gratuity_applicable = data.get('gratuity_applicable')
            config.save()
            
            # Calculate breakdown using prefetch_payroll_data utility
            configs_list = [config]
            structure_items_map, payroll_settings_map, employee_states_map, pt_rules_cache = prefetch_payroll_data(configs_list)
            
            breakdown = calculate_payroll_breakdown_optimized(
                config,
                structure_items_map.get(config.salary_structure_id, []),
                payroll_settings_map.get(config.salary_structure.organization_id),
                employee_states_map.get(config.employee_id),
                pt_rules_cache
            )
            
            # Calculate monthly gross salary (gross_salary is stored as annual in DB)
            from decimal import Decimal
            gross_salary_monthly = config.gross_salary / Decimal('12') if config.gross_salary else Decimal('0.00')
            response_data = {
                "id": config.id,
                "employee_id": str(config.employee.id),
                "salary_structure": config.salary_structure.id,
                "effective_month": config.effective_month,
                "effective_year": config.effective_year,
                "gross_salary": str(config.gross_salary),
                "gross_salary_monthly": str(round(gross_salary_monthly, 2)),
                "earnings": breakdown['earnings'],
                "deductions": breakdown['deductions'],
                "total_earnings": str(breakdown['total_earnings']),
                "total_deductions": str(breakdown['total_deductions']),
                "net_pay": str(breakdown['net_pay'])
            }
            
            return Response({
                "status": True,
                "message": "Employee payroll configuration updated successfully",
                "data": response_data
            }, status=status.HTTP_200_OK)
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error updating employee payroll configuration: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def delete(self, request, site_id, pk):
        """DELETE - Soft delete EmployeePayrollConfig"""
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "status": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Single O(1) query using index payconfig_id_adm_idx
            config = EmployeePayrollConfig.objects.filter(
                id=pk, 
                admin_id=admin.id, 
                is_active=True
            ).only('id', 'site_id', 'is_active').first()
            
            if not config:
                return Response({
                    "status": False,
                    "message": "Employee payroll configuration not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) site check
            if config.site_id != site_id:
                return Response({
                    "status": False,
                    "message": "Employee payroll configuration not found for this site",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) optimized update - only update is_active field using index
            EmployeePayrollConfig.objects.filter(id=pk, admin_id=admin.id).update(is_active=False)
            
            return Response({
                "status": True,
                "message": "Employee payroll configuration deleted successfully",
                "data": None
            }, status=status.HTTP_200_OK)
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error deleting employee payroll configuration: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== EMPLOYEE BANK INFO VIEWS ====================

class EmployeeBankInfoAPIView(APIView):
    """
    Unified CRUD API View for Employee Bank Information
    URL patterns:
    - prefix/<admin_id>              → GET (list), POST
    - prefix/<admin_id>/<int:pk>     → GET (detail), PUT, DELETE
    """
    
    permission_classes = [IsAuthenticated]
    
    def get(self, request, site_id, pk=None):
        """
        GET - List all bank info for employees under admin or get single bank info - O(1) queries
        Query parameters:
        - search: Search in employee_name, custom_employee_id, bank_name, account_number, ifsc_code, pan_card_number, aadhar_card_number
        - page: Page number (default: 1)
        - page_size: Items per page (default: 10, max: 100)
        - is_active: Filter by active status (true/false)
        - is_primary: Filter by primary status (true/false)
        """
        try:
            admin, site, error_response = get_admin_and_site_for_payroll(request, site_id)
            if error_response:
                return error_response
            
            admin_id = admin.id
            
            if pk:
                # Single O(1) query using index bankinfo_id_emp_idx with select_related
                bank_info = EmployeeBankInfo.objects.select_related(
                    'employee', 'employee__own_user_profile', 'employee__own_user_profile__admin'
                ).filter(
                    id=pk, 
                    employee__own_user_profile__admin_id=admin_id
                ).only(
                    'id', 'employee_id', 'pan_card_number', 'pan_card_name', 'aadhar_card_number',
                    'aadhar_card_name', 'bank_name', 'account_number', 'account_holder_name',
                    'account_type', 'ifsc_code', 'bank_address', 'branch_name', 'city', 'state',
                    'pincode', 'is_primary', 'is_active', 'created_at', 'updated_at',
                    'employee__own_user_profile__site_id'
                ).first()
                
                if not bank_info:
                    return Response({
                        "status": False,
                        "message": "Bank information not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
                
                # O(1) site check - access via select_related
                if bank_info.employee.own_user_profile.site_id != site_id:
                    return Response({
                        "status": False,
                        "message": "Bank information not found for this site",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
                
                serializer = EmployeeBankInfoSerializer(bank_info)
                return Response({
                    "status": True,
                    "message": "Bank information fetched successfully",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)
            else:
                # Get all bank info for employees under this admin
                # Base queryset - employees under this admin
                employees_under_admin = BaseUserModel.objects.filter(
                    role='user',
                    own_user_profile__admin=admin
                ).values_list('id', flat=True)
                
                queryset = EmployeeBankInfo.objects.filter(
                    employee_id__in=employees_under_admin
                ).select_related('employee', 'employee__own_user_profile')
                
                # Filter by site
                queryset = filter_queryset_by_site(queryset, site_id, 'site')
                queryset = queryset.order_by('-created_at')
                
                # Filter by is_active
                is_active_param = request.query_params.get('is_active')
                if is_active_param is not None:
                    is_active = is_active_param.lower() == 'true'
                    queryset = queryset.filter(is_active=is_active)
                
                # Filter by is_primary
                is_primary_param = request.query_params.get('is_primary')
                if is_primary_param is not None:
                    is_primary = is_primary_param.lower() == 'true'
                    queryset = queryset.filter(is_primary=is_primary)
                
                # Search filter
                search_query = request.query_params.get('search', '').strip()
                if search_query:
                    from django.db.models import Q
                    queryset = queryset.filter(
                        Q(employee__own_user_profile__user_name__icontains=search_query) |
                        Q(employee__own_user_profile__custom_employee_id__icontains=search_query) |
                        Q(bank_name__icontains=search_query) |
                        Q(account_number__icontains=search_query) |
                        Q(ifsc_code__icontains=search_query) |
                        Q(pan_card_number__icontains=search_query) |
                        Q(aadhar_card_number__icontains=search_query) |
                        Q(account_holder_name__icontains=search_query) |
                        Q(branch_name__icontains=search_query) |
                        Q(city__icontains=search_query) |
                        Q(state__icontains=search_query)
                    )
                
                # Pagination
                page = request.query_params.get('page', '1')
                page_size = request.query_params.get('page_size', '10')
                
                try:
                    page = int(page)
                    page_size = int(page_size)
                    if page < 1:
                        page = 1
                    if page_size < 1:
                        page_size = 10
                    if page_size > 100:
                        page_size = 100
                except ValueError:
                    page = 1
                    page_size = 10
                
                # Calculate pagination
                total_count = queryset.count()
                total_pages = (total_count + page_size - 1) // page_size if total_count > 0 else 0
                
                # Slice queryset
                start_index = (page - 1) * page_size
                end_index = start_index + page_size
                paginated_queryset = queryset[start_index:end_index]
                
                # Serialize
                serializer = EmployeeBankInfoListSerializer(paginated_queryset, many=True)
                
                return Response({
                    "status": True,
                    "message": "Bank information list fetched successfully",
                    "data": {
                        "results": serializer.data,
                        "pagination": {
                            "current_page": page,
                            "page_size": page_size,
                            "total_count": total_count,
                            "total_pages": total_pages,
                            "has_next": page < total_pages,
                            "has_previous": page > 1
                        }
                    }
                }, status=status.HTTP_200_OK)
                
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error fetching bank information: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request, site_id):
        """
        POST - Create new bank information for an employee
        Request body should include employee_id and all bank info fields
        """
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "status": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validate site belongs to admin
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            serializer = EmployeeBankInfoCreateSerializer(data=request.data)
            if serializer.is_valid():
                # Verify employee belongs to this admin
                employee_id = serializer.validated_data.get('employee_id')
                try:
                    employee = BaseUserModel.objects.select_related('own_user_profile').get(
                        id=employee_id,
                        role='user',
                        own_user_profile__admin=admin
                    )
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Employee not found or does not belong to this admin",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
                
                # Check if bank info already exists for this employee
                if EmployeeBankInfo.objects.filter(employee_id=employee_id).exists():
                    return Response({
                        "status": False,
                        "message": "Bank information already exists for this employee. Use PUT to update.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Create bank info
                bank_info = serializer.save()
                
                # Return created bank info
                response_serializer = EmployeeBankInfoSerializer(bank_info)
                return Response({
                    "status": True,
                    "message": "Bank information created successfully",
                    "data": response_serializer.data
                }, status=status.HTTP_201_CREATED)
            else:
                return Response({
                    "status": False,
                    "message": "Validation error",
                    "data": serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error creating bank information: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def put(self, request, site_id, pk=None):
        """
        PUT - Update existing bank information
        Requires pk in URL
        """
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "status": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validate site belongs to admin
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            if not pk:
                return Response({
                    "status": False,
                    "message": "Bank info ID (pk) is required for update",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get bank info
            try:
                bank_info = EmployeeBankInfo.objects.select_related(
                    'employee', 'employee__own_user_profile'
                ).get(id=pk, employee__own_user_profile__admin=admin)
            except EmployeeBankInfo.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Bank information not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Update
            serializer = EmployeeBankInfoUpdateSerializer(bank_info, data=request.data, partial=False)
            if serializer.is_valid():
                updated_bank_info = serializer.save()
                
                # Return updated bank info
                response_serializer = EmployeeBankInfoSerializer(updated_bank_info)
                return Response({
                    "status": True,
                    "message": "Bank information updated successfully",
                    "data": response_serializer.data
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    "status": False,
                    "message": "Validation error",
                    "data": serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error updating bank information: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def delete(self, request, site_id, pk=None):
        """
        DELETE - Delete bank information
        Requires pk in URL
        """
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "status": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validate site belongs to admin
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            if not pk:
                return Response({
                    "status": False,
                    "message": "Bank info ID (pk) is required for deletion",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get bank info
            try:
                bank_info = EmployeeBankInfo.objects.select_related(
                    'employee', 'employee__own_user_profile'
                ).get(id=pk, employee__own_user_profile__admin=admin)
            except EmployeeBankInfo.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Bank information not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Delete
            bank_info.delete()
            
            return Response({
                "status": True,
                "message": "Bank information deleted successfully",
                "data": None
            }, status=status.HTTP_200_OK)
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error deleting bank information: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== EMPLOYEE ADVANCE VIEWS ====================

class EmployeeAdvanceAPIView(APIView):
    """
    API View for Employee Advance (GET and POST only)
    URL patterns:
    - prefix/<admin_id>              → GET (list), POST
    - prefix/<admin_id>/<int:pk>     → GET (detail)
    """
    
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def get(self, request, site_id, pk=None):
        """
        GET - List all advances for employees under admin or get single advance - O(1) queries
        Query parameters:
        - search: Search in employee_name, custom_employee_id
        - status: Filter by status (active, partially_paid, settled, cancelled)
        - employee_id: Filter by specific employee
        - page: Page number (default: 1)
        - page_size: Items per page (default: 10, max: 100)
        """
        try:
            admin, site, error_response = get_admin_and_site_for_payroll(request, site_id)
            if error_response:
                return error_response
            
            admin_id = admin.id
            
            if pk:
                # Single O(1) query using index advance_id_adm_idx
                advance = EmployeeAdvance.objects.filter(
                    id=pk, 
                    admin_id=admin_id
                ).select_related(
                    'employee', 'employee__own_user_profile', 'created_by', 'admin'
                ).only(
                    'id', 'admin_id', 'site_id', 'employee_id', 'created_by_id',
                    'advance_amount', 'request_date', 'purpose', 'status',
                    'paid_amount', 'remaining_amount', 'is_settled', 'settlement_date',
                    'notes', 'attachment', 'created_at', 'updated_at'
                ).first()
                
                if not advance:
                    return Response({
                        "status": False,
                        "message": "Advance not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
                
                # O(1) site check
                if advance.site_id != site_id:
                    return Response({
                        "status": False,
                        "message": "Advance not found for this site",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
                
                serializer = EmployeeAdvanceSerializer(advance)
                return Response({
                    "status": True,
                    "message": "Advance fetched successfully",
                    "data": serializer.data
                }, status=status.HTTP_200_OK)
            else:
                # List query with index optimization - uses advance_adm_status_date_idx
                queryset = EmployeeAdvance.objects.filter(
                    admin_id=admin_id
                ).select_related('employee', 'employee__own_user_profile', 'created_by', 'admin').only(
                    'id', 'admin_id', 'site_id', 'employee_id', 'created_by_id',
                    'advance_amount', 'request_date', 'purpose', 'status',
                    'paid_amount', 'remaining_amount', 'is_settled', 'settlement_date',
                    'notes', 'attachment', 'created_at', 'updated_at'
                )
                
                # Filter by site - O(1) with index advance_site_adm_st_dt_idx
                queryset = filter_queryset_by_site(queryset, site_id, 'site')
                queryset = queryset.order_by('-request_date', '-created_at')
                
                # Filter by employee_id
                employee_id = request.query_params.get('employee_id')
                if employee_id:
                    try:
                        queryset = queryset.filter(employee_id=employee_id)
                    except ValueError:
                        pass
                
                # Filter by status
                status_filter = request.query_params.get('status')
                if status_filter:
                    queryset = queryset.filter(status=status_filter)
                
                # Search filter
                search_query = request.query_params.get('search', '').strip()
                if search_query:
                    from django.db.models import Q
                    queryset = queryset.filter(
                        Q(employee__own_user_profile__user_name__icontains=search_query) |
                        Q(employee__own_user_profile__custom_employee_id__icontains=search_query)
                    )
                
                # Pagination
                page = request.query_params.get('page', '1')
                page_size = request.query_params.get('page_size', '10')
                
                try:
                    page = int(page)
                    page_size = int(page_size)
                    if page < 1:
                        page = 1
                    if page_size < 1:
                        page_size = 10
                    if page_size > 100:
                        page_size = 100
                except ValueError:
                    page = 1
                    page_size = 10
                
                # Calculate pagination
                total_count = queryset.count()
                total_pages = (total_count + page_size - 1) // page_size if total_count > 0 else 0
                
                # Slice queryset
                start_index = (page - 1) * page_size
                end_index = start_index + page_size
                paginated_queryset = queryset[start_index:end_index]
                
                # Serialize
                serializer = EmployeeAdvanceListSerializer(paginated_queryset, many=True)
                
                return Response({
                    "status": True,
                    "message": "Advances fetched successfully",
                    "data": {
                        "results": serializer.data,
                        "pagination": {
                            "current_page": page,
                            "page_size": page_size,
                            "total_count": total_count,
                            "total_pages": total_pages,
                            "has_next": page < total_pages,
                            "has_previous": page > 1
                        }
                    }
                }, status=status.HTTP_200_OK)
                
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error fetching advances: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request, site_id):
        """
        POST - Create new advance for an employee
        Admin or organization can create advance
        """
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "status": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validate site belongs to admin
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            admin = BaseUserModel.objects.get(id=admin_id, role='admin')
            
            # Get created_by from request user
            created_by = request.user
            if created_by.role not in ['admin', 'organization']:
                return Response({
                    "status": False,
                    "message": "Only admin or organization can create advances",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            serializer = EmployeeAdvanceCreateSerializer(
                data=request.data,
                context={'admin': admin, 'created_by': created_by}
            )
            
            if serializer.is_valid():
                # Verify employee belongs to this admin
                employee_id = serializer.validated_data.get('employee_id')
                try:
                    employee = BaseUserModel.objects.select_related('own_user_profile').get(
                        id=employee_id,
                        role='user',
                        own_user_profile__admin=admin
                    )
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Employee not found or does not belong to this admin",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
                
                # Create advance
                advance = serializer.save()
                
                # Return created advance
                response_serializer = EmployeeAdvanceSerializer(advance)
                return Response({
                    "status": True,
                    "message": "Advance created successfully",
                    "data": response_serializer.data
                }, status=status.HTTP_201_CREATED)
            else:
                return Response({
                    "status": False,
                    "message": "Validation error",
                    "data": serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error creating advance: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== EMPLOYEE CUSTOM MONTHLY EARNINGS EXCEL ====================
class EmployeeEarningsExcelAPIView(APIView):
    """
    API View for Employee Earnings Excel Operations
    GET - Download Excel template
    POST - Upload Excel file and store data
    """
    
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def get(self, request, site_id):
        """
        GET - Download Excel template for earnings upload
        Query parameters: month (1-12), year (required)
        """
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "status": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validate site belongs to admin
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get month and year from query parameters
            month = request.query_params.get('month')
            year = request.query_params.get('year')
            
            if not month or not year:
                return Response({
                    "status": False,
                    "message": "Month and Year are required as query parameters",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                month = int(month)
                year = int(year)
            except ValueError:
                return Response({
                    "status": False,
                    "message": "Month and Year must be valid integers",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate month
            if month < 1 or month > 12:
                return Response({
                    "status": False,
                    "message": "Month must be between 1 and 12",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Earnings Template"
            
            # Header styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers (removed Month and Year columns)
            headers = [
                "Employee ID",
                "Overtime Pay",
                "Incentives",
                "Impact Award",
                "Bonus",
                "Expenses",
                "Leave Encashment",
                "Adjustments",
                "Arrears",
                "Performance Allowance",
                "Other Allowances",
                "Notes"
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            
            # Add info row about month and year
            info_cell = ws.cell(row=2, column=1, value=f"Month: {month}, Year: {year}")
            info_cell.font = Font(bold=True, italic=True, color="366092")
            info_cell.border = border
            
            # Add sample row
            sample_row = [
                "EMP001",
                0.00,  # Overtime Pay
                0.00,  # Incentives
                0.00,  # Impact Award
                0.00,  # Bonus
                0.00,  # Expenses
                0.00,  # Leave Encashment
                0.00,  # Adjustments
                0.00,  # Arrears
                0.00,  # Performance Allowance
                0.00,  # Other Allowances
                "Sample notes"  # Notes
            ]
            
            for col, value in enumerate(sample_row, 1):
                cell = ws.cell(row=3, column=col, value=value)
                cell.border = border
                if col in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11]:  # Decimal columns
                    cell.number_format = '0.00'
            
            # Auto width
            for col in ws.columns:
                max_len = 0
                letter = col[0].column_letter
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[letter].width = min(max_len + 2, 30)
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Create response
            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="earnings_upload_template.xlsx"'
            
            return response
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error generating template: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request, site_id):
        """
        POST - Upload Excel file and store earnings data
        Query parameters: month (1-12), year (required)
        """
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "status": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validate site belongs to admin
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get month and year from query parameters
            month = request.query_params.get('month')
            year = request.query_params.get('year')
            
            if not month or not year:
                return Response({
                    "status": False,
                    "message": "Month and Year are required as query parameters",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                month = int(month)
                year = int(year)
            except ValueError:
                return Response({
                    "status": False,
                    "message": "Month and Year must be valid integers",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate month
            if month < 1 or month > 12:
                return Response({
                    "status": False,
                    "message": "Month must be between 1 and 12",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check if file is provided
            if 'file' not in request.FILES:
                return Response({
                    "status": False,
                    "message": "Excel file is required",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            excel_file = request.FILES['file']
            
            # Validate file extension
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                return Response({
                    "status": False,
                    "message": "Invalid file format. Please upload an Excel file (.xlsx or .xls)",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Read Excel file
            try:
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
            except Exception as e:
                return Response({
                    "status": False,
                    "message": f"Error reading Excel file: {str(e)}",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Process rows (skip header row and info row if present)
            success_count = 0
            error_count = 0
            errors = []
            start_row = 2  # Start from row 2 (skip header)
            
            # Check if row 2 is info row (contains "Month:" or "Year:")
            try:
                first_data_row = ws[2][0].value if ws[2][0].value else ""
                if "Month:" in str(first_data_row) or "Year:" in str(first_data_row):
                    start_row = 3  # Skip info row too
            except:
                pass
            
            with transaction.atomic():
                for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
                    # Skip empty rows
                    if not row or not row[0]:
                        continue
                    
                    try:
                        # Extract data (month and year now come from URL)
                        custom_employee_id = str(row[0]).strip() if row[0] else None
                        
                        if not custom_employee_id:
                            error_count += 1
                            errors.append(f"Row {row_idx}: Missing required field (Employee ID)")
                            continue
                        
                        # Get employee by custom_employee_id
                        try:
                            employee_qs = BaseUserModel.objects.select_related('own_user_profile').filter(
                                role='user',
                                own_user_profile__custom_employee_id=custom_employee_id,
                                own_user_profile__admin=admin
                            )
                            # Filter by site
                            employee_qs = employee_qs.filter(own_user_profile__site_id=site_id)
                            employee = employee_qs.get()
                        except BaseUserModel.DoesNotExist:
                            error_count += 1
                            errors.append(f"Row {row_idx}: Employee not found with Custom Employee ID: {custom_employee_id}")
                            continue
                        except BaseUserModel.MultipleObjectsReturned:
                            error_count += 1
                            errors.append(f"Row {row_idx}: Multiple employees found with Custom Employee ID: {custom_employee_id}")
                            continue
                        
                        # Extract earnings values (columns shifted by 2 since month and year removed)
                        overtime_pay = Decimal(str(row[1])) if len(row) > 1 and row[1] is not None else Decimal('0.00')
                        incentives = Decimal(str(row[2])) if len(row) > 2 and row[2] is not None else Decimal('0.00')
                        impact_award = Decimal(str(row[3])) if len(row) > 3 and row[3] is not None else Decimal('0.00')
                        bonus = Decimal(str(row[4])) if len(row) > 4 and row[4] is not None else Decimal('0.00')
                        expenses = Decimal(str(row[5])) if len(row) > 5 and row[5] is not None else Decimal('0.00')
                        leave_encashment = Decimal(str(row[6])) if len(row) > 6 and row[6] is not None else Decimal('0.00')
                        adjustments = Decimal(str(row[7])) if len(row) > 7 and row[7] is not None else Decimal('0.00')
                        arrears = Decimal(str(row[8])) if len(row) > 8 and row[8] is not None else Decimal('0.00')
                        performance_allowance = Decimal(str(row[9])) if len(row) > 9 and row[9] is not None else Decimal('0.00')
                        other_allowances = Decimal(str(row[10])) if len(row) > 10 and row[10] is not None else Decimal('0.00')
                        notes = str(row[11]).strip() if len(row) > 11 and row[11] else None
                        
                        # Validate decimal values
                        for val in [overtime_pay, incentives, impact_award, bonus, expenses, 
                                   leave_encashment, adjustments, arrears, performance_allowance, other_allowances]:
                            if val < Decimal('0.00'):
                                raise ValueError("Earnings values cannot be negative")
                        
                        # Create or update earnings record
                        earnings, created = EmployeeCustomMonthlyEarning.objects.update_or_create(
                            employee=employee,
                            admin=admin,
                            month=month,
                            year=year,
                            defaults={
                                'overtime_pay': overtime_pay,
                                'incentives': incentives,
                                'impact_award': impact_award,
                                'bonus': bonus,
                                'expenses': expenses,
                                'leave_encashment': leave_encashment,
                                'adjustments': adjustments,
                                'arrears': arrears,
                                'performance_allowance': performance_allowance,
                                'other_allowances': other_allowances,
                                'notes': notes,
                                'site': site,
                            }
                        )
                        
                        success_count += 1
                        
                    except ValueError as e:
                        error_count += 1
                        errors.append(f"Row {row_idx}: {str(e)}")
                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {row_idx}: {str(e)}")
            
            # Prepare response
            response_data = {
                "status": True,
                "message": f"Excel file processed. Success: {success_count}, Errors: {error_count}",
                "data": {
                    "success_count": success_count,
                    "error_count": error_count,
                    "errors": errors[:10] if errors else []  # Limit to first 10 errors
                }
            }
            
            if error_count > 0:
                response_data["message"] += f". First {min(10, len(errors))} errors shown."
            
            return Response(response_data, status=status.HTTP_200_OK)
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error processing Excel file: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== EMPLOYEE CUSTOM MONTHLY DEDUCTIONS EXCEL ====================
class EmployeeDeductionsExcelAPIView(APIView):
    """
    API View for Employee Deductions Excel Operations
    GET - Download Excel template
    POST - Upload Excel file and store data
    """
    
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def get(self, request, site_id):
        """
        GET - Download Excel template for deductions upload
        Query parameters: month (1-12), year (required)
        """
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "status": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validate site belongs to admin
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get month and year from query parameters
            month = request.query_params.get('month')
            year = request.query_params.get('year')
            
            if not month or not year:
                return Response({
                    "status": False,
                    "message": "Month and Year are required as query parameters",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                month = int(month)
                year = int(year)
            except ValueError:
                return Response({
                    "status": False,
                    "message": "Month and Year must be valid integers",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate month
            if month < 1 or month > 12:
                return Response({
                    "status": False,
                    "message": "Month must be between 1 and 12",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Deductions Template"
            
            # Header styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = [
                "Employee ID",
                "Income Tax (TDS)",
                "Advance Deduction",
                "Labour Welfare Fund (LWF)",
                "Uniform Charges",
                "Canteen/Food Charges",
                "Late Mark Fine",
                "Penalty",
                "Employee Welfare Fund",
                "Other Deductions",
                "Notes"
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            
            # Add info row about month and year
            info_cell = ws.cell(row=2, column=1, value=f"Month: {month}, Year: {year}")
            info_cell.font = Font(bold=True, italic=True, color="366092")
            info_cell.border = border
            
            # Add sample row
            sample_row = [
                "EMP001",
                0.00,  # Income Tax (TDS)
                0.00,  # Advance Deduction
                0.00,  # Labour Welfare Fund (LWF)
                0.00,  # Uniform Charges
                0.00,  # Canteen/Food Charges
                0.00,  # Late Mark Fine
                0.00,  # Penalty
                0.00,  # Employee Welfare Fund
                0.00,  # Other Deductions
                "Sample notes"  # Notes
            ]
            
            for col, value in enumerate(sample_row, 1):
                cell = ws.cell(row=3, column=col, value=value)
                cell.border = border
                if col in [2, 3, 4, 5, 6, 7, 8, 9, 10]:  # Decimal columns
                    cell.number_format = '0.00'
            
            # Auto width
            for col in ws.columns:
                max_len = 0
                letter = col[0].column_letter
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[letter].width = min(max_len + 2, 30)
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Create response
            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="deductions_upload_template.xlsx"'
            
            return response
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error generating template: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request, site_id):
        """
        POST - Upload Excel file and store deductions data
        Query parameters: month (1-12), year (required)
        """
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "status": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validate site belongs to admin
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get month and year from query parameters
            month = request.query_params.get('month')
            year = request.query_params.get('year')
            
            if not month or not year:
                return Response({
                    "status": False,
                    "message": "Month and Year are required as query parameters",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                month = int(month)
                year = int(year)
            except ValueError:
                return Response({
                    "status": False,
                    "message": "Month and Year must be valid integers",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate month
            if month < 1 or month > 12:
                return Response({
                    "status": False,
                    "message": "Month must be between 1 and 12",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check if file is provided
            if 'file' not in request.FILES:
                return Response({
                    "status": False,
                    "message": "Excel file is required",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            excel_file = request.FILES['file']
            
            # Validate file extension
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                return Response({
                    "status": False,
                    "message": "Invalid file format. Please upload an Excel file (.xlsx or .xls)",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Read Excel file
            try:
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
            except Exception as e:
                return Response({
                    "status": False,
                    "message": f"Error reading Excel file: {str(e)}",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Process rows (skip header row and info row if present)
            success_count = 0
            error_count = 0
            errors = []
            start_row = 2  # Start from row 2 (skip header)
            
            # Check if row 2 is info row (contains "Month:" or "Year:")
            try:
                first_data_row = ws[2][0].value if ws[2][0].value else ""
                if "Month:" in str(first_data_row) or "Year:" in str(first_data_row):
                    start_row = 3  # Skip info row too
            except:
                pass
            
            with transaction.atomic():
                for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
                    # Skip empty rows
                    if not row or not row[0]:
                        continue
                    
                    try:
                        # Extract data (month and year now come from URL)
                        custom_employee_id = str(row[0]).strip() if row[0] else None
                        
                        if not custom_employee_id:
                            error_count += 1
                            errors.append(f"Row {row_idx}: Missing required field (Employee ID)")
                            continue
                        
                        # Get employee by custom_employee_id
                        try:
                            employee_qs = BaseUserModel.objects.select_related('own_user_profile').filter(
                                role='user',
                                own_user_profile__custom_employee_id=custom_employee_id,
                                own_user_profile__admin=admin
                            )
                            # Filter by site
                            employee_qs = employee_qs.filter(own_user_profile__site_id=site_id)
                            employee = employee_qs.get()
                        except BaseUserModel.DoesNotExist:
                            error_count += 1
                            errors.append(f"Row {row_idx}: Employee not found with Custom Employee ID: {custom_employee_id}")
                            continue
                        except BaseUserModel.MultipleObjectsReturned:
                            error_count += 1
                            errors.append(f"Row {row_idx}: Multiple employees found with Custom Employee ID: {custom_employee_id}")
                            continue
                        
                        # Extract deductions values
                        income_tax = Decimal(str(row[1])) if len(row) > 1 and row[1] is not None else Decimal('0.00')
                        advance = Decimal(str(row[2])) if len(row) > 2 and row[2] is not None else Decimal('0.00')
                        lwf = Decimal(str(row[3])) if len(row) > 3 and row[3] is not None else Decimal('0.00')
                        uniform = Decimal(str(row[4])) if len(row) > 4 and row[4] is not None else Decimal('0.00')
                        canteen_food = Decimal(str(row[5])) if len(row) > 5 and row[5] is not None else Decimal('0.00')
                        late_mark_fine = Decimal(str(row[6])) if len(row) > 6 and row[6] is not None else Decimal('0.00')
                        penalty = Decimal(str(row[7])) if len(row) > 7 and row[7] is not None else Decimal('0.00')
                        employee_welfare_fund = Decimal(str(row[8])) if len(row) > 8 and row[8] is not None else Decimal('0.00')
                        other_deductions = Decimal(str(row[9])) if len(row) > 9 and row[9] is not None else Decimal('0.00')
                        notes = str(row[10]).strip() if len(row) > 10 and row[10] else None
                        
                        # Validate decimal values
                        for val in [income_tax, advance, lwf, uniform, canteen_food, 
                                   late_mark_fine, penalty, employee_welfare_fund, other_deductions]:
                            if val < Decimal('0.00'):
                                raise ValueError("Deductions values cannot be negative")
                        
                        # Create or update deductions record
                        deductions, created = EmployeeCustomMonthlyDeduction.objects.update_or_create(
                            employee=employee,
                            admin=admin,
                            month=month,
                            year=year,
                            defaults={
                                'income_tax': income_tax,
                                'advance': advance,
                                'lwf': lwf,
                                'uniform': uniform,
                                'canteen_food': canteen_food,
                                'late_mark_fine': late_mark_fine,
                                'penalty': penalty,
                                'employee_welfare_fund': employee_welfare_fund,
                                'other_deductions': other_deductions,
                                'notes': notes,
                                'site': site,
                            }
                        )
                        
                        success_count += 1
                        
                    except ValueError as e:
                        error_count += 1
                        errors.append(f"Row {row_idx}: {str(e)}")
                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {row_idx}: {str(e)}")
            
            # Prepare response
            response_data = {
                "status": True,
                "message": f"Excel file processed. Success: {success_count}, Errors: {error_count}",
                "data": {
                    "success_count": success_count,
                    "error_count": error_count,
                    "errors": errors[:10] if errors else []  # Limit to first 10 errors
                }
            }
            
            if error_count > 0:
                response_data["message"] += f". First {min(10, len(errors))} errors shown."
            
            return Response(response_data, status=status.HTTP_200_OK)
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error processing Excel file: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== DEMO ATTENDANCE SHEET DOWNLOAD ====================
class DemoAttendanceSheetDownloadAPIView(APIView):
    """
    API View for Downloading Demo Attendance Sheet
    GET - Download Excel template with all employees under admin
    """
    
    permission_classes = [IsAuthenticated]
    
    def get(self, request, site_id):
        """
        GET - Download demo attendance sheet with all employees
        Automatically includes: Employee ID, Employee Name, Mobile number
        Payable days column is left empty for user input
        """
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "status": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validate site belongs to admin
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get all employees under this admin
            employees = UserProfile.objects.filter(
                admin=admin,
                user__is_active=True
            ).select_related('user')
            
            # Filter by site
            employees = employees.filter(site_id=site_id)
            
            employees = employees.order_by('custom_employee_id', 'user_name')
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance Sheet"
            
            # Header styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = [
                "Employee ID",
                "Employee Name",
                "Mobile number",
                "Payable days"
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            
            # Write employee data
            for row_idx, employee in enumerate(employees, start=2):
                ws.cell(row=row_idx, column=1, value=employee.custom_employee_id or "")
                ws.cell(row=row_idx, column=2, value=employee.user_name or "")
                ws.cell(row=row_idx, column=3, value=employee.user.phone_number if employee.user else "")
                ws.cell(row=row_idx, column=4, value="")  # Payable days - empty for user input
                
                # Apply border to all cells
                for col in range(1, 5):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = border
            
            # Auto width for columns
            for col in ws.columns:
                max_len = 0
                letter = col[0].column_letter
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
                    except:
                        pass
                ws.column_dimensions[letter].width = min(max_len + 2, 30)
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Create response
            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="demo_attendance_sheet.xlsx"'
            
            return response
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error generating demo attendance sheet: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== GENERATE PAYROLL FROM ATTENDANCE SHEET ====================
class GeneratePayrollFromAttendanceAPIView(APIView):
    """
    API View for Generating Payroll from Attendance Sheet
    POST - Upload Excel file with attendance data and generate payroll records
    """
    
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def post(self, request, site_id):
        """
        POST - Generate payroll from attendance Excel file (OPTIMIZED for 1000+ employees)
        Query parameters: month (1-12), year (required)
        Excel format: Employee ID, Employee Name, Mobile number, Payable days
        
        Optimizations:
        - Batch fetch all employees, payroll configs, custom earnings/deductions
        - Process all calculations in memory
        - Use bulk_create/bulk_update for database operations
        """
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "status": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validate site belongs to admin
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get month and year from query parameters
            month = request.query_params.get('month')
            year = request.query_params.get('year')
            
            if not month or not year:
                return Response({
                    "status": False,
                    "message": "Month and Year are required as query parameters",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                month = int(month)
                year = int(year)
            except ValueError:
                return Response({
                    "status": False,
                    "message": "Month and Year must be valid integers",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate month
            if month < 1 or month > 12:
                return Response({
                    "status": False,
                    "message": "Month must be between 1 and 12",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate file
            if 'file' not in request.FILES:
                return Response({
                    "status": False,
                    "message": "Excel file is required",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            excel_file = request.FILES['file']
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                return Response({
                    "status": False,
                    "message": "Only Excel files (.xlsx, .xls) are allowed",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Read Excel file - load all data into memory first
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            # Get total days in month
            total_days_in_month = monthrange(year, month)[1]
            
            # Get organization and payroll settings (once)
            try:
                admin_profile = AdminProfile.objects.select_related('organization').get(user=admin)
                organization = admin_profile.organization
                payroll_settings = OrganizationPayrollSettings.objects.get(organization=organization)
            except (AdminProfile.DoesNotExist, OrganizationPayrollSettings.DoesNotExist):
                payroll_settings = None
            
            # Build PT rules cache for batch processing (once)
            pt_rules_cache = {}
            if payroll_settings and payroll_settings.pt_enabled:
                pt_rules = ProfessionalTaxRule.objects.filter(is_active=True).values(
                    'state_name', 'salary_from', 'salary_to', 'tax_amount', 'applicable_month'
                )
                for rule in pt_rules:
                    state = rule['state_name']
                    if state not in pt_rules_cache:
                        pt_rules_cache[state] = []
                    pt_rules_cache[state].append(rule)
            
            # ========== STEP 1: Read all Excel data into memory ==========
            # Use dict to handle duplicates - last occurrence wins
            excel_data_dict = {}
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0]:
                    continue
                
                custom_employee_id = str(row[0]).strip() if row[0] else None
                payable_days = Decimal(str(row[3])) if len(row) > 3 and row[3] is not None else Decimal('0.00')
                
                if not custom_employee_id:
                    continue
                
                # Validate payable days
                if payable_days < Decimal('0.00') or payable_days > Decimal(str(total_days_in_month)):
                    continue
                
                # Store/update - last occurrence wins for duplicates
                excel_data_dict[custom_employee_id] = {
                    'row_idx': row_idx,
                    'custom_employee_id': custom_employee_id,
                    'payable_days': payable_days
                }
            
            excel_data = list(excel_data_dict.values())
            
            if not excel_data:
                return Response({
                    "status": False,
                    "message": "No valid data found in Excel file",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # ========== STEP 2: Batch fetch all employees ==========
            custom_employee_ids = [item['custom_employee_id'] for item in excel_data]
            employees_qs = BaseUserModel.objects.filter(
                                role='user',
                own_user_profile__custom_employee_id__in=custom_employee_ids,
                                own_user_profile__admin=admin
            ).select_related('own_user_profile')
            
            # Filter by site
            employees_qs = employees_qs.filter(own_user_profile__site_id=site_id)
            
            # Create mapping: custom_employee_id -> employee
            employee_map = {}
            employee_ids = []
            for employee in employees_qs:
                custom_id = employee.own_user_profile.custom_employee_id
                if custom_id:
                    employee_map[custom_id] = employee
                    employee_ids.append(employee.id)
            
            # ========== STEP 3: Batch fetch all payroll configs ==========
            # Get all active configs for these employees, effective on or before payroll month/year
            payroll_configs_qs = EmployeePayrollConfig.objects.filter(
                employee_id__in=employee_ids,
                admin=admin,
                is_active=True
            ).filter(
                db_models.Q(effective_year__lt=year) | 
                db_models.Q(effective_year=year, effective_month__lte=month)
            )
            
            # Filter by site
            payroll_configs_qs = payroll_configs_qs.filter(site_id=site_id)
            
            payroll_configs_qs = payroll_configs_qs.select_related('salary_structure').prefetch_related(
                'salary_structure__items__component',
                'salary_structure__items__calculation_base'
            ).order_by('employee_id', '-effective_year', '-effective_month')
            
            # Group configs by employee_id and take the most recent one for each
            payroll_config_map = {}
            structure_ids = set()
            for config in payroll_configs_qs:
                emp_id = config.employee_id
                if emp_id not in payroll_config_map:
                    payroll_config_map[emp_id] = config
                    structure_ids.add(config.salary_structure_id)
            
            # ========== STEP 4: Batch fetch all salary structure items ==========
            structure_items_map = {}
            if structure_ids:
                structure_items_qs = SalaryStructureItem.objects.filter(
                    structure_id__in=structure_ids
                ).select_related('component', 'calculation_base').order_by('order', 'id')
                
                for item in structure_items_qs:
                    structure_id = item.structure_id
                    if structure_id not in structure_items_map:
                        structure_items_map[structure_id] = []
                    structure_items_map[structure_id].append(item)
            
            # ========== STEP 5: Batch fetch all custom earnings and deductions ==========
            custom_earnings_map = {}
            custom_deductions_map = {}
            
            if employee_ids:
                custom_earnings_qs = EmployeeCustomMonthlyEarning.objects.filter(
                    employee_id__in=employee_ids,
                    admin=admin,
                    month=month,
                    year=year
                )
                # Filter by site
                custom_earnings_qs = custom_earnings_qs.filter(site_id=site_id)
                for earning in custom_earnings_qs:
                    custom_earnings_map[earning.employee_id] = earning
                
                custom_deductions_qs = EmployeeCustomMonthlyDeduction.objects.filter(
                    employee_id__in=employee_ids,
                    admin=admin,
                    month=month,
                    year=year
                )
                # Filter by site
                custom_deductions_qs = custom_deductions_qs.filter(site_id=site_id)
                for deduction in custom_deductions_qs:
                    custom_deductions_map[deduction.employee_id] = deduction
            
            # ========== STEP 6: Get existing payroll records for bulk update ==========
            existing_records = {}
            existing_records_qs = GeneratedPayrollRecord.objects.filter(
                employee_id__in=employee_ids,
                admin=admin,
                month=month,
                year=year
            )
            # Filter by site
            existing_records_qs = existing_records_qs.filter(site_id=site_id)
            for record in existing_records_qs:
                existing_records[record.employee_id] = record
            
            # ========== STEP 7: Process all calculations in memory ==========
            success_count = 0
            error_count = 0
            errors = []
            payroll_records_to_create = []
            payroll_records_to_update = []
            
            for excel_item in excel_data:
                row_idx = excel_item['row_idx']
                custom_employee_id = excel_item['custom_employee_id']
                payable_days = excel_item['payable_days']
                
                try:
                    # Get employee
                    employee = employee_map.get(custom_employee_id)
                    if not employee:
                        error_count += 1
                        errors.append(f"Row {row_idx}: Employee not found with Custom Employee ID: {custom_employee_id}")
                        continue
                    
                    # Get employee profile for state
                    employee_profile = employee.own_user_profile
                    employee_state = employee_profile.state if employee_profile else None
                    
                    # Get payroll config
                    payroll_config = payroll_config_map.get(employee.id)
                    if not payroll_config:
                        error_count += 1
                        errors.append(f"Row {row_idx}: Payroll config not found for employee: {custom_employee_id}")
                        continue
                    
                    # Get salary structure items
                    structure_items = structure_items_map.get(payroll_config.salary_structure_id, [])
                    if not structure_items:
                        error_count += 1
                        errors.append(f"Row {row_idx}: No salary structure items found for employee: {custom_employee_id}")
                        continue
                    
                    # Calculate payroll breakdown
                    breakdown = calculate_payroll_breakdown_optimized(
                        config=payroll_config,
                        structure_items=structure_items,
                        payroll_settings=payroll_settings,
                        employee_state=employee_state,
                        pt_rules_cache=pt_rules_cache
                    )
                    
                    # Get custom earnings and deductions
                    custom_earnings = custom_earnings_map.get(employee.id)
                    custom_deductions = custom_deductions_map.get(employee.id)
                    
                    # Calculate pro-rata factor
                    prorata_factor = payable_days / Decimal(str(total_days_in_month))
                    
                    # Create component calculation type map
                    component_calculation_type_map = {}
                    for item in structure_items:
                        component_code = item.component.code
                        component_calculation_type_map[component_code] = item.calculation_type
                    
                    # Adjust earnings and deductions for payable days
                    adjusted_earnings = []
                    adjusted_deductions = []
                    
                    # Process earnings (pro-rata for all standard components)
                    for earning in breakdown['earnings']:
                        component_code = earning.get('component', '')
                        amount = Decimal(str(earning.get('amount', 0)))
                        calc_type = component_calculation_type_map.get(component_code, '')
                        
                        if calc_type == 'auto':
                            adjusted_amount = amount
                        else:
                            adjusted_amount = amount * prorata_factor
                        
                        adjusted_earnings.append({
                            'name': component_code,
                            'amount': float(round(adjusted_amount, 2)),
                            'type': 'standard'
                        })
                    
                    # Add custom earnings
                    if custom_earnings:
                        custom_earnings_list = [
                            {'name': 'Overtime Pay', 'amount': float(custom_earnings.overtime_pay), 'type': 'custom'},
                            {'name': 'Incentives', 'amount': float(custom_earnings.incentives), 'type': 'custom'},
                            {'name': 'Impact Award', 'amount': float(custom_earnings.impact_award), 'type': 'custom'},
                            {'name': 'Bonus', 'amount': float(custom_earnings.bonus), 'type': 'custom'},
                            {'name': 'Expenses', 'amount': float(custom_earnings.expenses), 'type': 'custom'},
                            {'name': 'Leave Encashment', 'amount': float(custom_earnings.leave_encashment), 'type': 'custom'},
                            {'name': 'Adjustments', 'amount': float(custom_earnings.adjustments), 'type': 'custom'},
                            {'name': 'Arrears', 'amount': float(custom_earnings.arrears), 'type': 'custom'},
                            {'name': 'Performance Allowance', 'amount': float(custom_earnings.performance_allowance), 'type': 'custom'},
                            {'name': 'Other Allowances', 'amount': float(custom_earnings.other_allowances), 'type': 'custom'},
                        ]
                        adjusted_earnings.extend([e for e in custom_earnings_list if e['amount'] > 0])
                    
                    # Get gross salary (config stores annual, convert to monthly)
                    gross_salary_annual = payroll_config.gross_salary
                    gross_salary_monthly = gross_salary_annual / Decimal('12')
                    gross_salary_pro_rated = gross_salary_monthly * prorata_factor
                    
                    # Get basic salary and DA from adjusted earnings for PF calculation
                    basic_salary_pro_rated = Decimal('0.00')
                    da_salary_pro_rated = Decimal('0.00')
                    for earning in adjusted_earnings:
                        if is_basic_component(earning['name'], ''):
                            basic_salary_pro_rated = Decimal(str(earning['amount']))
                        elif is_da_component(earning['name'], ''):
                            da_salary_pro_rated = Decimal(str(earning['amount']))
                    
                    # Process deductions
                    for deduction in breakdown['deductions']:
                        component_code = deduction.get('component', '')
                        amount = Decimal(str(deduction.get('amount', 0)))
                        calc_type = component_calculation_type_map.get(component_code, '')
                        
                        if calc_type == 'fixed':
                            adjusted_amount = amount * prorata_factor
                            deduction_type = 'standard'
                        elif calc_type == 'auto':
                            # Statutory deductions: Recalculate based on pro-rated gross salary
                            component_item = None
                            for item in structure_items:
                                if item.component.code == component_code:
                                    component_item = item
                                    break
                            
                            if component_item and component_item.component.statutory_type:
                                statutory_type = component_item.component.statutory_type
                                component_name = component_item.component.name
                                
                                if statutory_type == 'PF':
                                    if is_employee_component(component_code, component_name):
                                        pf_base = basic_salary_pro_rated + da_salary_pro_rated
                                        if pf_base <= 0:
                                            pf_base = gross_salary_pro_rated
                                        adjusted_amount = calculate_pf_employee(
                                            pf_base, payroll_settings, payroll_config.pf_applicable
                                        )
                                    else:
                                        adjusted_amount = Decimal('0.00')
                                elif statutory_type == 'ESI':
                                    if is_employee_component(component_code, component_name):
                                        adjusted_amount = calculate_esi_employee(
                                            gross_salary_pro_rated, payroll_settings, payroll_config.esi_applicable
                                        )
                                    else:
                                        adjusted_amount = Decimal('0.00')
                                elif statutory_type == 'PT':
                                    adjusted_amount = calculate_pt(
                                        gross_salary_pro_rated, employee_state, payroll_config.effective_month,
                                        payroll_settings, payroll_config.pt_applicable, pt_rules_cache
                                    )
                                elif statutory_type == 'GRATUITY':
                                    adjusted_amount = amount
                                else:
                                    adjusted_amount = amount
                            else:
                                adjusted_amount = amount
                            
                            deduction_type = 'statutory'
                        else:
                            adjusted_amount = amount
                            deduction_type = 'standard'
                        
                        adjusted_deductions.append({
                            'name': component_code,
                            'amount': float(round(adjusted_amount, 2)),
                            'type': deduction_type
                        })
                    
                    # Add custom deductions
                    if custom_deductions:
                        custom_deductions_list = [
                            {'name': 'Income Tax (TDS)', 'amount': float(custom_deductions.income_tax), 'type': 'custom'},
                            {'name': 'Advance Deduction', 'amount': float(custom_deductions.advance), 'type': 'custom'},
                            {'name': 'Labour Welfare Fund (LWF)', 'amount': float(custom_deductions.lwf), 'type': 'custom'},
                            {'name': 'Uniform Charges', 'amount': float(custom_deductions.uniform), 'type': 'custom'},
                            {'name': 'Canteen/Food Charges', 'amount': float(custom_deductions.canteen_food), 'type': 'custom'},
                            {'name': 'Late Mark Fine', 'amount': float(custom_deductions.late_mark_fine), 'type': 'custom'},
                            {'name': 'Penalty', 'amount': float(custom_deductions.penalty), 'type': 'custom'},
                            {'name': 'Employee Welfare Fund', 'amount': float(custom_deductions.employee_welfare_fund), 'type': 'custom'},
                            {'name': 'Other Deductions', 'amount': float(custom_deductions.other_deductions), 'type': 'custom'},
                        ]
                        adjusted_deductions.extend([d for d in custom_deductions_list if d['amount'] > 0])
                    
                    # Calculate totals
                    total_earnings = sum(Decimal(str(e['amount'])) for e in adjusted_earnings)
                    total_deductions = sum(Decimal(str(d['amount'])) for d in adjusted_deductions)
                    net_pay = total_earnings - total_deductions
                    
                    # Calculate basic salary (for reference)
                    basic_salary = Decimal('0.00')
                    for earning in adjusted_earnings:
                        if is_basic_component(earning['name'], ''):
                            basic_salary = Decimal(str(earning['amount']))
                            break
                    
                    # Convert breakdown to JSON-serializable format
                    breakdown_serializable = {
                        'earnings': breakdown.get('earnings', []),
                        'deductions': breakdown.get('deductions', []),
                        'total_earnings': float(breakdown.get('total_earnings', Decimal('0.00'))),
                        'total_deductions': float(breakdown.get('total_deductions', Decimal('0.00'))),
                        'net_pay': float(breakdown.get('net_pay', Decimal('0.00')))
                    }
                    
                    # Store calculation breakdown
                    calculation_breakdown = {
                        'prorata_factor': float(prorata_factor),
                        'total_days_in_month': total_days_in_month,
                        'payable_days': float(payable_days),
                        'gross_salary_annual': float(gross_salary_annual),
                        'gross_salary_monthly': float(gross_salary_monthly),
                        'gross_salary_pro_rated': float(gross_salary_pro_rated),
                        'breakdown': breakdown_serializable
                    }
                    
                    # Prepare payroll record data
                    payroll_data = {
                        'payable_days': payable_days,
                        'total_days_in_month': total_days_in_month,
                        'gross_salary': gross_salary_monthly,
                        'basic_salary': basic_salary,
                        'earnings': adjusted_earnings,
                        'deductions': adjusted_deductions,
                        'total_earnings': float(round(total_earnings, 2)),
                        'total_deductions': float(round(total_deductions, 2)),
                        'net_pay': float(round(net_pay, 2)),
                        'calculation_breakdown': calculation_breakdown,
                        'payroll_config': payroll_config,
                        'custom_earnings_record': custom_earnings,
                        'custom_deductions_record': custom_deductions,
                        'site': site,
                    }
                    
                    # Check if record exists for bulk update
                    existing_record = existing_records.get(employee.id)
                    if existing_record:
                        # Update existing record
                        for key, value in payroll_data.items():
                            setattr(existing_record, key, value)
                        payroll_records_to_update.append(existing_record)
                    else:
                        # Create new record
                        payroll_records_to_create.append(
                            GeneratedPayrollRecord(
                                employee=employee,
                                admin=admin,
                                month=month,
                                year=year,
                                **payroll_data
                            )
                        )
                    
                    success_count += 1
                    
                except ValueError as e:
                    error_count += 1
                    errors.append(f"Row {row_idx}: {str(e)}")
                except Exception as e:
                    error_count += 1
                    errors.append(f"Row {row_idx}: {str(e)}")
            
            # ========== STEP 8: Bulk database operations ==========
            with transaction.atomic():
                # Bulk create new records
                if payroll_records_to_create:
                    GeneratedPayrollRecord.objects.bulk_create(
                        payroll_records_to_create,
                        batch_size=500,
                        ignore_conflicts=False
                    )
                
                # Bulk update existing records
                if payroll_records_to_update:
                    GeneratedPayrollRecord.objects.bulk_update(
                        payroll_records_to_update,
                        fields=[
                            'payable_days', 'total_days_in_month', 'gross_salary', 'basic_salary',
                            'earnings', 'deductions', 'total_earnings', 'total_deductions', 'net_pay',
                            'calculation_breakdown', 'payroll_config', 'custom_earnings_record',
                            'custom_deductions_record'
                        ],
                        batch_size=500
                    )
            
            # Prepare response
            response_data = {
                "status": True,
                "message": f"Payroll generated. Success: {success_count}, Errors: {error_count}",
                "data": {
                    "success_count": success_count,
                    "error_count": error_count,
                    "errors": errors[:10] if errors else []  # Limit to first 10 errors
                }
            }
            
            if error_count > 0:
                response_data["message"] += f". First {min(10, len(errors))} errors shown."
            
            return Response(response_data, status=status.HTTP_200_OK)
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error generating payroll: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def get(self, request, site_id):
        """
        GET - Fetch generated payroll records for a specific month/year
        Query parameters: 
        - month (1-12), year (required)
        - download=true or format=excel - Returns Excel file instead of JSON
        """
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "status": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validate site belongs to admin
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get month and year from query parameters
            month = request.query_params.get('month')
            year = request.query_params.get('year')
            
            if not month or not year:
                return Response({
                    "status": False,
                    "message": "Month and Year are required as query parameters",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                month = int(month)
                year = int(year)
            except ValueError:
                return Response({
                    "status": False,
                    "message": "Month and Year must be valid integers",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate month
            if month < 1 or month > 12:
                return Response({
                    "status": False,
                    "message": "Month must be between 1 and 12",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check if Excel download is requested
            download_excel = request.query_params.get('download', '').lower() == 'true' or \
                           request.query_params.get('format', '').lower() == 'excel'
            
            # Get generated payroll records
            payroll_records = GeneratedPayrollRecord.objects.filter(
                admin=admin,
                month=month,
                year=year
            ).select_related(
                'employee',
                'employee__own_user_profile',
                'payroll_config',
                'custom_earnings_record',
                'custom_deductions_record'
            ).order_by('employee__own_user_profile__custom_employee_id')
            
            # If Excel download is requested, return Excel file
            if download_excel:
                return self._generate_payroll_excel(payroll_records, month, year)
            
            # Serialize data - clean and simple format
            records_data = []
            for record in payroll_records:
                # Safely get employee profile
                employee_profile = None
                if record.employee:
                    try:
                        employee_profile = record.employee.own_user_profile
                    except:
                        employee_profile = None
                
                # Get earnings and deductions
                earnings_breakdown = record.earnings if record.earnings else []
                deductions_breakdown = record.deductions if record.deductions else []
                
                # Calculate pro-rated gross if not in breakdown
                calculation_breakdown = record.calculation_breakdown if record.calculation_breakdown else {}
                if 'gross_salary_pro_rated' in calculation_breakdown:
                    gross_salary_pro_rated = calculation_breakdown.get('gross_salary_pro_rated')
                else:
                    if record.total_days_in_month > 0:
                        prorata_factor = float(record.payable_days) / float(record.total_days_in_month)
                        gross_salary_pro_rated = float(record.gross_salary) * prorata_factor
                    else:
                        gross_salary_pro_rated = float(record.gross_salary)
                
                # Clean earnings list
                earnings_list = []
                for earning in earnings_breakdown:
                    amount = earning.get('amount', 0)
                    if isinstance(amount, (int, float, Decimal)):
                        amount = float(amount)
                    earnings_list.append({
                        'name': earning.get('name', ''),
                        'amount': amount,
                        'type': earning.get('type', 'standard')
                    })
                
                # Clean deductions list
                deductions_list = []
                for deduction in deductions_breakdown:
                    amount = deduction.get('amount', 0)
                    if isinstance(amount, (int, float, Decimal)):
                        amount = float(amount)
                    deductions_list.append({
                        'name': deduction.get('name', ''),
                        'amount': amount,
                        'type': deduction.get('type', 'statutory')
                    })
                
                # Simple, clean record structure
                records_data.append({
                    'id': record.id,
                    'employee_id': str(record.employee.id) if record.employee else None,
                    'employee_name': employee_profile.user_name if employee_profile else '',
                    'custom_employee_id': employee_profile.custom_employee_id if employee_profile else '',
                    'designation': employee_profile.designation if employee_profile else '',
                    'department': employee_profile.job_title if employee_profile else '',
                    'payable_days': float(record.payable_days),
                    'total_days_in_month': record.total_days_in_month,
                    'gross_salary': float(record.gross_salary),
                    'gross_salary_pro_rated': float(gross_salary_pro_rated),
                    'basic_salary': float(record.basic_salary),
                    'earnings': earnings_list,
                    'deductions': deductions_list,
                    'total_earnings': float(record.total_earnings),
                    'total_deductions': float(record.total_deductions),
                    'net_pay': float(record.net_pay),
                    'payslip_number': record.payslip_number,
                    'notes': record.notes,
                    'generated_at': record.generated_at.isoformat() if record.generated_at else None,
                    'updated_at': record.updated_at.isoformat() if record.updated_at else None,
                })
            
            return Response({
                "status": True,
                "message": f"Generated payroll records fetched successfully",
                "data": {
                    "records": records_data,
                    "total_records": len(records_data),
                    "month": month,
                    "year": year
                }
            }, status=status.HTTP_200_OK)
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error fetching payroll records: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _generate_payroll_excel(self, payroll_records, month, year):
        """
        Generate Excel file for payroll report
        """
        from datetime import datetime
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                      'July', 'August', 'September', 'October', 'November', 'December']
        month_name = month_names[month] if month <= 12 else f'Month-{month}'
        ws.title = f"Payroll Report {month_name} {year}"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Main headers
        headers = [
            "Employee ID",
            "Employee Name",
            "Designation",
            "Department",
            "Payable Days",
            "Total Days",
            "Gross Salary",
            "Basic Salary",
            "Total Earnings",
            "Total Deductions",
            "Net Pay",
            "Payslip Number"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Write payroll data
        for row_idx, record in enumerate(payroll_records, start=2):
            employee_profile = record.employee.own_user_profile if record.employee else None
            
            # Calculate pro-rated gross
            calculation_breakdown = record.calculation_breakdown if record.calculation_breakdown else {}
            if 'gross_salary_pro_rated' in calculation_breakdown:
                gross_salary_pro_rated = calculation_breakdown.get('gross_salary_pro_rated', 0)
            else:
                if record.total_days_in_month > 0:
                    prorata_factor = float(record.payable_days) / float(record.total_days_in_month)
                    gross_salary_pro_rated = float(record.gross_salary) * prorata_factor
                else:
                    gross_salary_pro_rated = float(record.gross_salary)
            
            # Write main data
            ws.cell(row=row_idx, column=1, value=employee_profile.custom_employee_id if employee_profile else "N/A")
            ws.cell(row=row_idx, column=2, value=employee_profile.user_name if employee_profile else "N/A")
            ws.cell(row=row_idx, column=3, value=employee_profile.designation if employee_profile else "N/A")
            ws.cell(row=row_idx, column=4, value=employee_profile.job_title if employee_profile else "N/A")
            ws.cell(row=row_idx, column=5, value=float(record.payable_days))
            ws.cell(row=row_idx, column=6, value=record.total_days_in_month)
            ws.cell(row=row_idx, column=7, value=float(gross_salary_pro_rated))
            ws.cell(row=row_idx, column=8, value=float(record.basic_salary))
            ws.cell(row=row_idx, column=9, value=float(record.total_earnings))
            ws.cell(row=row_idx, column=10, value=float(record.total_deductions))
            ws.cell(row=row_idx, column=11, value=float(record.net_pay))
            ws.cell(row=row_idx, column=12, value=record.payslip_number or "N/A")
            
            # Apply border to all cells
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = border
                # Format currency columns
                if col in [7, 8, 9, 10, 11]:  # Salary columns
                    cell.number_format = '#,##0.00'
        
        # Auto width for columns
        for col in ws.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[letter].width = min(max_len + 2, 30)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        filename = f"Payroll_Report_{month_name}_{year}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        
        return response


# ==================== PAYSLIP GENERATION FROM PAYROLL RECORD ====================

class GeneratePayslipFromPayrollRecordAPIView(APIView):
    """
    API View for Generating Payslips from GeneratedPayrollRecord
    GET - Generate payslips for all employees of an admin for a specific month/year
    Returns all generated payslip data in PayslipGenerator format
    """
    
    permission_classes = [IsAuthenticated]
    
    def _create_payslip_from_record(self, payroll_record, admin, month_name_str, company_name, company_address, company_logo, site=None):
        """Helper method to create payslip from payroll record"""
        from datetime import date
        
        employee = payroll_record.employee
        employee_profile = employee.own_user_profile if employee else None
        
        # Prepare earnings list
        earnings_list = []
        if payroll_record.earnings:
            for earning in payroll_record.earnings:
                earnings_list.append({
                    'name': earning.get('name', ''),
                    'amount': float(earning.get('amount', 0))
                })
        
        # Prepare deductions list
        deductions_list = []
        if payroll_record.deductions:
            for deduction in payroll_record.deductions:
                deductions_list.append({
                    'name': deduction.get('name', ''),
                    'amount': float(deduction.get('amount', 0))
                })
        
        # Store gross salary in custom_pay_summary_fields for display
        custom_pay_summary = {
            'month_gross': float(payroll_record.gross_salary) if payroll_record.gross_salary else float(payroll_record.total_earnings)
        }
        
        # Create payslip
        payslip = PayslipGenerator.objects.create(
            admin=admin,
            employee=employee,
            site=site,
            month=month_name_str,
            year=payroll_record.year,
            pay_date=date(payroll_record.year, payroll_record.month, 1),
            paid_days=int(payroll_record.payable_days),
            loss_of_pay_days=int(payroll_record.total_days_in_month - payroll_record.payable_days),
            template='classic',  # Default template
            currency='INR',
            company_name=company_name,
            company_address=company_address,
            company_logo=company_logo,
            employee_name=employee_profile.user_name if employee_profile else '',
            employee_code=employee_profile.custom_employee_id if employee_profile else '',
            designation=employee_profile.designation if employee_profile else '',
            department=employee_profile.job_title if employee_profile else '',
            pan_number=employee_profile.pan_number if employee_profile and hasattr(employee_profile, 'pan_number') else '',
            earnings=earnings_list,
            deductions=deductions_list,
            custom_pay_summary_fields=custom_pay_summary,
            total_earnings=payroll_record.total_earnings,
            total_deductions=payroll_record.total_deductions,
            net_pay=payroll_record.net_pay,
            notes=payroll_record.notes if hasattr(payroll_record, 'notes') else ''
        )
        # Payslip number will be auto-generated in save() method
        return payslip
    
    def get(self, request, site_id):
        """
        GET - Generate payslips for all employees from GeneratedPayrollRecord
        Query parameters:
        - month (1-12) - Required
        - year - Required
        Returns all generated payslip data in PayslipGenerator format
        """
        try:
            from datetime import date
            
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "status": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validate site belongs to admin
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get month and year from query parameters
            month = request.query_params.get('month')
            year = request.query_params.get('year')
            
            if not month or not year:
                return Response({
                    "status": False,
                    "message": "Month and Year are required as query parameters",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                month = int(month)
                year = int(year)
            except ValueError:
                return Response({
                    "status": False,
                    "message": "Month and Year must be valid integers",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate month
            if month < 1 or month > 12:
                return Response({
                    "status": False,
                    "message": "Month must be between 1 and 12",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get month name for matching
            month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                          'July', 'August', 'September', 'October', 'November', 'December']
            month_name_str = month_names[month] if month <= 12 else f'Month-{month}'
            
            # Delete all existing payslips for this admin, month, and year first
            # This ensures no UNIQUE constraint issues when regenerating
            PayslipGenerator.objects.filter(
                admin=admin,
                month=month_name_str,
                year=year
            ).delete()
            
            # Get all payroll records for this admin, month, and year
            payroll_records = GeneratedPayrollRecord.objects.filter(
                admin=admin,
                month=month,
                year=year
            ).select_related(
                'employee',
                'employee__own_user_profile',
                'payroll_config',
                'admin'
            )
            
            # Filter by site
            payroll_records = filter_queryset_by_site(payroll_records, site_id, 'site')
            
            if not payroll_records.exists():
                return Response({
                    "status": False,
                    "message": f"No payroll records found for {month_name_str} {year}",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Get organization details once (same for all payslips)
            company_name = None
            company_address = None
            company_logo = None
            try:
                admin_profile = AdminProfile.objects.select_related('organization').get(user=admin)
                organization_user = admin_profile.organization
                
                # Get organization profile for name and address
                try:
                    org_profile = organization_user.own_organization_profile
                    company_name = org_profile.organization_name if org_profile else None
                    # Construct address from state and city
                    if org_profile:
                        address_parts = []
                        if org_profile.city:
                            address_parts.append(org_profile.city)
                        if org_profile.state:
                            address_parts.append(org_profile.state)
                        if address_parts:
                            company_address = ", ".join(address_parts)
                except:
                    pass
                
                # Get organization logo from settings
                try:
                    org_settings = organization_user.own_organization_profile_setting
                    if org_settings and org_settings.organization_logo:
                        company_logo = org_settings.organization_logo
                except:
                    pass
            except Exception as e:
                # If organization details not found, leave as None
                pass
            
            # Generate payslips for all payroll records
            generated_payslips = []
            errors = []
            
            for payroll_record in payroll_records:
                try:
                    payslip = self._create_payslip_from_record(
                        payroll_record, admin, month_name_str,
                        company_name, company_address, company_logo, site
                    )
                    generated_payslips.append(payslip)
                except Exception as e:
                    errors.append({
                        'employee_id': str(payroll_record.employee.id) if payroll_record.employee else None,
                        'error': str(e)
                    })
            
            # Serialize all generated payslips
            serializer = PayslipGeneratorSerializer(generated_payslips, many=True, context={'request': request})
            
            return Response({
                "status": True,
                "message": f"Successfully generated {len(generated_payslips)} payslip(s)" + (f", {len(errors)} failed" if errors else ""),
                "data": {
                    "payslips": serializer.data,
                    "total_generated": len(generated_payslips),
                    "errors": errors if errors else None
                }
            }, status=status.HTTP_200_OK)
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except GeneratedPayrollRecord.DoesNotExist:
            return Response({
                "status": False,
                "message": "Payroll record not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error generating payslip: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== PAYSLIP LIST VIEW ====================

class PayslipListView(APIView):
    """
    API View for Listing Payslips
    GET - Get all payslips for an admin or for a specific employee
    DELETE - Delete all payslips for an admin
    """
    
    permission_classes = [IsAuthenticated]
    
    def get(self, request, site_id):
        """
        GET - Get payslips
        Query parameters:
        - employee_id (optional) - Filter by employee
        - month (optional) - Filter by month
        - year (optional) - Filter by year
        """
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": False,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": False,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "status": False,
                    "message": "Unauthorized access. Only admin and organization roles can access this endpoint",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validate site belongs to admin
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": False,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Get query parameters
            employee_id = request.query_params.get('employee_id')
            month = request.query_params.get('month')
            year = request.query_params.get('year')
            
            # Build query
            payslips_qs = PayslipGenerator.objects.filter(admin=admin)
            
            # Filter by employee if provided
            if employee_id:
                try:
                    employee = BaseUserModel.objects.get(id=employee_id, role='user')
                    payslips_qs = payslips_qs.filter(employee=employee)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": False,
                        "message": "Employee not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            
            # Filter by month if provided
            if month:
                try:
                    month_int = int(month)
                    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                                  'July', 'August', 'September', 'October', 'November', 'December']
                    month_name = month_names[month_int] if 1 <= month_int <= 12 else None
                    if month_name:
                        payslips_qs = payslips_qs.filter(month=month_name)
                except ValueError:
                    pass
            
            # Filter by year if provided
            if year:
                try:
                    year_int = int(year)
                    payslips_qs = payslips_qs.filter(year=year_int)
                except ValueError:
                    pass
            
            # Order by latest first
            payslips_qs = payslips_qs.select_related('employee', 'employee__own_user_profile').order_by('-year', '-created_at')
            
            # Serialize payslips
            serializer = PayslipGeneratorSerializer(payslips_qs, many=True, context={'request': request})
            
            return Response({
                "status": True,
                "message": f"Payslips fetched successfully",
                "data": {
                    "payslips": serializer.data,
                    "total_count": len(serializer.data)
                }
            }, status=status.HTTP_200_OK)
            
        except BaseUserModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                "status": False,
                "message": f"Error fetching payslips: {str(e)}",
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)