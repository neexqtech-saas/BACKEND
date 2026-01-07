"""
Expenditure Views
Optimized for high-traffic, low-cost, future-proof architecture
All queries O(1) or using proper database indexes
"""

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from django.http import HttpResponse
from io import BytesIO
from datetime import date, timedelta, datetime, time
from decimal import Decimal
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from .models import ExpenseCategory, ExpenseProject, Expense
from .serializers import (
    ExpenseCategorySerializer, ExpenseProjectSerializer,
    ExpenseSerializer, ExpenseCreateSerializer
)
from AuthN.models import BaseUserModel, AdminProfile
from SiteManagement.models import Site, EmployeeAdminSiteAssignment
from utils.pagination_utils import CustomPagination
from utils.site_filter_utils import filter_queryset_by_site


def get_admin_and_site_for_expense(request, site_id, user_id=None):
    """
    Optimized admin and site validation - O(1) queries with select_related
    Handles admin, organization, and user roles
    Returns: (admin, site, None) tuple or (None, None, Response with error)
    """
    user = request.user
    
    # Fast path for admin role - O(1) query
    if user.role == 'admin':
        admin_id = user.id
        admin = user
        # Single O(1) query with index on (id, created_by_admin, is_active)
        try:
            site = Site.objects.only('id', 'site_name', 'created_by_admin_id', 'is_active').get(
                id=site_id, 
                created_by_admin_id=admin_id, 
                is_active=True
            )
            return admin, site, None
        except Site.DoesNotExist:
            return None, None, Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Site not found or you don't have permission to access this site",
                "data": []
            }, status=status.HTTP_403_FORBIDDEN)
    
    # Organization role - O(1) queries with select_related
    elif user.role == 'organization':
        admin_id = request.query_params.get('admin_id')
        if not admin_id:
            return None, None, Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                "data": []
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Single O(1) query with select_related to avoid N+1 - uses index on (id, role)
        try:
            admin = BaseUserModel.objects.select_related('own_admin_profile').only(
                'id', 'role', 'email'
            ).get(id=admin_id, role='admin')
        except BaseUserModel.DoesNotExist:
            return None, None, Response({
                "status": status.HTTP_404_NOT_FOUND,
                "message": "Admin not found",
                "data": []
            }, status=status.HTTP_404_NOT_FOUND)
        
        # O(1) query - verify admin belongs to organization using select_related
        admin_profile = AdminProfile.objects.select_related('user', 'organization').only(
            'id', 'user_id', 'organization_id'
        ).filter(
            user_id=admin_id,
            organization_id=user.id
        ).first()
        
        if not admin_profile:
            return None, None, Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Admin does not belong to your organization",
                "data": []
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Single O(1) query with index on (id, created_by_admin, is_active)
        try:
            site = Site.objects.only('id', 'site_name', 'created_by_admin_id', 'is_active').get(
                id=site_id, 
                created_by_admin_id=admin_id, 
                is_active=True
            )
            return admin, site, None
        except Site.DoesNotExist:
            return None, None, Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Site not found or you don't have permission to access this site",
                "data": []
            }, status=status.HTTP_403_FORBIDDEN)
    
    # User role - O(1) queries with select_related
    elif user.role == 'user':
        # Validate user_id matches logged-in user
        if user_id and str(user_id) != str(user.id):
            return None, None, Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "You can only access your own expenses",
                "data": []
            }, status=status.HTTP_403_FORBIDDEN)
        
        # O(1) query - Get admin from employee assignment using select_related
        assignment = EmployeeAdminSiteAssignment.objects.filter(
            employee_id=user.id,
            site_id=site_id,
            is_active=True
        ).select_related('admin', 'site').only(
            'admin_id', 'site_id', 'admin__id', 'site__id'
        ).first()
        
        if not assignment:
            return None, None, Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "You are not assigned to this site",
                "data": []
            }, status=status.HTTP_403_FORBIDDEN)
        
        admin = assignment.admin
        # O(1) query - Validate site exists
        try:
            site = Site.objects.only('id', 'site_name', 'is_active').get(
                id=site_id, 
                is_active=True
            )
            return admin, site, None
        except Site.DoesNotExist:
            return None, None, Response({
                "status": status.HTTP_404_NOT_FOUND,
                "message": "Site not found",
                "data": []
            }, status=status.HTTP_404_NOT_FOUND)
    
    else:
        return None, None, Response({
            "status": status.HTTP_403_FORBIDDEN,
            "message": "Unauthorized access. Only admin, organization, and user roles can access this endpoint",
            "data": []
        }, status=status.HTTP_403_FORBIDDEN)


class ExpenseCategoryAPIView(APIView):
    """Expense Category CRUD - Optimized"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request, site_id, pk=None):
        """Get expense categories - O(1) queries with index optimization"""
        try:
            admin, site, error_response = get_admin_and_site_for_expense(request, site_id)
            if error_response:
                return error_response
            
            if pk:
                # Single O(1) query using index expcat_id_adm_idx (id, admin)
                category = ExpenseCategory.objects.filter(
                    id=pk, 
                    admin_id=admin.id, 
                    is_active=True
                ).only('id', 'admin_id', 'name', 'code', 'description', 'is_active', 'created_at', 'updated_at', 'site_id').first()
                
                if not category:
                    return Response({
                        "status": status.HTTP_404_NOT_FOUND,
                        "message": "Expense category not found",
                        "data": []
                    }, status=status.HTTP_404_NOT_FOUND)
                
                # O(1) site check
                if site_id and category.site_id != site_id:
                    return Response({
                        "status": status.HTTP_404_NOT_FOUND,
                        "message": "Expense category not found for this site"
                    }, status=status.HTTP_404_NOT_FOUND)
                
                serializer = ExpenseCategorySerializer(category)
                return Response({
                    "status": status.HTTP_200_OK,
                    "message": "Expense category fetched successfully",
                    "data": serializer.data
                })
            else:
                # Single O(1) query using index expcat_adm_active_idx (admin, is_active)
                categories = ExpenseCategory.objects.filter(
                    admin_id=admin.id, 
                    is_active=True
                ).only('id', 'admin_id', 'name', 'code', 'description', 'is_active', 'created_at', 'updated_at', 'site_id')
                
                # Filter by site - O(1) with index
                categories = filter_queryset_by_site(categories, site_id, 'site')
                
                serializer = ExpenseCategorySerializer(categories, many=True)
                return Response({
                    "status": status.HTTP_200_OK,
                    "message": "Expense categories fetched successfully",
                    "data": serializer.data
                })
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request, site_id):
        """Create expense category - Optimized"""
        try:
            admin, site, error_response = get_admin_and_site_for_expense(request, site_id)
            if error_response:
                return error_response
            
            data = request.data.copy()
            data['admin'] = str(admin.id)
            if site_id:
                data['site'] = str(site.id)
            
            serializer = ExpenseCategorySerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    "status": status.HTTP_201_CREATED,
                    "message": "Expense category created successfully",
                    "data": serializer.data
                }, status=status.HTTP_201_CREATED)
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "Validation error",
                "data": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def put(self, request, site_id, pk=None):
        """Update expense category - Optimized"""
        try:
            admin, site, error_response = get_admin_and_site_for_expense(request, site_id)
            if error_response:
                return error_response
            
            # Single O(1) query using index
            category = ExpenseCategory.objects.filter(
                id=pk, 
                admin_id=admin.id
            ).only('id', 'site_id').first()
            
            if not category:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Expense category not found",
                    "data": []
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) site check
            if site_id and category.site_id != site_id:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Expense category not found for this site"
                }, status=status.HTTP_404_NOT_FOUND)
            
            serializer = ExpenseCategorySerializer(category, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    "status": status.HTTP_200_OK,
                    "message": "Expense category updated successfully",
                    "data": serializer.data
                })
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "Validation error",
                "data": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def delete(self, request, site_id, pk=None):
        """Delete expense category (soft delete) - Optimized O(1) update"""
        try:
            admin, site, error_response = get_admin_and_site_for_expense(request, site_id)
            if error_response:
                return error_response
            
            # Single O(1) query to check existence
            category = ExpenseCategory.objects.filter(
                id=pk, 
                admin_id=admin.id
            ).only('id', 'site_id', 'is_active').first()
            
            if not category:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Expense category not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) site check
            if site_id and category.site_id != site_id:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Expense category not found for this site"
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) optimized update - only update is_active field using index
            ExpenseCategory.objects.filter(id=pk, admin_id=admin.id).update(is_active=False)
            
            return Response({
                "status": status.HTTP_200_OK,
                "message": "Expense category deleted successfully",
                "data": None
            })
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ExpenseProjectAPIView(APIView):
    """Expense Project CRUD - Optimized"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request, site_id, pk=None):
        """Get expense projects - O(1) queries with index optimization"""
        try:
            admin, site, error_response = get_admin_and_site_for_expense(request, site_id)
            if error_response:
                return error_response
            
            if pk:
                # Single O(1) query using index expproj_id_adm_idx (id, admin)
                project = ExpenseProject.objects.filter(
                    id=pk, 
                    admin_id=admin.id, 
                    is_active=True
                ).only('id', 'admin_id', 'name', 'code', 'description', 'is_active', 'created_at', 'updated_at', 'site_id').first()
                
                if not project:
                    return Response({
                        "status": status.HTTP_404_NOT_FOUND,
                        "message": "Expense project not found",
                        "data": []
                    }, status=status.HTTP_404_NOT_FOUND)
                
                # O(1) site check
                if site_id and project.site_id != site_id:
                    return Response({
                        "status": status.HTTP_404_NOT_FOUND,
                        "message": "Expense project not found for this site"
                    }, status=status.HTTP_404_NOT_FOUND)
                
                serializer = ExpenseProjectSerializer(project)
                return Response({
                    "status": status.HTTP_200_OK,
                    "message": "Expense project fetched successfully",
                    "data": serializer.data
                })
            else:
                # Single O(1) query using index expproj_adm_active_idx (admin, is_active)
                projects = ExpenseProject.objects.filter(
                    admin_id=admin.id, 
                    is_active=True
                ).only('id', 'admin_id', 'name', 'code', 'description', 'is_active', 'created_at', 'updated_at', 'site_id')
                
                # Filter by site - O(1) with index
                projects = filter_queryset_by_site(projects, site_id, 'site')
                
                serializer = ExpenseProjectSerializer(projects, many=True)
                return Response({
                    "status": status.HTTP_200_OK,
                    "message": "Expense projects fetched successfully",
                    "data": serializer.data
                })
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request, site_id):
        """Create expense project - Optimized"""
        try:
            admin, site, error_response = get_admin_and_site_for_expense(request, site_id)
            if error_response:
                return error_response
            
            data = request.data.copy()
            data['admin'] = str(admin.id)
            if site_id:
                data['site'] = str(site.id)
            
            serializer = ExpenseProjectSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    "status": status.HTTP_201_CREATED,
                    "message": "Expense project created successfully",
                    "data": serializer.data
                }, status=status.HTTP_201_CREATED)
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "Validation error",
                "data": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def put(self, request, site_id, pk=None):
        """Update expense project - Optimized"""
        try:
            admin, site, error_response = get_admin_and_site_for_expense(request, site_id)
            if error_response:
                return error_response
            
            # Single O(1) query using index
            project = ExpenseProject.objects.filter(
                id=pk, 
                admin_id=admin.id
            ).only('id', 'site_id').first()
            
            if not project:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Expense project not found",
                    "data": []
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) site check
            if site_id and project.site_id != site_id:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Expense project not found for this site"
                }, status=status.HTTP_404_NOT_FOUND)
            
            serializer = ExpenseProjectSerializer(project, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    "status": status.HTTP_200_OK,
                    "message": "Expense project updated successfully",
                    "data": serializer.data
                })
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "Validation error",
                "data": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def delete(self, request, site_id, pk=None):
        """Delete expense project (soft delete) - Optimized O(1) update"""
        try:
            admin, site, error_response = get_admin_and_site_for_expense(request, site_id)
            if error_response:
                return error_response
            
            # Single O(1) query to check existence
            project = ExpenseProject.objects.filter(
                id=pk, 
                admin_id=admin.id
            ).only('id', 'site_id', 'is_active').first()
            
            if not project:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Expense project not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) site check
            if site_id and project.site_id != site_id:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Expense project not found for this site"
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) optimized update - only update is_active field using index
            ExpenseProject.objects.filter(id=pk, admin_id=admin.id).update(is_active=False)
            
            return Response({
                "status": status.HTTP_200_OK,
                "message": "Expense project deleted successfully",
                "data": None
            })
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ExpenseAPIView(APIView):
    """Expense CRUD - Employee can submit, Admin can view all - Optimized"""
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    
    def get(self, request, site_id, user_id=None):
        """Get expenses - O(1) queries with index optimization"""
        try:
            admin, site, error_response = get_admin_and_site_for_expense(request, site_id, user_id)
            if error_response:
                return error_response
            
            admin_id = admin.id
            
            # Base queryset with index optimization
            if user_id:
                # Uses index expense_emp_status_date_idx (employee, status, expense_date)
                employee = BaseUserModel.objects.filter(id=user_id, role='user').only('id', 'role').first()
                if not employee:
                    return Response({
                        "status": status.HTTP_404_NOT_FOUND,
                        "message": "Employee not found",
                        "data": []
                    }, status=status.HTTP_404_NOT_FOUND)
                expenses = Expense.objects.filter(admin_id=admin_id, employee_id=user_id)
            else:
                # Uses index expense_adm_status_date_idx (admin, status, expense_date)
                expenses = Expense.objects.filter(admin_id=admin_id)
            
            # Filter by site - O(1) with index
            expenses = filter_queryset_by_site(expenses, site_id, 'site')
            
            # Optimized date filtering - use date range for index usage
            date_from_str = request.query_params.get('date_from')
            date_to_str = request.query_params.get('date_to')
            
            if not date_from_str and not date_to_str:
                # Default to last 10 days - uses index efficiently
                today = date.today()
                ten_days_ago = today - timedelta(days=10)
                expenses = expenses.filter(expense_date__gte=ten_days_ago, expense_date__lte=today)
            else:
                if date_from_str:
                    try:
                        expenses = expenses.filter(expense_date__gte=date_from_str)
                    except ValueError:
                        pass
                
                if date_to_str:
                    try:
                        expenses = expenses.filter(expense_date__lte=date_to_str)
                    except ValueError:
                        pass
            
            # Status filter - uses expense_adm_status_date_idx or expense_emp_status_date_idx index
            status_filter = request.query_params.get('status')
            if status_filter:
                expenses = expenses.filter(status=status_filter)
            
            # Category filter - uses expense_adm_cat_status_idx index
            category_id = request.query_params.get('category_id')
            if category_id:
                expenses = expenses.filter(category_id=category_id)
            
            # Search functionality - optimized with proper field selection
            search_query = request.query_params.get('search', '').strip()
            if search_query:
                expenses = expenses.filter(
                    Q(title__icontains=search_query) |
                    Q(description__icontains=search_query) |
                    Q(employee__email__icontains=search_query) |
                    Q(employee__own_user_profile__user_name__icontains=search_query) |
                    Q(employee__own_user_profile__custom_employee_id__icontains=search_query) |
                    Q(category__name__icontains=search_query) |
                    Q(project__name__icontains=search_query)
                )
            
            # Optimize queryset with select_related to avoid N+1
            expenses = expenses.select_related(
                'employee', 'category', 'project', 'approved_by', 'rejected_by',
                'employee__own_user_profile', 'approved_by__own_user_profile', 'rejected_by__own_user_profile'
            ).order_by('-expense_date', '-created_at')
            
            # Check if Excel export is requested
            export_excel = request.query_params.get('export', '').lower() == 'true'
            if export_excel:
                # Limit export to prevent memory issues - max 10K records
                export_limit = 10000
                export_expenses = expenses[:export_limit]
                serializer = ExpenseSerializer(export_expenses, many=True)
                return self.generate_excel_export_optimized(export_expenses)
            
            # Apply range/limit parameter for user_id endpoint (default 50)
            if user_id:
                limit_param = request.query_params.get('limit') or request.query_params.get('range')
                if limit_param:
                    try:
                        limit = int(limit_param)
                        if limit > 0:
                            expenses = expenses[:limit]
                    except ValueError:
                        pass
                
                # If no limit specified, default to 50
                if not limit_param:
                    expenses = expenses[:50]
                
                serializer = ExpenseSerializer(expenses, many=True)
                return Response({
                    "status": status.HTTP_200_OK,
                    "message": "Expenses fetched successfully",
                    "data": serializer.data
                })
            
            # Fetch only required fields for serialization - reduces data transfer
            expenses = expenses.only(
                'id', 'admin_id', 'employee_id', 'category_id', 'project_id', 'site_id',
                'title', 'description', 'expense_date', 'amount', 'currency', 'status',
                'submitted_at', 'approved_at', 'approved_by_id', 'rejected_at', 'rejected_by_id',
                'reimbursement_amount', 'reimbursement_date', 'reimbursement_mode',
                'reimbursement_reference', 'receipts', 'supporting_documents', 'remarks',
                'created_at', 'updated_at', 'created_by_id',
                'employee__email', 'employee__own_user_profile__user_name', 'employee__own_user_profile__custom_employee_id',
                'category__name', 'project__name',
                'approved_by__email', 'approved_by__own_user_profile__user_name',
                'rejected_by__email', 'rejected_by__own_user_profile__user_name'
            )
            
            # Pagination for admin view (all expenses) - single query with LIMIT/OFFSET using index
            paginator = self.pagination_class()
            paginated_qs = paginator.paginate_queryset(expenses, request)
            serializer = ExpenseSerializer(paginated_qs, many=True)
            pagination_data = paginator.get_paginated_response(serializer.data)
            pagination_data["results"] = serializer.data
            pagination_data["message"] = "Expenses fetched successfully"
            
            return Response(pagination_data)
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def generate_excel_export_optimized(self, expenses_queryset):
        """
        Generate Excel export for expenses - Highly Optimized
        Uses .values() to fetch only required fields directly from DB
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expenses"
        
        headers = [
            "Expense ID", "Title", "Description", "Employee Name", "Custom Employee ID", "Employee Email",
            "Category", "Project", "Expense Date", "Amount", "Currency", "Status",
            "Submitted At", "Approved At", "Approved By", "Rejected At", "Rejected By",
            "Reimbursement Amount", "Reimbursement Date", "Reimbursement Mode", "Reimbursement Reference",
            "Remarks", "Created At", "Updated At"
        ]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center")
        
        # Header Row
        for col, head in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=head)
            c.fill = header_fill
            c.font = header_font
            c.alignment = header_alignment
        
        # Helper function to convert values to Excel-compatible format
        def to_excel_value(val):
            if val is None:
                return "N/A"
            if hasattr(val, 'strftime'):
                return val.strftime('%Y-%m-%d %H:%M:%S') if hasattr(val, 'hour') else val.strftime('%Y-%m-%d')
            try:
                if not isinstance(val, (str, int, float, bool)):
                    return str(val)
            except:
                return "N/A"
            return val
        
        # Use .values() to fetch only required fields - avoids model instantiation
        expense_values = expenses_queryset.values(
            'id', 'title', 'description', 'expense_date', 'amount', 'currency', 'status',
            'submitted_at', 'approved_at', 'rejected_at',
            'reimbursement_amount', 'reimbursement_date', 'reimbursement_mode',
            'reimbursement_reference', 'remarks', 'created_at', 'updated_at',
            'employee__email', 'employee__own_user_profile__user_name', 'employee__own_user_profile__custom_employee_id',
            'category__name', 'project__name',
            'approved_by__own_user_profile__user_name', 'rejected_by__own_user_profile__user_name'
        )
        
        # Data Rows - iterate over values directly (O(N) but optimized)
        row_data = []
        for expense in expense_values:
            row = [
                to_excel_value(expense.get("id", "N/A")),
                to_excel_value(expense.get("title", "N/A")),
                to_excel_value(expense.get("description", "N/A")),
                to_excel_value(expense.get("employee__own_user_profile__user_name", "N/A")),
                to_excel_value(expense.get("employee__own_user_profile__custom_employee_id", "N/A")),
                to_excel_value(expense.get("employee__email", "N/A")),
                to_excel_value(expense.get("category__name", "N/A")),
                to_excel_value(expense.get("project__name", "N/A")),
                to_excel_value(expense.get("expense_date", "N/A")),
                to_excel_value(expense.get("amount", "N/A")),
                to_excel_value(expense.get("currency", "N/A")),
                to_excel_value(expense.get("status", "N/A")),
                to_excel_value(expense.get("submitted_at", "N/A")),
                to_excel_value(expense.get("approved_at", "N/A")),
                to_excel_value(expense.get("approved_by__own_user_profile__user_name", "N/A")),
                to_excel_value(expense.get("rejected_at", "N/A")),
                to_excel_value(expense.get("rejected_by__own_user_profile__user_name", "N/A")),
                to_excel_value(expense.get("reimbursement_amount", "N/A")),
                to_excel_value(expense.get("reimbursement_date", "N/A")),
                to_excel_value(expense.get("reimbursement_mode", "N/A")),
                to_excel_value(expense.get("reimbursement_reference", "N/A")),
                to_excel_value(expense.get("remarks", "N/A")),
                to_excel_value(expense.get("created_at", "N/A")),
                to_excel_value(expense.get("updated_at", "N/A")),
            ]
            row_data.append(row)
        
        # Batch write to worksheet for better performance
        for i, row in enumerate(row_data, 2):
            for col, val in enumerate(row, 1):
                ws.cell(row=i, column=col).value = val
        
        # Auto width - optimized to only check first 100 rows for performance
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = 0
            letter = col[0].column_letter
            check_rows = min(100, len(col))
            for cell in col[:check_rows]:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[letter].width = min(max_len + 2, 50)  # Cap at 50 chars
        
        # Save in memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="expenses.xlsx"'
        
        return response
    
    @transaction.atomic
    def post(self, request, site_id, user_id=None):
        """Create expense (Employee submits expense request) - Optimized"""
        try:
            admin, site, error_response = get_admin_and_site_for_expense(request, site_id, user_id)
            if error_response:
                return error_response
            
            # Validate user_id is provided
            if not user_id:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "user_id is required in URL",
                    "data": []
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # O(1) query - Get employee
            employee = BaseUserModel.objects.filter(id=user_id, role='user').only('id', 'role').first()
            if not employee:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Employee not found",
                    "data": []
                }, status=status.HTTP_404_NOT_FOUND)
            
            data = request.data.copy()
            data['admin'] = str(admin.id)
            data['employee'] = str(employee.id)
            if site_id:
                data['site'] = str(site.id)
            data['status'] = 'pending'
            data['submitted_at'] = timezone.now()
            data['created_by'] = str(request.user.id)
            
            serializer = ExpenseCreateSerializer(data=data)
            if serializer.is_valid():
                expense = serializer.save()
                return Response({
                    "status": status.HTTP_201_CREATED,
                    "message": "Expense submitted successfully",
                    "data": ExpenseSerializer(expense).data
                }, status=status.HTTP_201_CREATED)
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "Validation error",
                "data": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ExpenseApprovalAPIView(APIView):
    """Admin approves or rejects expense - Optimized"""
    permission_classes = [IsAuthenticated]
    
    @transaction.atomic
    def put(self, request, site_id, expense_id=None, action=None):
        """Approve or reject expense - Optimized O(1) query"""
        try:
            admin, site, error_response = get_admin_and_site_for_expense(request, site_id)
            if error_response:
                return error_response
            
            # Single O(1) query using index expense_id_adm_idx (id, admin)
            expense = Expense.objects.filter(
                id=expense_id, 
                admin_id=admin.id
            ).only('id', 'site_id', 'amount', 'status').first()
            
            if not expense:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Expense not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) site check
            if site_id and expense.site_id != site_id:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Expense not found for this site"
                }, status=status.HTTP_404_NOT_FOUND)
            
            if action == 'approve':
                # Approve expense
                approval_amount = request.data.get('approval_amount')
                description = request.data.get('description', '')
                
                if approval_amount is None:
                    # If no approval_amount provided, use amount
                    approval_amount = float(expense.amount)
                else:
                    approval_amount = float(approval_amount)
                
                # O(1) optimized update
                Expense.objects.filter(id=expense_id).update(
                    status='approved',
                    approved_by_id=request.user.id,
                    approved_at=timezone.now(),
                    reimbursement_amount=Decimal(str(approval_amount)),
                    remarks=description if description else expense.remarks
                )
                
                # Reload expense for response
                expense.refresh_from_db()
                
                return Response({
                    "status": status.HTTP_200_OK,
                    "message": "Expense approved successfully",
                    "data": ExpenseSerializer(expense).data
                })
            
            elif action == 'reject':
                # Reject expense
                description = request.data.get('description', '')
                
                if not description:
                    return Response({
                        "status": status.HTTP_400_BAD_REQUEST,
                        "message": "Rejection reason is required",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # O(1) optimized update
                Expense.objects.filter(id=expense_id).update(
                    status='rejected',
                    rejected_by_id=request.user.id,
                    rejected_at=timezone.now(),
                    rejection_reason=description
                )
                
                # Reload expense for response
                expense.refresh_from_db()
                
                return Response({
                    "status": status.HTTP_200_OK,
                    "message": "Expense rejected successfully",
                    "data": ExpenseSerializer(expense).data
                })
            
            else:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "Invalid action. Use 'approve' or 'reject'",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
