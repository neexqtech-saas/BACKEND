import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from io import BytesIO
from django.http import HttpResponse


class EmployeeExcelExportService:

    @staticmethod
    def generate(active_employees, deactivated_employees, admin_id):
        """
        Generate Excel file with 2 sheets: Active Employees and Deactivated Employees
        """
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create sheets
        ws_active = wb.create_sheet("Active Employees")
        ws_deactivated = wb.create_sheet("Deactivated Employees")
        
        # Headers
        headers = [
            "Employee Name", "Custom Employee ID", "Email", "Username",
            "Phone Number", "Designation", "Job Title", "Date of Joining",
            "Date of Birth", "Gender", "Status"
        ]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Helper function to set headers
        def set_headers(worksheet):
            for col, head in enumerate(headers, 1):
                c = worksheet.cell(row=1, column=col, value=head)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center")
        
        # Set headers for both sheets
        set_headers(ws_active)
        set_headers(ws_deactivated)
        
        # Helper function to convert values to Excel-compatible format
        def to_excel_value(val):
            """Convert value to Excel-compatible format"""
            if val is None:
                return "N/A"
            # Convert datetime objects to string
            if hasattr(val, 'strftime'):
                return val.strftime('%Y-%m-%d')
            # Convert any other non-serializable objects to string
            try:
                if not isinstance(val, (str, int, float, bool)):
                    return str(val)
            except:
                return "N/A"
            return val
        
        # Helper function to add employee data
        def add_employee_row(worksheet, row_num, employee):
            row = [
                to_excel_value(employee.get("user_name", "N/A")),
                to_excel_value(employee.get("custom_employee_id", "N/A")),
                to_excel_value(employee.get("email", "N/A")),
                to_excel_value(employee.get("username", "N/A")),
                to_excel_value(employee.get("phone_number", "N/A")),
                to_excel_value(employee.get("designation", "N/A")),
                to_excel_value(employee.get("job_title", "N/A")),
                to_excel_value(employee.get("date_of_joining", "N/A")),
                to_excel_value(employee.get("date_of_birth", "N/A")),
                to_excel_value(employee.get("gender", "N/A")),
                "Active" if employee.get("is_active", False) else "Inactive"
            ]
            for col, val in enumerate(row, 1):
                worksheet.cell(row=row_num, column=col).value = to_excel_value(val)
        
        # Add active employees
        for i, employee in enumerate(active_employees, 2):
            add_employee_row(ws_active, i, employee)
        
        # Add deactivated employees
        for i, employee in enumerate(deactivated_employees, 2):
            add_employee_row(ws_deactivated, i, employee)
        
        # Auto width for both sheets
        for ws in [ws_active, ws_deactivated]:
            for col in ws.columns:
                max_len = 0
                letter = col[0].column_letter
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[letter].width = max_len + 2
        
        # Save in memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="employees_export_{admin_id}.xlsx"'
        
        return response

