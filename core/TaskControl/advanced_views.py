"""
Advanced Task Management Views
"""

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.db.models import Q, Count, Sum, Avg
from django.utils import timezone
from datetime import datetime, date, timedelta
from decimal import Decimal
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import traceback

from .models import Task, TaskComment
from .serializers import (
    TaskSerializer, TaskCommentSerializer
)
from AuthN.models import BaseUserModel, UserProfile, AdminProfile
from SiteManagement.models import Site
from utils.pagination_utils import CustomPagination
from utils.site_filter_utils import filter_queryset_by_site


class TaskAPIView(APIView):
    """
    Task CRUD Operations
    - Admin can create tasks and assign them to employees
    - Employees can create their own tasks
    - Admin can see all tasks
    - Employees can only see their assigned or self-created tasks
    """
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    
    def get(self, request, site_id, user_id=None, pk=None):
        """Get tasks - filtered by role - Single optimized query"""
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": status.HTTP_400_BAD_REQUEST,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": status.HTTP_403_FORBIDDEN,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": status.HTTP_404_NOT_FOUND,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                # For user role, admin_id is not required
                admin_id = None
                admin = None
            
            # Validate site belongs to admin (if admin_id exists)
            if admin_id:
                try:
                    site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
                except Site.DoesNotExist:
                    return Response({
                        "status": status.HTTP_403_FORBIDDEN,
                        "message": "Site not found or you don't have permission to access this site",
                        "data": None
                    }, status=status.HTTP_403_FORBIDDEN)
            
            user = request.user
            
            if pk:
                # Single query with index optimization (id, admin)
                task = Task.objects.filter(
                    id=pk,
                    admin_id=admin_id
                ).select_related('task_type', 'assigned_to', 'assigned_by').first()
                
                if not task:
                    return Response({
                        "status": status.HTTP_404_NOT_FOUND,
                        "message": "Task not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
                
                # Filter by site if provided
                if site_id and task.site_id != site_id:
                    return Response({
                        "status": status.HTTP_404_NOT_FOUND,
                        "message": "Task not found for this site"
                    }, status=status.HTTP_404_NOT_FOUND)
                
                # Check access permission
                if user.role == 'user' and task.assigned_to_id != user.id:
                    return Response({
                        "status": status.HTTP_403_FORBIDDEN,
                        "message": "You don't have permission to view this task",
                        "data": None
                    }, status=status.HTTP_403_FORBIDDEN)
                
                serializer = TaskSerializer(task)
                return Response({
                    "status": status.HTTP_200_OK,
                    "message": "Task fetched successfully",
                    "data": serializer.data
                })
            else:
                # Base queryset with index optimization
                if user.role == 'admin':
                    queryset = Task.objects.filter(admin_id=admin_id)
                    if user_id:
                        queryset = queryset.filter(assigned_to_id=user_id)
                elif user.role == 'user':
                    # Employee view: show all tasks assigned to them regardless of admin assignment
                    queryset = Task.objects.filter(assigned_to_id=user.id)
                else:
                    queryset = Task.objects.none()
                
                # Filter by site if provided
                queryset = filter_queryset_by_site(queryset, site_id, 'site')
                
                # Apply filters - uses indexes
                status_filter = request.query_params.get('status')
                if status_filter:
                    queryset = queryset.filter(status=status_filter)
                
                priority_filter = request.query_params.get('priority')
                if priority_filter:
                    queryset = queryset.filter(priority=priority_filter)
                
                # Optimized date filtering - use datetime range for index usage
                from_date_str = request.query_params.get('from_date')
                to_date_str = request.query_params.get('to_date')
                
                if not from_date_str and not to_date_str:
                    # Default to last 10 days
                    today_end = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
                    ten_days_ago_start = (timezone.now() - timedelta(days=10)).replace(hour=0, minute=0, second=0, microsecond=0)
                    queryset = queryset.filter(created_at__gte=ten_days_ago_start, created_at__lte=today_end)
                else:
                    if from_date_str:
                        try:
                            from_date_obj = datetime.strptime(from_date_str, '%Y-%m-%d').date()
                            from_date_dt = timezone.make_aware(datetime.combine(from_date_obj, datetime.min.time()))
                            queryset = queryset.filter(created_at__gte=from_date_dt)
                        except ValueError:
                            pass
                    
                    if to_date_str:
                        try:
                            to_date_obj = datetime.strptime(to_date_str, '%Y-%m-%d').date()
                            to_date_dt = timezone.make_aware(datetime.combine(to_date_obj, datetime.max.time()))
                            queryset = queryset.filter(created_at__lte=to_date_dt)
                        except ValueError:
                            pass
                
                # Search functionality
                search_query = request.query_params.get('search', '').strip()
                if search_query:
                    queryset = queryset.filter(
                        Q(title__icontains=search_query) |
                        Q(description__icontains=search_query) |
                        Q(assigned_to__email__icontains=search_query) |
                        Q(assigned_to__own_user_profile__user_name__icontains=search_query) |
                        Q(assigned_to__own_user_profile__custom_employee_id__icontains=search_query) |
                        Q(assigned_by__email__icontains=search_query) |
                        Q(assigned_by__own_user_profile__user_name__icontains=search_query) |
                        Q(assigned_by__own_user_profile__custom_employee_id__icontains=search_query) |
                        Q(task_type__name__icontains=search_query)
                    )
                
                # Optimize queryset with select_related
                queryset = queryset.select_related(
                    'assigned_to', 'assigned_by', 'task_type',
                    'assigned_to__own_user_profile', 'assigned_by__own_user_profile'
                ).order_by('-created_at')
                
                # Fetch only required fields
                queryset = queryset.only(
                    'id', 'admin_id', 'task_type_id', 'title', 'description', 'priority', 'status',
                    'assigned_to_id', 'assigned_by_id', 'start_date', 'due_date', 'start_time',
                    'end_time', 'actual_hours', 'completed_at', 'progress_percentage',
                    'created_at', 'updated_at'
                )
                
                # Check if Excel export is requested
                export_excel = request.query_params.get('export', '').lower() == 'true'
                if export_excel:
                    # Limit export to prevent memory issues - max 10K records
                    export_limit = 10000
                    export_queryset = queryset[:export_limit]
                    serializer = TaskSerializer(export_queryset, many=True)
                    return self.generate_excel_export(serializer.data)
                
                # Pagination
                paginator = self.pagination_class()
                paginated_qs = paginator.paginate_queryset(queryset, request)
                serializer = TaskSerializer(paginated_qs, many=True)
                pagination_data = paginator.get_paginated_response(serializer.data)
                pagination_data["results"] = serializer.data
                pagination_data["message"] = "Tasks fetched successfully"
                
                return Response(pagination_data)
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @transaction.atomic
    def post(self, request, site_id, user_id=None):
        """Create task - Admin can create and assign to employee"""
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": status.HTTP_400_BAD_REQUEST,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": status.HTTP_403_FORBIDDEN,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": status.HTTP_404_NOT_FOUND,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                # For user role, admin_id is not required
                admin_id = None
                admin = None
            
            # Validate site belongs to admin (if admin_id exists)
            if admin_id:
                try:
                    site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
                except Site.DoesNotExist:
                    return Response({
                        "status": status.HTTP_403_FORBIDDEN,
                        "message": "Site not found or you don't have permission to access this site",
                        "data": None
                    }, status=status.HTTP_403_FORBIDDEN)
            else:
                # For user role, just validate site exists
                try:
                    site = Site.objects.get(id=site_id, is_active=True)
                except Site.DoesNotExist:
                    return Response({
                        "status": status.HTTP_404_NOT_FOUND,
                        "message": "Site not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            
            user = request.user
            
            data = request.data.copy()
            if admin_id:
                data['admin'] = str(admin.id)
            data['assigned_by'] = str(user.id)
            # Set site if provided
            if site_id:
                data['site'] = str(site.id)
            
            # If user_id is provided in URL and user is admin, assign to that employee
            if user_id and user.role == 'admin':
                employee = get_object_or_404(BaseUserModel, id=user_id, role='user')
                data['assigned_to'] = str(employee.id)
            elif 'assigned_to' not in data or not data['assigned_to']:
                # If assigned_to is not provided and user is employee, assign to themselves
                if user.role == 'user':
                    data['assigned_to'] = str(user.id)
            
            serializer = TaskSerializer(data=data)
            if serializer.is_valid():
                task = serializer.save()
                
                # Add dependencies if provided
                dependency_ids = request.data.get('dependency_ids', [])
                if dependency_ids and admin_id:
                    dependencies = Task.objects.filter(id__in=dependency_ids, admin=admin)
                    task.dependencies.set(dependencies)
                
                # Schedule task if frequency is not onetime
                if task.schedule_frequency != 'onetime':
                    from .tasks import create_scheduled_task
                    # Schedule first instance immediately if start_date is today or past
                    if task.start_date and task.start_date <= date.today():
                        create_scheduled_task.delay(task.id)
                
                return Response({
                    "status": status.HTTP_201_CREATED,
                    "message": "Task created successfully",
                    "data": TaskSerializer(task).data
                }, status=status.HTTP_201_CREATED)
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "Validation error",
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @transaction.atomic
    def put(self, request, site_id, user_id=None, pk=None):
        """Update task"""
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": status.HTTP_400_BAD_REQUEST,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": status.HTTP_403_FORBIDDEN,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": status.HTTP_404_NOT_FOUND,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                # For user role, admin_id is not required
                admin_id = None
                admin = None
            
            if not pk:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "Task ID is required",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate site belongs to admin (if admin_id exists)
            if admin_id:
                try:
                    site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
                except Site.DoesNotExist:
                    return Response({
                        "status": status.HTTP_403_FORBIDDEN,
                        "message": "Site not found or you don't have permission to access this site",
                        "data": None
                    }, status=status.HTTP_403_FORBIDDEN)
            
            # Get task - filter by admin_id if provided
            if admin_id:
                task = get_object_or_404(Task, id=pk, admin_id=admin_id)
            else:
                task = get_object_or_404(Task, id=pk)
            
            # Filter by site if provided
            if site_id and task.site_id != site_id:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Task not found for this site"
                }, status=status.HTTP_404_NOT_FOUND)
            
            user = request.user
            
            # Check permission - only admin or assigned employee can update
            if user.role == 'user' and task.assigned_to != user:
                return Response({
                    "status": status.HTTP_403_FORBIDDEN,
                    "message": "You don't have permission to update this task",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # If status changed to completed, set completed_at
            if 'status' in request.data and request.data['status'] == 'completed' and task.status != 'completed':
                request.data['completed_at'] = timezone.now()
            
            serializer = TaskSerializer(task, data=request.data, partial=True)
            if serializer.is_valid():
                old_frequency = task.schedule_frequency
                task = serializer.save()
                
                # Update dependencies if provided
                if 'dependency_ids' in request.data and admin_id:
                    dependency_ids = request.data.get('dependency_ids', [])
                    dependencies = Task.objects.filter(id__in=dependency_ids, admin=admin)
                    task.dependencies.set(dependencies)
                
                # Reschedule if frequency changed
                if 'schedule_frequency' in request.data and old_frequency != task.schedule_frequency:
                    if task.schedule_frequency != 'onetime':
                        from .tasks import create_scheduled_task
                        if task.start_date and task.start_date <= date.today():
                            create_scheduled_task.delay(task.id)
                
                return Response({
                    "status": status.HTTP_200_OK,
                    "message": "Task updated successfully",
                    "data": TaskSerializer(task).data
                })
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "Validation error",
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def delete(self, request, site_id, user_id=None, pk=None):
        """Delete task - Only admin can delete"""
        try:
            # Get admin_id based on role
            if request.user.role == 'admin':
                admin_id = request.user.id
                admin = request.user
            elif request.user.role == 'organization':
                # Organization role: get admin_id from query params
                admin_id = request.query_params.get('admin_id')
                if not admin_id:
                    return Response({
                        "status": status.HTTP_400_BAD_REQUEST,
                        "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                        "data": None
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validate admin exists and belongs to organization
                try:
                    admin = BaseUserModel.objects.get(id=admin_id, role='admin')
                    # Verify admin belongs to organization
                    admin_profile = AdminProfile.objects.filter(
                        user=admin,
                        organization=request.user
                    ).first()
                    if not admin_profile:
                        return Response({
                            "status": status.HTTP_403_FORBIDDEN,
                            "message": "Admin does not belong to your organization",
                            "data": None
                        }, status=status.HTTP_403_FORBIDDEN)
                except BaseUserModel.DoesNotExist:
                    return Response({
                        "status": status.HTTP_404_NOT_FOUND,
                        "message": "Admin not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({
                    "status": status.HTTP_403_FORBIDDEN,
                    "message": "Unauthorized access. Only admin and organization roles can delete tasks",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validate site belongs to admin
            try:
                site = Site.objects.get(id=site_id, created_by_admin=admin, is_active=True)
            except Site.DoesNotExist:
                return Response({
                    "status": status.HTTP_403_FORBIDDEN,
                    "message": "Site not found or you don't have permission to access this site",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            if not pk:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "Task ID is required",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            task = get_object_or_404(Task, id=pk, admin_id=admin_id)
            
            # Filter by site if provided
            if site_id and task.site_id != site_id:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Task not found for this site"
                }, status=status.HTTP_404_NOT_FOUND)
            
            user = request.user
            
            # Check permission - only admin can delete
            if user.role != 'admin':
                return Response({
                    "status": status.HTTP_403_FORBIDDEN,
                    "message": "You don't have permission to delete this task",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            task.delete()
            return Response({
                "status": status.HTTP_200_OK,
                "message": "Task deleted successfully",
                "data": None
            })
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def generate_excel_export(self, task_data):
        """Generate Excel export for tasks"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tasks"
        
        headers = [
            "Task ID", "Title", "Description", "Task Type", "Priority", "Status",
            "Assigned To Name", "Custom Employee ID (Assigned To)", "Assigned To Email",
            "Assigned By Name", "Custom Employee ID (Assigned By)", "Assigned By Email",
            "Start Date", "Due Date", "Start Time", "End Time", "Actual Hours",
            "Schedule Frequency", "Week Day", "Month Date", "Schedule End Date",
            "Progress Percentage", "Created At", "Updated At"
        ]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Header Row
        for col, head in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=head)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")
        
        # Helper function to convert values to Excel-compatible format
        def to_excel_value(val):
            if val is None:
                return "N/A"
            if hasattr(val, 'strftime'):
                return val.strftime('%Y-%m-%d %H:%M:%S') if hasattr(val, 'hour') else val.strftime('%Y-%m-%d')
            try:
                if not isinstance(val, (str, int, float, bool)):
                    return str(val)
            except:
                return "N/A"
            return val
        
        # Data Rows
        for i, task in enumerate(task_data, 2):
            row = [
                to_excel_value(task.get("id", "N/A")),
                to_excel_value(task.get("title", "N/A")),
                to_excel_value(task.get("description", "N/A")),
                to_excel_value(task.get("task_type_name", "N/A")),
                to_excel_value(task.get("priority", "N/A")),
                to_excel_value(task.get("status", "N/A")),
                to_excel_value(task.get("assigned_to_name", "N/A")),
                to_excel_value(task.get("assigned_to_custom_employee_id", "N/A")),
                to_excel_value(task.get("assigned_to_email", "N/A")),
                to_excel_value(task.get("assigned_by_name", "N/A")),
                to_excel_value(task.get("assigned_by_custom_employee_id", "N/A")),
                to_excel_value(task.get("assigned_by_email", "N/A")),
                to_excel_value(task.get("start_date", "N/A")),
                to_excel_value(task.get("due_date", "N/A")),
                to_excel_value(task.get("start_time", "N/A")),
                to_excel_value(task.get("end_time", "N/A")),
                to_excel_value(task.get("actual_hours", "N/A")),
                to_excel_value(task.get("schedule_frequency", "N/A")),
                to_excel_value(task.get("week_day", "N/A")),
                to_excel_value(task.get("month_date", "N/A")),
                to_excel_value(task.get("schedule_end_date", "N/A")),
                to_excel_value(task.get("progress_percentage", "N/A")),
                to_excel_value(task.get("created_at", "N/A")),
                to_excel_value(task.get("updated_at", "N/A")),
            ]
            for col, val in enumerate(row, 1):
                ws.cell(row=i, column=col).value = to_excel_value(val)
        
        # Auto width
        for col in ws.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[letter].width = max_len + 2
        
        # Save in memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="tasks.xlsx"'
        
        return response


class TaskCommentAPIView(APIView):
    """Task Comments"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request, task_id):
        """Get task comments"""
        try:
            task = get_object_or_404(Task, id=task_id)
            comments = TaskComment.objects.filter(task=task)
            serializer = TaskCommentSerializer(comments, many=True)
            return Response({
                "status": status.HTTP_200_OK,
                "message": "Comments fetched successfully",
                "data": serializer.data
            })
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request, task_id):
        """Create comment"""
        try:
            task = get_object_or_404(Task, id=task_id)
            data = request.data.copy()
            data['task'] = str(task.id)
            data['admin'] = str(request.user.id)
            
            serializer = TaskCommentSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    "status": status.HTTP_201_CREATED,
                    "message": "Comment added successfully",
                    "data": serializer.data
                }, status=status.HTTP_201_CREATED)
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "Validation error",
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


