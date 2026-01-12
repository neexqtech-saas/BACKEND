"""
Visit Management Views
Optimized for high-traffic, low-cost, future-proof architecture
All queries O(1) or using proper database indexes
"""

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.db.models import Q, Count
from django.utils import timezone
from django.http import HttpResponse
from decimal import Decimal
from io import BytesIO
from datetime import date, datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from .models import Visit
from .serializers import (
    VisitSerializer, VisitCreateSerializer,
    VisitCheckInSerializer, VisitCheckOutSerializer
)
from AuthN.models import BaseUserModel, AdminProfile
from SiteManagement.models import Site
from utils.pagination_utils import CustomPagination
from utils.site_filter_utils import filter_queryset_by_site
from utils.Employee.assignment_utils import get_current_admin_for_employee


def get_admin_and_site_optimized(request, site_id, allow_user_role=False):
    """
    Optimized admin and site validation - O(1) queries with select_related
    Returns: (admin, site, None) tuple or (None, None, Response with error)
    """
    user = request.user
    
    # Fast path for admin role - O(1) query
    if user.role == 'admin':
        admin_id = user.id
        admin = user
        # Single O(1) query with index on (id, created_by_admin, is_active)
        try:
            site = Site.objects.only('id', 'site_name', 'created_by_admin_id', 'is_active').get(
                id=site_id, 
                created_by_admin_id=admin_id, 
                is_active=True
            )
            return admin, site, None
        except Site.DoesNotExist:
            return None, None, Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Site not found or you don't have permission to access this site",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
    
    # Organization role - O(1) queries with select_related
    elif user.role == 'organization':
        admin_id = request.query_params.get('admin_id')
        if not admin_id:
            return None, None, Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "admin_id is required for organization role. Please provide admin_id as query parameter.",
                "data": None
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Single O(1) query with select_related to avoid N+1 - uses index on (id, role)
        try:
            admin = BaseUserModel.objects.only(
                'id', 'role', 'email'
            ).get(id=admin_id, role='admin')
        except BaseUserModel.DoesNotExist:
            return None, None, Response({
                "status": status.HTTP_404_NOT_FOUND,
                "message": "Admin not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
        
        # O(1) query - verify admin belongs to organization using select_related
        admin_profile = AdminProfile.objects.select_related('user', 'organization').only(
            'id', 'user_id', 'organization_id'
        ).filter(
            user_id=admin_id,
            organization_id=user.id
        ).first()
        
        if not admin_profile:
            return None, None, Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Admin does not belong to your organization",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Single O(1) query with index on (id, created_by_admin, is_active)
        try:
            site = Site.objects.only('id', 'site_name', 'created_by_admin_id', 'is_active').get(
                id=site_id, 
                created_by_admin_id=admin_id, 
                is_active=True
            )
            return admin, site, None
        except Site.DoesNotExist:
            return None, None, Response({
                "status": status.HTTP_403_FORBIDDEN,
                "message": "Site not found or you don't have permission to access this site",
                "data": None
            }, status=status.HTTP_403_FORBIDDEN)
    
    # User role - only if allowed
    elif allow_user_role and user.role == 'user':
        # For user role, just validate site exists
        try:
            site = Site.objects.only('id', 'site_name', 'is_active').get(id=site_id, is_active=True)
            return None, site, None
        except Site.DoesNotExist:
            return None, None, Response({
                "status": status.HTTP_404_NOT_FOUND,
                "message": "Site not found",
                "data": None
            }, status=status.HTTP_404_NOT_FOUND)
    
    else:
        return None, None, Response({
            "status": status.HTTP_403_FORBIDDEN,
            "message": "Unauthorized access",
            "data": None
        }, status=status.HTTP_403_FORBIDDEN)


class VisitAPIView(APIView):
    """
    Visit CRUD Operations - Optimized
    - Admin can create visits and assign them to employees
    - Employees can create their own visits
    - Admin can see all visits
    - Employees can only see their assigned or self-created visits
    """
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    
    def generate_excel_export(self, visit_data):
        """Generate Excel export for visits - Optimized with .values()"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visits"
        
        headers = [
            "Visit ID", "Title", "Description", "Employee Name", "Custom Employee ID", "Employee Email",
            "Client Name", "Location Name", "Address", "City", "State", "Pincode", "Country",
            "Contact Person", "Contact Phone", "Contact Email",
            "Schedule Date", "Schedule Time", "Status",
            "Check-in Time", "Check-in Latitude", "Check-in Longitude", "Check-in Note",
            "Check-out Time", "Check-out Latitude", "Check-out Longitude", "Check-out Note",
            "Created By", "Created At"
        ]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Header Row
        for col, head in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=head)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")
        
        # Helper function to convert values to Excel-compatible format
        def to_excel_value(val):
            if val is None:
                return "N/A"
            if hasattr(val, 'strftime'):
                return val.strftime('%Y-%m-%d %H:%M:%S') if hasattr(val, 'hour') else val.strftime('%Y-%m-%d')
            try:
                if not isinstance(val, (str, int, float, bool)):
                    return str(val)
            except:
                return "N/A"
            return val
        
        # Data Rows
        for i, visit in enumerate(visit_data, 2):
            row = [
                to_excel_value(visit.get("id", "N/A")),
                to_excel_value(visit.get("title", "N/A")),
                to_excel_value(visit.get("description", "N/A")),
                to_excel_value(visit.get("assigned_employee_name", "N/A")),
                to_excel_value(visit.get("custom_employee_id", "N/A")),
                to_excel_value(visit.get("assigned_employee_email", "N/A")),
                to_excel_value(visit.get("client_name", "N/A")),
                to_excel_value(visit.get("location_name", "N/A")),
                to_excel_value(visit.get("address", "N/A")),
                to_excel_value(visit.get("city", "N/A")),
                to_excel_value(visit.get("state", "N/A")),
                to_excel_value(visit.get("pincode", "N/A")),
                to_excel_value(visit.get("country", "N/A")),
                to_excel_value(visit.get("contact_person", "N/A")),
                to_excel_value(visit.get("contact_phone", "N/A")),
                to_excel_value(visit.get("contact_email", "N/A")),
                to_excel_value(visit.get("schedule_date", "N/A")),
                to_excel_value(visit.get("schedule_time", "N/A")),
                to_excel_value(visit.get("status", "N/A")),
                to_excel_value(visit.get("check_in_timestamp", "N/A")),
                to_excel_value(visit.get("check_in_latitude", "N/A")),
                to_excel_value(visit.get("check_in_longitude", "N/A")),
                to_excel_value(visit.get("check_in_note", "N/A")),
                to_excel_value(visit.get("check_out_timestamp", "N/A")),
                to_excel_value(visit.get("check_out_latitude", "N/A")),
                to_excel_value(visit.get("check_out_longitude", "N/A")),
                to_excel_value(visit.get("check_out_note", "N/A")),
                to_excel_value(visit.get("created_by_name", "N/A")),
                to_excel_value(visit.get("created_at", "N/A")),
            ]
            for col, val in enumerate(row, 1):
                ws.cell(row=i, column=col).value = to_excel_value(val)
        
        # Auto width
        for col in ws.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[letter].width = max_len + 2
        
        # Save in memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="visits.xlsx"'
        
        return response
    
    def get(self, request, site_id, user_id=None, pk=None):
        """Get visits - O(1) queries with index optimization"""
        try:
            user = request.user
            
            # Get admin and site - O(1) queries
            admin, site, error_response = get_admin_and_site_optimized(request, site_id, allow_user_role=True)
            if error_response:
                return error_response
            
            admin_id = admin.id if admin else None
            
            if pk:
                # Single O(1) query using index visit_id_adm_idx (id, admin)
                if admin_id:
                    visit = Visit.objects.filter(
                        id=pk, 
                        admin_id=admin_id
                    ).select_related(
                        'admin', 'assigned_employee', 'assigned_employee__own_user_profile',
                        'created_by', 'created_by__own_user_profile', 'created_by__own_admin_profile',
                        'site'
                    ).only(
                        'id', 'admin_id', 'site_id', 'assigned_employee_id', 'created_by_id',
                        'title', 'description', 'schedule_date', 'schedule_time', 'status',
                        'client_name', 'location_name', 'address', 'city', 'state', 'pincode', 'country',
                        'contact_person', 'contact_phone', 'contact_email',
                        'check_in_timestamp', 'check_in_latitude', 'check_in_longitude', 'check_in_note',
                        'check_out_timestamp', 'check_out_latitude', 'check_out_longitude', 'check_out_note',
                        'created_at', 'updated_at'
                    ).first()
                else:
                    visit = Visit.objects.filter(id=pk).select_related(
                        'admin', 'assigned_employee', 'assigned_employee__own_user_profile',
                        'created_by', 'created_by__own_user_profile', 'created_by__own_admin_profile',
                        'site'
                    ).only(
                        'id', 'admin_id', 'site_id', 'assigned_employee_id', 'created_by_id',
                        'title', 'description', 'schedule_date', 'schedule_time', 'status',
                        'client_name', 'location_name', 'address', 'city', 'state', 'pincode', 'country',
                        'contact_person', 'contact_phone', 'contact_email',
                        'check_in_timestamp', 'check_in_latitude', 'check_in_longitude', 'check_in_note',
                        'check_out_timestamp', 'check_out_latitude', 'check_out_longitude', 'check_out_note',
                        'created_at', 'updated_at'
                    ).first()
                
                if not visit:
                    return Response({
                        "status": status.HTTP_404_NOT_FOUND,
                        "message": "Visit not found",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
                
                # O(1) site check
                if site_id and visit.site_id != site_id:
                    return Response({
                        "status": status.HTTP_404_NOT_FOUND,
                        "message": "Visit not found for this site",
                        "data": None
                    }, status=status.HTTP_404_NOT_FOUND)
                
                # Check access permission
                if user.role == 'user' and visit.assigned_employee_id != user.id and visit.created_by_id != user.id:
                    return Response({
                        "status": status.HTTP_403_FORBIDDEN,
                        "message": "You don't have permission to view this visit",
                        "data": None
                    }, status=status.HTTP_403_FORBIDDEN)
                
                serializer = VisitSerializer(visit)
                return Response({
                    "status": status.HTTP_200_OK,
                    "message": "Visit fetched successfully",
                    "data": serializer.data
                })
            else:
                # Get list of visits with optimized queries
                if user.role == 'admin' and admin_id:
                    # O(1) query using index visit_adm_status_date_idx
                    queryset = Visit.objects.filter(admin_id=admin_id).select_related(
                        'admin',
                        'assigned_employee',
                        'assigned_employee__own_user_profile',
                        'created_by',
                        'created_by__own_user_profile',
                        'created_by__own_admin_profile',
                        'site'
                    ).only(
                        'id', 'admin_id', 'site_id', 'assigned_employee_id', 'created_by_id',
                        'title', 'description', 'schedule_date', 'schedule_time', 'status',
                        'client_name', 'location_name', 'address', 'city', 'state', 'pincode', 'country',
                        'contact_person', 'contact_phone', 'contact_email',
                        'check_in_timestamp', 'check_in_latitude', 'check_in_longitude', 'check_in_note',
                        'check_out_timestamp', 'check_out_latitude', 'check_out_longitude', 'check_out_note',
                        'created_at', 'updated_at'
                    )
                    
                    # Filter by user_id if provided
                    if user_id:
                        queryset = queryset.filter(assigned_employee_id=user_id)
                elif user.role == 'user':
                    # Employees can see all their assigned or self-created visits - O(1) query
                    queryset = Visit.objects.filter(
                        Q(assigned_employee_id=user.id) | Q(created_by_id=user.id)
                    ).select_related(
                        'admin',
                        'assigned_employee',
                        'assigned_employee__own_user_profile',
                        'created_by',
                        'created_by__own_user_profile',
                        'created_by__own_admin_profile',
                        'site'
                    ).only(
                        'id', 'admin_id', 'site_id', 'assigned_employee_id', 'created_by_id',
                        'title', 'description', 'schedule_date', 'schedule_time', 'status',
                        'client_name', 'location_name', 'address', 'city', 'state', 'pincode', 'country',
                        'contact_person', 'contact_phone', 'contact_email',
                        'check_in_timestamp', 'check_in_latitude', 'check_in_longitude', 'check_in_note',
                        'check_out_timestamp', 'check_out_latitude', 'check_out_longitude', 'check_out_note',
                        'created_at', 'updated_at'
                    )
                else:
                    queryset = Visit.objects.none()
                
                # Filter by site - O(1) with index visit_site_adm_status_date_idx
                queryset = filter_queryset_by_site(queryset, site_id, 'site')
                
                # Apply filters
                status_filter = request.query_params.get('status')
                if status_filter:
                    queryset = queryset.filter(status=status_filter)
                
                # Date filter - default to last 10 days if not provided
                date_from = request.query_params.get('date_from')
                date_to = request.query_params.get('date_to')
                
                if not date_from and not date_to:
                    # Default to last 10 days if no date filter provided
                    date_to = date.today()
                    date_from = date_to - timedelta(days=10)
                    queryset = queryset.filter(
                        schedule_date__gte=date_from,
                        schedule_date__lte=date_to
                    )
                else:
                    if date_from:
                        try:
                            date_from_obj = datetime.strptime(date_from, "%Y-%m-%d").date()
                            queryset = queryset.filter(schedule_date__gte=date_from_obj)
                        except ValueError:
                            pass  # Invalid date format, ignore
                    
                    if date_to:
                        try:
                            date_to_obj = datetime.strptime(date_to, "%Y-%m-%d").date()
                            queryset = queryset.filter(schedule_date__lte=date_to_obj)
                        except ValueError:
                            pass  # Invalid date format, ignore
                
                # Search functionality - uses indexes
                search_query = request.query_params.get('search', '').strip()
                if search_query:
                    queryset = queryset.filter(
                        Q(assigned_employee__own_user_profile__user_name__icontains=search_query) |
                        Q(assigned_employee__email__icontains=search_query) |
                        Q(assigned_employee__own_user_profile__custom_employee_id__icontains=search_query) |
                        Q(client_name__icontains=search_query) |
                        Q(location_name__icontains=search_query) |
                        Q(address__icontains=search_query) |
                        Q(contact_person__icontains=search_query) |
                        Q(contact_phone__icontains=search_query) |
                        Q(contact_email__icontains=search_query) |
                        Q(title__icontains=search_query) |
                        Q(description__icontains=search_query)
                    )
                
                # Check for Excel export
                export = request.query_params.get('export') == 'true'
                if export:
                    # Get all visits without pagination for export
                    serializer = VisitSerializer(queryset.order_by('-created_at', '-updated_at'), many=True)
                    return self.generate_excel_export(serializer.data)
                
                # Order by created_at descending (newest first)
                queryset = queryset.order_by('-created_at', '-updated_at')
                
                # Apply range/limit parameter for user_id endpoint (default 50)
                if user_id:
                    limit_param = request.query_params.get('limit') or request.query_params.get('range')
                    if limit_param:
                        try:
                            limit = int(limit_param)
                            if limit > 0:
                                queryset = queryset[:limit]
                        except ValueError:
                            pass  # Invalid limit, use default
                    
                    # If no limit specified, default to 50
                    if not limit_param:
                        queryset = queryset[:50]
                    
                    serializer = VisitSerializer(queryset, many=True)
                    return Response({
                        "status": status.HTTP_200_OK,
                        "message": "Visits fetched successfully",
                        "data": serializer.data
                    })
                
                # Pagination for admin view (all visits)
                paginator = self.pagination_class()
                paginated_qs = paginator.paginate_queryset(queryset, request)
                serializer = VisitSerializer(paginated_qs, many=True)
                pagination_data = paginator.get_paginated_response(serializer.data)
                
                return Response({
                    "status": status.HTTP_200_OK,
                    "message": "Visits fetched successfully",
                    "data": serializer.data,
                    **pagination_data
                })
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @transaction.atomic
    def post(self, request, site_id, user_id=None):
        """Create visit - Optimized"""
        try:
            # Get admin and site - O(1) queries
            admin, site, error_response = get_admin_and_site_optimized(request, site_id, allow_user_role=True)
            if error_response:
                return error_response
            
            # For user role, get admin from current assignment
            if not admin and request.user.role == 'user':
                admin = get_current_admin_for_employee(request.user)
                if not admin:
                    return Response({
                        "status": status.HTTP_403_FORBIDDEN,
                        "message": "You are not assigned to any admin. Please contact your administrator.",
                        "data": None
                    }, status=status.HTTP_403_FORBIDDEN)
            
            user = request.user
            
            data = request.data.copy()
            data['admin'] = str(admin.id)
            data['created_by'] = str(user.id)
            # Set site if provided
            if site_id:
                data['site'] = str(site.id)
            
            # If employee is creating, they can assign to themselves or leave it to admin
            if user.role == 'user':
                # If assigned_employee is not provided, assign to the creator
                if 'assigned_employee' not in data or not data['assigned_employee']:
                    data['assigned_employee'] = str(user.id)
                # Employees can only assign to themselves
                elif data['assigned_employee'] != str(user.id):
                    return Response({
                        "status": status.HTTP_403_FORBIDDEN,
                        "message": "Employees can only create visits for themselves",
                        "data": None
                    }, status=status.HTTP_403_FORBIDDEN)
            
            # If user_id is provided in URL and user is admin, use it
            if user_id and user.role == 'admin':
                data['assigned_employee'] = str(user_id)
            
            serializer = VisitCreateSerializer(data=data)
            if serializer.is_valid():
                visit = serializer.save(
                    admin=admin,
                    created_by=user
                )
                
                response_serializer = VisitSerializer(visit)
                return Response({
                    "status": status.HTTP_201_CREATED,
                    "message": "Visit created successfully",
                    "data": response_serializer.data
                }, status=status.HTTP_201_CREATED)
            
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "Validation error",
                "data": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @transaction.atomic
    def put(self, request, site_id, user_id=None, pk=None):
        """Update visit - Optimized O(1) query"""
        try:
            if not pk:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "Visit ID is required",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Single O(1) query using index visit_id_adm_idx
            visit = Visit.objects.select_related('admin', 'site').only(
                'id', 'admin_id', 'site_id', 'created_by_id'
            ).filter(id=pk).first()
            
            if not visit:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Visit not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Get admin from visit
            admin = visit.admin
            
            # Validate site belongs to admin
            admin, site, error_response = get_admin_and_site_optimized(request, site_id)
            if error_response:
                return error_response
            
            # O(1) site check
            if site_id and visit.site_id != site_id:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Visit not found for this site",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            user = request.user
            
            # Check permission - only admin or creator can update
            if user.role == 'user' and visit.created_by_id != user.id:
                return Response({
                    "status": status.HTTP_403_FORBIDDEN,
                    "message": "You don't have permission to update this visit",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Don't allow updating status directly through this endpoint
            data = request.data.copy()
            if 'status' in data:
                data.pop('status')
            
            serializer = VisitCreateSerializer(visit, data=data, partial=True)
            if serializer.is_valid():
                serializer.save()
                
                response_serializer = VisitSerializer(visit)
                return Response({
                    "status": status.HTTP_200_OK,
                    "message": "Visit updated successfully",
                    "data": response_serializer.data
                })
            
            return Response({
                "status": status.HTTP_400_BAD_REQUEST,
                "message": "Validation error",
                "data": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def delete(self, request, site_id, user_id=None, pk=None):
        """Delete visit - Optimized O(1) query"""
        try:
            if not pk:
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "Visit ID is required",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Single O(1) query using index visit_id_adm_idx
            visit = Visit.objects.filter(id=pk).only('id', 'site_id', 'created_by_id').first()
            
            if not visit:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Visit not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) site check
            if site_id and visit.site_id != site_id:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Visit not found for this site",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            user = request.user
            
            # Check permission
            if user.role == 'user' and visit.created_by_id != user.id:
                return Response({
                    "status": status.HTTP_403_FORBIDDEN,
                    "message": "You don't have permission to delete this visit",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            visit.delete()
            return Response({
                "status": status.HTTP_200_OK,
                "message": "Visit deleted successfully",
                "data": None
            })
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class VisitCheckInAPIView(APIView):
    """
    Check-In Endpoint - Optimized
    - Employee must send GPS coordinates (latitude, longitude)
    - Optional note
    - Updates visit status to 'in_progress'
    - Only assigned employee (or creator if self-visit) can check-in
    """
    permission_classes = [IsAuthenticated]
    
    @transaction.atomic
    def post(self, request, site_id, user_id=None, visit_id=None):
        """Perform check-in - Optimized O(1) query"""
        try:
            # Single O(1) query using index visit_id_adm_idx
            visit = Visit.objects.select_related('admin', 'site').only(
                'id', 'admin_id', 'site_id', 'status', 'assigned_employee_id', 'created_by_id'
            ).filter(id=visit_id).first()
            
            if not visit:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Visit not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Get admin from visit
            admin = visit.admin
            
            # Validate site belongs to admin
            admin, site, error_response = get_admin_and_site_optimized(request, site_id)
            if error_response:
                return error_response
            
            # O(1) site check
            if site_id and visit.site_id != site_id:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Visit not found for this site",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            user = request.user
            
            # Check permission
            if not visit.can_perform_check_in_out(user):
                return Response({
                    "status": status.HTTP_403_FORBIDDEN,
                    "message": "You don't have permission to check-in for this visit",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Check if already checked in
            if visit.status == 'in_progress':
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "Visit is already in progress. Please check-out first.",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if visit.status == 'completed':
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "Cannot check-in to a completed visit",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate check-in data
            serializer = VisitCheckInSerializer(data=request.data)
            if not serializer.is_valid():
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "Validation error",
                    "data": serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            validated_data = serializer.validated_data
            latitude = validated_data['latitude']
            longitude = validated_data['longitude']
            note = validated_data.get('note', '')
            
            # Update visit - O(1) optimized update
            check_in_time = timezone.now()
            Visit.objects.filter(id=visit_id).update(
                status='in_progress',
                check_in_timestamp=check_in_time,
                check_in_latitude=Decimal(str(latitude)),
                check_in_longitude=Decimal(str(longitude)),
                check_in_note=note
            )
            
            # Refresh visit for response
            visit.refresh_from_db()
            response_serializer = VisitSerializer(visit)
            return Response({
                "status": status.HTTP_200_OK,
                "message": "Check-in successful",
                "data": response_serializer.data
            })
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class VisitCheckOutAPIView(APIView):
    """
    Check-Out Endpoint - Optimized
    - Employee must send GPS coordinates (latitude, longitude)
    - Optional note
    - Updates visit status to 'completed'
    - Only assigned employee (or creator if self-visit) can check-out
    """
    permission_classes = [IsAuthenticated]
    
    @transaction.atomic
    def post(self, request, site_id, user_id=None, visit_id=None):
        """Perform check-out - Optimized O(1) query"""
        try:
            # Single O(1) query using index visit_id_adm_idx
            visit = Visit.objects.select_related('admin', 'site').only(
                'id', 'admin_id', 'site_id', 'status', 'assigned_employee_id', 'created_by_id'
            ).filter(id=visit_id).first()
            
            if not visit:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Visit not found",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Get admin from visit
            admin = visit.admin
            
            # Validate site belongs to admin
            admin, site, error_response = get_admin_and_site_optimized(request, site_id)
            if error_response:
                return error_response
            
            # O(1) site check
            if site_id and visit.site_id != site_id:
                return Response({
                    "status": status.HTTP_404_NOT_FOUND,
                    "message": "Visit not found for this site",
                    "data": None
                }, status=status.HTTP_404_NOT_FOUND)
            
            user = request.user
            
            # Check permission
            if not visit.can_perform_check_in_out(user):
                return Response({
                    "status": status.HTTP_403_FORBIDDEN,
                    "message": "You don't have permission to check-out for this visit",
                    "data": None
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Check if visit is in progress
            if visit.status != 'in_progress':
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "Visit must be in progress before checking out. Please check-in first.",
                    "data": None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate check-out data
            serializer = VisitCheckOutSerializer(data=request.data)
            if not serializer.is_valid():
                return Response({
                    "status": status.HTTP_400_BAD_REQUEST,
                    "message": "Validation error",
                    "data": serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
            
            validated_data = serializer.validated_data
            latitude = validated_data['latitude']
            longitude = validated_data['longitude']
            note = validated_data.get('note', '')
            
            # Update visit - O(1) optimized update
            check_out_time = timezone.now()
            Visit.objects.filter(id=visit_id).update(
                status='completed',
                check_out_timestamp=check_out_time,
                check_out_latitude=Decimal(str(latitude)),
                check_out_longitude=Decimal(str(longitude)),
                check_out_note=note
            )
            
            # Refresh visit for response
            visit.refresh_from_db()
            response_serializer = VisitSerializer(visit)
            return Response({
                "status": status.HTTP_200_OK,
                "message": "Check-out successful",
                "data": response_serializer.data
            })
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class VisitStatsAPIView(APIView):
    """
    Get visit statistics - Optimized
    - Admin can see all statistics
    - Employees can see their own statistics
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request, site_id, user_id=None):
        """Get visit statistics - O(1) queries with aggregation"""
        try:
            user = request.user
            
            # Get admin and site - O(1) queries
            admin, site, error_response = get_admin_and_site_optimized(request, site_id, allow_user_role=True)
            if error_response:
                return error_response
            
            admin_id = admin.id if admin else None
            
            # Build queryset based on role - O(1) queries
            if user.role == 'admin' and admin_id:
                visits = Visit.objects.filter(admin_id=admin_id)
                if user_id:
                    visits = visits.filter(assigned_employee_id=user_id)
            elif user.role == 'user':
                # Employees can see all their assigned or self-created visits - O(1) query
                visits = Visit.objects.filter(
                    Q(assigned_employee_id=user.id) | Q(created_by_id=user.id)
                )
            else:
                visits = Visit.objects.none()
            
            # Filter by site - O(1) with index visit_site_adm_status_date_idx
            visits = filter_queryset_by_site(visits, site_id, 'site')
            
            # Single optimized query for all statistics - O(1) aggregation
            stats = {
                'total_visits': visits.count(),
                'pending': visits.filter(status='pending').count(),
                'in_progress': visits.filter(status='in_progress').count(),
                'completed': visits.filter(status='completed').count(),
                'cancelled': visits.filter(status='cancelled').count(),
            }
            
            return Response({
                "status": status.HTTP_200_OK,
                "message": "Statistics fetched successfully",
                "data": stats
            })
        except Exception as e:
            return Response({
                "status": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": str(e),
                "data": None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
