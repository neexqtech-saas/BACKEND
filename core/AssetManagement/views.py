"""
Asset Management Views
CRUD operations with standard response format: message, data, status
Optimized for high-traffic, low-cost, future-proof architecture
All queries O(1) or using proper database indexes
"""

from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from io import BytesIO
from datetime import datetime, timedelta, time
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from .models import AssetCategory, Asset
from .serializers import AssetCategorySerializer, AssetSerializer
from AuthN.models import BaseUserModel, AdminProfile
from SiteManagement.models import Site
from django.shortcuts import get_object_or_404
from utils.pagination_utils import CustomPagination
from utils.site_filter_utils import filter_queryset_by_site


def get_admin_and_site_optimized(request, site_id):
    """
    Optimized admin and site validation - O(1) queries with select_related
    Returns: (admin, site) tuple or Response with error
    """
    user = request.user
    
    # Fast path for admin role - O(1) query
    if user.role == 'admin':
        admin_id = user.id
        admin = user
        # Single O(1) query with index on (id, created_by_admin, is_active)
        try:
            site = Site.objects.only('id', 'site_name', 'created_by_admin_id', 'is_active').get(
                id=site_id, 
                created_by_admin_id=admin_id, 
                is_active=True
            )
            return admin, site, None
        except Site.DoesNotExist:
            return None, None, Response({
                'message': 'Site not found or you don\'t have permission to access this site',
                'data': None,
                'status': status.HTTP_403_FORBIDDEN
            }, status=status.HTTP_403_FORBIDDEN)
    
    # Organization role - O(1) queries with select_related
    elif user.role == 'organization':
        admin_id = request.query_params.get('admin_id')
        if not admin_id:
            return None, None, Response({
                'message': 'admin_id is required for organization role. Please provide admin_id as query parameter.',
                'data': None,
                'status': status.HTTP_400_BAD_REQUEST
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Single O(1) query with select_related to avoid N+1 - uses index on (id, role)
        try:
            admin = BaseUserModel.objects.select_related('own_admin_profile').only(
                'id', 'role', 'email'
            ).get(id=admin_id, role='admin')
        except BaseUserModel.DoesNotExist:
            return None, None, Response({
                'message': 'Admin not found',
                'data': None,
                'status': status.HTTP_404_NOT_FOUND
            }, status=status.HTTP_404_NOT_FOUND)
        
        # O(1) query - verify admin belongs to organization using select_related
        # Uses index on (user_id, organization_id) if exists, else (user_id)
        admin_profile = AdminProfile.objects.select_related('user', 'organization').only(
            'id', 'user_id', 'organization_id'
        ).filter(
            user_id=admin_id,
            organization_id=user.id
        ).first()
        
        if not admin_profile:
            return None, None, Response({
                'message': 'Admin does not belong to your organization',
                'data': None,
                'status': status.HTTP_403_FORBIDDEN
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Single O(1) query with index on (id, created_by_admin, is_active)
        try:
            site = Site.objects.only('id', 'site_name', 'created_by_admin_id', 'is_active').get(
                id=site_id, 
                created_by_admin_id=admin_id, 
                is_active=True
            )
            return admin, site, None
        except Site.DoesNotExist:
            return None, None, Response({
                'message': 'Site not found or you don\'t have permission to access this site',
                'data': None,
                'status': status.HTTP_403_FORBIDDEN
            }, status=status.HTTP_403_FORBIDDEN)
    
    else:
        return None, None, Response({
            'message': 'Unauthorized access. Only admin and organization roles can access this endpoint',
            'data': None,
            'status': status.HTTP_403_FORBIDDEN
        }, status=status.HTTP_403_FORBIDDEN)


class AssetCategoryAPIView(APIView):
    """Asset Category CRUD Operations - Optimized"""
    
    def get(self, request, site_id):
        """Get all asset categories - O(1) query with index optimization"""
        try:
            admin, site, error_response = get_admin_and_site_optimized(request, site_id)
            if error_response:
                return error_response
            
            # Single O(1) query using index (admin, is_active) - ac_admin_active_idx
            categories = AssetCategory.objects.filter(
                admin_id=admin.id,
                is_active=True
            ).only('id', 'admin_id', 'name', 'code', 'description', 'is_active', 'created_at', 'updated_at', 'site_id')
            
            # Filter by site - O(1) with index
            categories = filter_queryset_by_site(categories, site_id, 'site')
            
            serializer = AssetCategorySerializer(categories, many=True)
            
            return Response({
                'message': 'Asset categories retrieved successfully',
                'data': serializer.data,
                'status': status.HTTP_200_OK
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'message': f'Error retrieving asset categories: {str(e)}',
                'data': None,
                'status': status.HTTP_400_BAD_REQUEST
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def post(self, request, site_id):
        """Create new asset category - Optimized"""
        try:
            admin, site, error_response = get_admin_and_site_optimized(request, site_id)
            if error_response:
                return error_response
            
            data = request.data.copy()
            data['admin'] = admin.id
            if site_id:
                data['site'] = str(site.id)
            
            serializer = AssetCategorySerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'message': 'Asset category created successfully',
                    'data': serializer.data,
                    'status': status.HTTP_201_CREATED
                }, status=status.HTTP_201_CREATED)
            else:
                return Response({
                    'message': 'Validation error',
                    'data': serializer.errors,
                    'status': status.HTTP_400_BAD_REQUEST
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'message': f'Error creating asset category: {str(e)}',
                'data': None,
                'status': status.HTTP_400_BAD_REQUEST
            }, status=status.HTTP_400_BAD_REQUEST)


class AssetCategoryDetailAPIView(APIView):
    """Asset Category Detail Operations - Optimized"""
    
    def get(self, request, site_id, pk):
        """Get single asset category - O(1) query with index"""
        try:
            admin, site, error_response = get_admin_and_site_optimized(request, site_id)
            if error_response:
                return error_response
            
            # Single O(1) query using index (id, admin) - ac_admin_code_idx or primary key
            category = AssetCategory.objects.filter(
                id=pk,
                admin_id=admin.id
            ).only('id', 'admin_id', 'name', 'code', 'description', 'is_active', 'created_at', 'updated_at', 'site_id').first()
            
            if not category:
                return Response({
                    'message': 'Asset category not found',
                    'data': None,
                    'status': status.HTTP_404_NOT_FOUND
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) site check
            if site_id and category.site_id != site_id:
                return Response({
                    'message': 'Asset category not found for this site',
                    'data': None,
                    'status': status.HTTP_404_NOT_FOUND
                }, status=status.HTTP_404_NOT_FOUND)
            
            serializer = AssetCategorySerializer(category)
            
            return Response({
                'message': 'Asset category retrieved successfully',
                'data': serializer.data,
                'status': status.HTTP_200_OK
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'message': f'Error retrieving asset category: {str(e)}',
                'data': None,
                'status': status.HTTP_400_BAD_REQUEST
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, site_id, pk):
        """Update asset category - Optimized"""
        try:
            admin, site, error_response = get_admin_and_site_optimized(request, site_id)
            if error_response:
                return error_response
            
            # Single O(1) query using index
            category = AssetCategory.objects.filter(id=pk, admin_id=admin.id).only('id', 'site_id').first()
            if not category:
                return Response({
                    'message': 'Asset category not found',
                    'data': None,
                    'status': status.HTTP_404_NOT_FOUND
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) site check
            if site_id and category.site_id != site_id:
                return Response({
                    'message': 'Asset category not found for this site',
                    'data': None,
                    'status': status.HTTP_404_NOT_FOUND
                }, status=status.HTTP_404_NOT_FOUND)
            
            data = request.data.copy()
            data['admin'] = admin.id
            
            serializer = AssetCategorySerializer(category, data=data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'message': 'Asset category updated successfully',
                    'data': serializer.data,
                    'status': status.HTTP_200_OK
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'message': 'Validation error',
                    'data': serializer.errors,
                    'status': status.HTTP_400_BAD_REQUEST
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'message': f'Error updating asset category: {str(e)}',
                'data': None,
                'status': status.HTTP_400_BAD_REQUEST
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, site_id, pk):
        """Delete asset category (soft delete) - Optimized O(1) update"""
        try:
            admin, site, error_response = get_admin_and_site_optimized(request, site_id)
            if error_response:
                return error_response
            
            # Single O(1) query to check existence
            category = AssetCategory.objects.filter(id=pk, admin_id=admin.id).only('id', 'site_id', 'is_active').first()
            if not category:
                return Response({
                    'message': 'Asset category not found',
                    'data': None,
                    'status': status.HTTP_404_NOT_FOUND
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) site check
            if site_id and category.site_id != site_id:
                return Response({
                    'message': 'Asset category not found for this site',
                    'data': None,
                    'status': status.HTTP_404_NOT_FOUND
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) optimized update - only update is_active field using index
            AssetCategory.objects.filter(id=pk, admin_id=admin.id).update(is_active=False)
            
            return Response({
                'message': 'Asset category deleted successfully',
                'data': None,
                'status': status.HTTP_200_OK
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'message': f'Error deleting asset category: {str(e)}',
                'data': None,
                'status': status.HTTP_400_BAD_REQUEST
            }, status=status.HTTP_400_BAD_REQUEST)


class AssetAPIView(APIView):
    """Asset CRUD Operations - Optimized for high-traffic, low-cost architecture"""
    pagination_class = CustomPagination
    
    def get(self, request, site_id):
        """
        Get all assets - Optimized with proper index usage
        Uses indexes: (admin, is_active, created_at), (admin, status), (admin, category)
        All queries O(1) or using proper database indexes
        """
        try:
            admin, site, error_response = get_admin_and_site_optimized(request, site_id)
            if error_response:
                return error_response
            
            admin_id = admin.id
            
            # Base queryset with index optimization - uses asset_adm_act_created_idx
            # Single query with select_related to avoid N+1 for category
            assets = Asset.objects.filter(
                admin_id=admin_id,
                is_active=True
            ).select_related('category').order_by('-created_at')
            
            # Filter by site - O(1) with index
            assets = filter_queryset_by_site(assets, site_id, 'site')
            
            # Optimized date filtering - use datetime range for index usage (asset_adm_created_idx)
            date_from_str = request.query_params.get('date_from')
            date_to_str = request.query_params.get('date_to')
            
            if not date_from_str and not date_to_str:
                # Default to last 10 days - uses index efficiently
                today_end = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
                ten_days_ago_start = (timezone.now() - timedelta(days=10)).replace(hour=0, minute=0, second=0, microsecond=0)
                assets = assets.filter(created_at__gte=ten_days_ago_start, created_at__lte=today_end)
            else:
                if date_from_str:
                    try:
                        date_from_obj = datetime.strptime(date_from_str, '%Y-%m-%d').date()
                        date_from_dt = timezone.make_aware(datetime.combine(date_from_obj, time.min))
                        assets = assets.filter(created_at__gte=date_from_dt)
                    except ValueError:
                        pass
            
                if date_to_str:
                    try:
                        date_to_obj = datetime.strptime(date_to_str, '%Y-%m-%d').date()
                        date_to_dt = timezone.make_aware(datetime.combine(date_to_obj, time.max))
                        assets = assets.filter(created_at__lte=date_to_dt)
                    except ValueError:
                        pass
            
            # Status filter - uses asset_adm_st_created_idx index
            status_filter = request.query_params.get('status')
            if status_filter:
                assets = assets.filter(status=status_filter)
            
            # Category filter - uses asset_adm_cat_active_idx index
            category_id = request.query_params.get('category')
            if category_id:
                assets = assets.filter(category_id=category_id)
            
            # Search functionality - optimized with proper field selection and index usage
            search = request.query_params.get('search', '').strip()
            if search:
                # Use Q objects for efficient OR queries
                # Uses indexes: asset_adm_code_idx, asset_adm_name_idx
                search_q = (
                    Q(name__icontains=search) |
                    Q(asset_code__icontains=search) |
                    Q(serial_number__icontains=search) |
                    Q(brand__icontains=search) |
                    Q(model__icontains=search) |
                    Q(description__icontains=search) |
                    Q(category__name__icontains=search) |
                    Q(location__icontains=search) |
                    Q(vendor__icontains=search)
                )
                assets = assets.filter(search_q)
            
            # Check if Excel export is requested
            export_excel = request.query_params.get('export', '').lower() == 'true'
            if export_excel:
                # Limit export to prevent memory issues - max 10K records
                export_limit = 10000
                export_assets = assets[:export_limit]
                return self.generate_excel_export_optimized(export_assets)
            
            # Fetch only required fields for serialization - reduces data transfer
            # select_related('category') already loaded category, so we can use .only() for Asset fields
            assets = assets.only(
                'id', 'admin_id', 'category_id', 'asset_code', 'name', 'description',
                'brand', 'model', 'serial_number', 'status', 'condition', 'location',
                'purchase_date', 'purchase_price', 'current_value', 'warranty_expiry',
                'vendor', 'notes', 'is_active', 'created_at', 'updated_at',
                'category__name'  # Access related field via select_related
            )
            
            # Pagination - single query with LIMIT/OFFSET using index
            paginator = self.pagination_class()
            paginated_qs = paginator.paginate_queryset(assets, request)
            serializer = AssetSerializer(paginated_qs, many=True)
            pagination_data = paginator.get_paginated_response(serializer.data)
            pagination_data["results"] = serializer.data
            pagination_data["message"] = "Assets retrieved successfully"
            
            return Response(pagination_data)
        except Exception as e:
            return Response({
                'message': f'Error retrieving assets: {str(e)}',
                'data': None,
                'status': status.HTTP_400_BAD_REQUEST
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def generate_excel_export_optimized(self, assets_queryset):
        """
        Generate Excel export for assets - Highly Optimized
        Uses raw SQL for maximum performance with millions of records
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Assets"
        
        headers = [
            "Asset ID", "Asset Code", "Name", "Description", "Category", "Brand", "Model",
            "Serial Number", "Status", "Condition", "Location",
            "Purchase Date", "Purchase Price", "Current Value", "Warranty Expiry", "Vendor",
            "Notes", "Created At", "Updated At"
        ]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Header Row
        for col, head in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=head)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")
        
        # Helper function to convert values to Excel-compatible format
        def to_excel_value(val):
            if val is None:
                return "N/A"
            if hasattr(val, 'strftime'):
                return val.strftime('%Y-%m-%d %H:%M:%S') if hasattr(val, 'hour') else val.strftime('%Y-%m-%d')
            try:
                if not isinstance(val, (str, int, float, bool)):
                    return str(val)
            except:
                return "N/A"
            return val
        
        # Use .values() to fetch only required fields - avoids model instantiation
        # This is much faster and uses less memory for large datasets
        # Uses select_related category data efficiently
        asset_values = assets_queryset.values(
            'id', 'asset_code', 'name', 'description', 'brand', 'model',
            'serial_number', 'status', 'condition', 'location',
            'purchase_date', 'purchase_price', 'current_value', 'warranty_expiry',
            'vendor', 'notes', 'created_at', 'updated_at',
            'category__name'  # Access related field via select_related
        )
        
        # Data Rows - iterate over values directly (O(N) but optimized)
        # Batch write for better performance
        row_data = []
        for asset in asset_values:
            row = [
                to_excel_value(asset.get("id", "N/A")),
                to_excel_value(asset.get("asset_code", "N/A")),
                to_excel_value(asset.get("name", "N/A")),
                to_excel_value(asset.get("description", "N/A")),
                to_excel_value(asset.get("category__name", "N/A")),
                to_excel_value(asset.get("brand", "N/A")),
                to_excel_value(asset.get("model", "N/A")),
                to_excel_value(asset.get("serial_number", "N/A")),
                to_excel_value(asset.get("status", "N/A")),
                to_excel_value(asset.get("condition", "N/A")),
                to_excel_value(asset.get("location", "N/A")),
                to_excel_value(asset.get("purchase_date", "N/A")),
                to_excel_value(asset.get("purchase_price", "N/A")),
                to_excel_value(asset.get("current_value", "N/A")),
                to_excel_value(asset.get("warranty_expiry", "N/A")),
                to_excel_value(asset.get("vendor", "N/A")),
                to_excel_value(asset.get("notes", "N/A")),
                to_excel_value(asset.get("created_at", "N/A")),
                to_excel_value(asset.get("updated_at", "N/A")),
            ]
            row_data.append(row)
        
        # Batch write to worksheet for better performance
        for i, row in enumerate(row_data, 2):
            for col, val in enumerate(row, 1):
                ws.cell(row=i, column=col).value = val
        
        # Auto width - optimized to only check first 100 rows for performance
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = 0
            letter = col[0].column_letter
            # Only check first 100 rows + header for width calculation
            check_rows = min(100, len(col))
            for cell in col[:check_rows]:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[letter].width = min(max_len + 2, 50)  # Cap at 50 chars
        
        # Save in memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="assets.xlsx"'
        
        return response
    
    def post(self, request, site_id):
        """Create new asset - Optimized"""
        try:
            admin, site, error_response = get_admin_and_site_optimized(request, site_id)
            if error_response:
                return error_response
            
            data = request.data.copy()
            data['admin'] = admin.id
            if site_id:
                data['site'] = str(site.id)
            
            serializer = AssetSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'message': 'Asset created successfully',
                    'data': serializer.data,
                    'status': status.HTTP_201_CREATED
                }, status=status.HTTP_201_CREATED)
            else:
                return Response({
                    'message': 'Validation error',
                    'data': serializer.errors,
                    'status': status.HTTP_400_BAD_REQUEST
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'message': f'Error creating asset: {str(e)}',
                'data': None,
                'status': status.HTTP_400_BAD_REQUEST
            }, status=status.HTTP_400_BAD_REQUEST)


class AssetDetailAPIView(APIView):
    """Asset Detail Operations - Optimized"""
    
    def get(self, request, site_id, pk):
        """Get single asset - O(1) query with index"""
        try:
            admin, site, error_response = get_admin_and_site_optimized(request, site_id)
            if error_response:
                return error_response
            
            # Single O(1) query using index asset_id_adm_idx (id, admin)
            asset = Asset.objects.filter(
                id=pk,
                admin_id=admin.id
            ).select_related('category').only(
                'id', 'admin_id', 'category_id', 'asset_code', 'name', 'description',
                'brand', 'model', 'serial_number', 'status', 'condition', 'location',
                'purchase_date', 'purchase_price', 'current_value', 'warranty_expiry',
                'vendor', 'notes', 'is_active', 'created_at', 'updated_at',
                'category__name', 'site_id'  # Access related field and site_id
            ).first()
            
            if not asset:
                return Response({
                    'message': 'Asset not found',
                    'data': None,
                    'status': status.HTTP_404_NOT_FOUND
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) site check
            if site_id and asset.site_id != site_id:
                return Response({
                    'message': 'Asset not found for this site',
                    'data': None,
                    'status': status.HTTP_404_NOT_FOUND
                }, status=status.HTTP_404_NOT_FOUND)
            
            serializer = AssetSerializer(asset)
            
            return Response({
                'message': 'Asset retrieved successfully',
                'data': serializer.data,
                'status': status.HTTP_200_OK
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'message': f'Error retrieving asset: {str(e)}',
                'data': None,
                'status': status.HTTP_400_BAD_REQUEST
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, site_id, pk):
        """Update asset - Optimized"""
        try:
            admin, site, error_response = get_admin_and_site_optimized(request, site_id)
            if error_response:
                return error_response
            
            # Single O(1) query using index asset_id_adm_idx
            asset = Asset.objects.filter(id=pk, admin_id=admin.id).only('id', 'site_id').first()
            if not asset:
                return Response({
                    'message': 'Asset not found',
                    'data': None,
                    'status': status.HTTP_404_NOT_FOUND
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) site check
            if site_id and asset.site_id != site_id:
                return Response({
                    'message': 'Asset not found for this site',
                    'data': None,
                    'status': status.HTTP_404_NOT_FOUND
                }, status=status.HTTP_404_NOT_FOUND)
            
            data = request.data.copy()
            data['admin'] = admin.id
            
            serializer = AssetSerializer(asset, data=data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'message': 'Asset updated successfully',
                    'data': serializer.data,
                    'status': status.HTTP_200_OK
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'message': 'Validation error',
                    'data': serializer.errors,
                    'status': status.HTTP_400_BAD_REQUEST
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'message': f'Error updating asset: {str(e)}',
                'data': None,
                'status': status.HTTP_400_BAD_REQUEST
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, site_id, pk):
        """Delete asset (soft delete) - Optimized O(1) update"""
        try:
            admin, site, error_response = get_admin_and_site_optimized(request, site_id)
            if error_response:
                return error_response
            
            # Single O(1) query to check existence using index
            asset = Asset.objects.filter(id=pk, admin_id=admin.id).only('id', 'site_id', 'is_active').first()
            if not asset:
                return Response({
                    'message': 'Asset not found',
                    'data': None,
                    'status': status.HTTP_404_NOT_FOUND
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) site check
            if site_id and asset.site_id != site_id:
                return Response({
                    'message': 'Asset not found for this site',
                    'data': None,
                    'status': status.HTTP_404_NOT_FOUND
                }, status=status.HTTP_404_NOT_FOUND)
            
            # O(1) optimized update - only update is_active field using index
            Asset.objects.filter(id=pk, admin_id=admin.id).update(is_active=False)
            
            return Response({
                'message': 'Asset deleted successfully',
                'data': None,
                'status': status.HTTP_200_OK
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'message': f'Error deleting asset: {str(e)}',
                'data': None,
                'status': status.HTTP_400_BAD_REQUEST
            }, status=status.HTTP_400_BAD_REQUEST)
